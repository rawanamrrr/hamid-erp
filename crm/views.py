from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from accounts.permissions import require_permission, require_granular_action, require_granular_action_open  # RBAC (Phase 3.1)
from django.db.models import Q, Sum, Value, DecimalField, Count, F
from django.db.models.functions import Coalesce
import json
import csv
from decimal import Decimal
from django.utils import timezone
from django.views import View
from django.views.generic import ListView
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.utils.http import urlencode
import io
from openpyxl import Workbook, load_workbook

from .models import Customer, CustomerPayment
from .forms import CustomerForm
from sales.models import Order
from settings.models import SystemSetting

@method_decorator(login_required, name='dispatch')
@method_decorator(require_granular_action_open('crm', 'list'), name='dispatch')
class CustomerListView(ListView):
    model = Customer
    template_name = 'crm/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20

    def get_filtered_queryset(self):
        qs = Customer.objects.all().order_by('-created_at')
        q = self.request.GET.get('q', '').strip()
        customer_type = self.request.GET.get('customer_type', '').strip()

        if q:
            qs = qs.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(phone__icontains=q)
            )
        if customer_type:
            qs = qs.filter(customer_type=customer_type)
        return qs

    def get_queryset(self):
        return self.get_filtered_queryset()

    def _export_excel(self, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(['الاسم الأول', 'الاسم الثاني', 'الهاتف', 'العنوان', 'نوع العميل', 'تاريخ الإضافة', 'الرصيد'])
        for cust in queryset:
            ws.append([
                cust.first_name,
                cust.last_name,
                cust.phone,
                cust.address or '',
                cust.get_customer_type_display(),
                cust.created_at.strftime('%Y-%m-%d'),
                float(cust.calculated_balance or 0),
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customers_list.xlsx"'
        return response

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get('export') == 'excel':
            export_qs = self.get_filtered_queryset()
            # Attach the SAME balance every other page (customer detail, POS badge) uses
            # instead of a hand-rolled Sum() — a plain Sum('order__total_amount') over the
            # `order` join both (a) counts void/item-less orders that get_balance()
            # deliberately excludes, and (b) double-counts any order with more than one
            # line item, since the order-items join fans out the row before the Sum.
            for cust in export_qs:
                cust.calculated_balance = cust.get_balance()
            return self._export_excel(export_qs)
        return super().render_to_response(context, **response_kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filtered_qs = self.get_filtered_queryset()

        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        stats = filtered_qs.aggregate(
            total_customers=Count('id'),
            new_this_month=Count('id', filter=Q(created_at__gte=month_start)),
            wholesale_count=Count('id', filter=Q(customer_type__in=['wholesale', 'semi_wholesale'])),
        )
        # total_balance must match the per-row balance shown on this same page (and the
        # customer detail page) exactly — see the comment above on why a Sum() annotation
        # can't safely replicate get_balance() here.
        stats['total_balance'] = sum((c.get_balance() for c in filtered_qs), Decimal('0.00'))

        query_dict = self.request.GET.copy()
        query_dict.pop('page', None)
        query_dict.pop('export', None)
        context.update({
            'title': 'قائمة العملاء',
            'search_query': self.request.GET.get('q', ''),
            'filter_type': self.request.GET.get('customer_type', ''),
            'customer_types': Customer.CUSTOMER_TYPES,
            'stats': stats,
            'export_query': urlencode(query_dict, doseq=True),
        })
        return context


@method_decorator(login_required, name='dispatch')
class CustomerBulkAddView(View):
    def post(self, request):
        try:
            payload = json.loads(request.body or '{}')
            rows = payload.get('rows', [])
            if not isinstance(rows, list) or not rows:
                return JsonResponse({'status': 'error', 'message': 'لا توجد بيانات لإضافتها'}, status=400)

            errors = []
            to_create = []
            seen_phones = set()
            existing_phones = set(Customer.objects.filter(phone__in=[(r.get('phone') or '').strip() for r in rows]).values_list('phone', flat=True))

            for idx, row in enumerate(rows, start=1):
                first_name = (row.get('first_name') or '').strip()
                last_name = (row.get('last_name') or '').strip()
                phone = (row.get('phone') or '').strip()
                address = (row.get('address') or '').strip()
                customer_type = (row.get('customer_type') or 'retail').strip()

                if not first_name:
                    errors.append({'row': idx, 'reason': 'الاسم الأول مطلوب'})
                    continue
                if not phone:
                    errors.append({'row': idx, 'reason': 'رقم الهاتف مطلوب'})
                    continue
                if customer_type not in dict(Customer.CUSTOMER_TYPES):
                    errors.append({'row': idx, 'reason': 'نوع العميل غير صالح'})
                    continue
                if phone in seen_phones:
                    errors.append({'row': idx, 'reason': 'رقم الهاتف مكرر داخل الجدول'})
                    continue
                if phone in existing_phones:
                    errors.append({'row': idx, 'reason': 'رقم الهاتف موجود مسبقا'})
                    continue

                seen_phones.add(phone)
                to_create.append(Customer(
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    address=address or None,
                    customer_type=customer_type,
                ))

            if to_create:
                Customer.objects.bulk_create(to_create)

            return JsonResponse({
                'status': 'success',
                'saved_count': len(to_create),
                'errors': errors,
            })
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'صيغة JSON غير صالحة'}, status=400)


def _normalize_header(value):
    return (value or '').strip().lower().replace(' ', '').replace('_', '')


def _map_customer_row(raw_row):
    row = {k: (v if v is not None else '') for k, v in raw_row.items()}
    header_map = {_normalize_header(k): v for k, v in row.items()}
    first_name = (header_map.get('firstname') or header_map.get('الاسمالأول') or header_map.get('الاسم') or '').strip()
    last_name = (header_map.get('lastname') or header_map.get('الاسمالثاني') or header_map.get('العائلة') or '').strip()
    phone = str(header_map.get('phone') or header_map.get('رقمالهاتف') or '').strip()
    address = (header_map.get('address') or header_map.get('العنوان') or '').strip()
    customer_type = (header_map.get('customertype') or header_map.get('نوعالعميل') or 'retail').strip()
    type_aliases = {
        'individual': 'retail',
        'retail': 'retail',
        'قطاعي': 'retail',
        'semi_wholesale': 'semi_wholesale',
        'semiwholesale': 'semi_wholesale',
        'نصجملة': 'semi_wholesale',
        'wholesale': 'wholesale',
        'جملة': 'wholesale',
    }
    customer_type = type_aliases.get(customer_type.lower().replace(' ', ''), customer_type)
    return {
        'first_name': first_name,
        'last_name': last_name,
        'phone': phone,
        'address': address,
        'customer_type': customer_type,
    }


@method_decorator(login_required, name='dispatch')
class CustomerImportView(View):
    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return JsonResponse({'status': 'error', 'message': 'يرجى اختيار ملف'}, status=400)

        name = upload.name.lower()
        parsed_rows = []
        try:
            if name.endswith('.csv'):
                content = upload.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                parsed_rows = [_map_customer_row(r) for r in reader]
            elif name.endswith('.json'):
                data = json.loads(upload.read().decode('utf-8-sig'))
                if not isinstance(data, list):
                    return JsonResponse({'status': 'error', 'message': 'ملف JSON يجب أن يكون قائمة'}, status=400)
                parsed_rows = [_map_customer_row(r) for r in data if isinstance(r, dict)]
            elif name.endswith('.xlsx'):
                wb = load_workbook(upload, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return JsonResponse({'status': 'error', 'message': 'الملف فارغ'}, status=400)
                headers = [str(h or '') for h in rows[0]]
                for values in rows[1:]:
                    raw = {headers[i]: values[i] if i < len(values) else '' for i in range(len(headers))}
                    parsed_rows.append(_map_customer_row(raw))
            else:
                return JsonResponse({'status': 'error', 'message': 'صيغة الملف غير مدعومة'}, status=400)
        except Exception as exc:
            return JsonResponse({'status': 'error', 'message': f'فشل قراءة الملف: {str(exc)}'}, status=400)

        errors = []
        to_create = []
        seen_phones = set()
        existing_phones = set(Customer.objects.filter(phone__in=[r['phone'] for r in parsed_rows if r.get('phone')]).values_list('phone', flat=True))

        for idx, row in enumerate(parsed_rows, start=2):
            if not row['first_name']:
                errors.append({'row': idx, 'reason': 'الاسم الأول مطلوب'})
                continue
            if not row['phone']:
                errors.append({'row': idx, 'reason': 'رقم الهاتف مطلوب'})
                continue
            if row['customer_type'] not in dict(Customer.CUSTOMER_TYPES):
                errors.append({'row': idx, 'reason': 'نوع العميل غير صالح'})
                continue
            if row['phone'] in seen_phones:
                errors.append({'row': idx, 'reason': 'رقم الهاتف مكرر داخل الملف'})
                continue
            if row['phone'] in existing_phones:
                errors.append({'row': idx, 'reason': 'رقم الهاتف موجود مسبقا'})
                continue

            seen_phones.add(row['phone'])
            to_create.append(Customer(
                first_name=row['first_name'],
                last_name=row['last_name'],
                phone=row['phone'],
                address=row['address'] or None,
                customer_type=row['customer_type'],
            ))

        if to_create:
            Customer.objects.bulk_create(to_create)

        return JsonResponse({
            'status': 'success',
            'success_count': len(to_create),
            'errors': errors,
            'total_rows': len(parsed_rows),
        })


@method_decorator(login_required, name='dispatch')
class CustomerTemplateDownloadView(View):
    def get(self, request):
        fmt = request.GET.get('format', 'xlsx').lower()
        sample_rows = [
            {'first_name': 'أحمد', 'last_name': 'محمد', 'phone': '01000000001', 'address': 'القاهرة', 'customer_type': 'retail'},
            {'first_name': 'شركة النور', 'last_name': '', 'phone': '01000000002', 'address': 'الجيزة', 'customer_type': 'wholesale'},
        ]

        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="customers_template.csv"'
            writer = csv.DictWriter(response, fieldnames=list(sample_rows[0].keys()))
            writer.writeheader()
            writer.writerows(sample_rows)
            return response

        if fmt == 'json':
            response = HttpResponse(
                json.dumps(sample_rows, ensure_ascii=False, indent=2),
                content_type='application/json; charset=utf-8'
            )
            response['Content-Disposition'] = 'attachment; filename="customers_template.json"'
            return response

        wb = Workbook()
        ws = wb.active
        ws.title = "Customers Template"
        headers = list(sample_rows[0].keys())
        ws.append(headers)
        for row in sample_rows:
            ws.append([row[h] for h in headers])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customers_template.xlsx"'
        return response

@login_required
@require_permission('crm', 'view')
def customer_list_print(request):
    """عرض قائمة العملاء للطباعة"""
    customers = Customer.objects.all().order_by('first_name')
    
    # تطبيق نفس الفلاتر عند الطباعة
    q = request.GET.get('q')
    customer_type = request.GET.get('customer_type')
    
    if q:
        customers = customers.filter(
            Q(first_name__icontains=q) | 
            Q(last_name__icontains=q) | 
            Q(phone__icontains=q)
        )
    if customer_type:
        customers = customers.filter(customer_type=customer_type)
        
    sys_settings = SystemSetting.objects.first()
    if not sys_settings:
        sys_settings = SystemSetting.objects.create(shop_name="متجري")

    context = {
        'customers': customers,
        'sys_settings': sys_settings,
        'print_type': request.GET.get('type', 'a4'),
        'print_date': timezone.now(),
        'filter_desc': f"بحث: {q}" if q else "الكل"
    }
    return render(request, 'crm/customer_list_print.html', context)

# ... (باقي الدوال كما هي دون تغيير)
@login_required
@require_permission('crm', 'view')
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    # Combined chronological ledger (exclude voided orders so the running balance
    # reconciles to get_balance()).
    real_orders = customer.order_set.filter(items__isnull=False).exclude(status='void').distinct().order_by('-created_at')
    payments = customer.payments_log.all().order_by('-created_at')
    
    pay_methods_map = {
        'cash': 'نقدي',
        'visa': 'فيزا',
        'bank': 'تحويل بنكي',
        'wallet': 'فودافون كاش',
        'instapay': 'إنستا باي',
        'custom': 'مختلط / مجزأ',
    }
    
    ledger = []
    for order in real_orders:
        method_str = pay_methods_map.get(order.payment_method.lower(), order.payment_method) if order.payment_method else '-'
        ledger.append({
            'type': 'order',
            'id': order.id,
            'date': order.created_at,
            'total': order.total_amount,
            'paid': order.received_amount,
            'method': method_str,
            'notes': order.notes,
            'url_name': 'order_invoice',
        })
    for p in payments:
        method_str = pay_methods_map.get(p.payment_method.lower(), p.payment_method) if p.payment_method else '-'
        ledger.append({
            'type': 'payment' if p.transaction_type == 'payment' else 'debt',
            'id': p.id,
            'date': p.created_at,
            'total': p.amount if p.transaction_type == 'debt' else Decimal('0.00'),
            'paid': p.amount if p.transaction_type == 'payment' else Decimal('0.00'),
            'method': method_str,
            'notes': p.notes,
            'url_name': 'payment_receipt',
        })
    # Running balance per row (positive = customer owes; negative = credit), oldest→newest,
    # starting from the opening balance. Each row moves the balance by (total − paid).
    opening = Decimal(str(customer.opening_balance or '0.00'))
    ledger.sort(key=lambda x: x['date'])
    _run = opening
    for e in ledger:
        _run += (e['total'] - e['paid'])
        e['running'] = _run
    ledger.sort(key=lambda x: x['date'], reverse=True)

    # Saved/recurring orders (Phase 6.11) with a quick shortage count (item stock < saved qty).
    saved_orders = []
    for so in customer.saved_orders.prefetch_related('items__product').all():
        so_items = list(so.items.all())
        short = sum(1 for it in so_items
                    if it.product and (it.product.stock_quantity or 0) < it.quantity)
        saved_orders.append({
            'id': so.id, 'name': so.name, 'count': len(so_items),
            'short': short, 'updated': so.updated_at,
        })

    context = {
        'customer': customer,
        'title': f'ملف العميل: {customer.first_name}',
        'balance': customer.get_balance(),
        'opening_balance': opening,
        'ledger': ledger,
        'saved_orders': saved_orders,
        'total_spent': customer.get_total_spent(),
        'most_purchased': customer.most_purchased_product(),
        'last_visit': customer.last_visit_date(),
        'avg_days': customer.average_purchase_frequency(),
        'highest_invoice': customer.highest_invoice_amount(),
    }
    return render(request, 'crm/customer_detail.html', context)

@login_required
@require_granular_action('crm', 'ar_aging', 'crm', 'view')
def ar_aging_report(request):
    """Accounts-receivable aging — who owes, bucketed by overdue age (Phase 8.1)."""
    from datetime import datetime
    from .aging import ar_aging

    as_of = None
    raw = request.GET.get('as_of')
    if raw:
        try:
            as_of = datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            as_of = None

    rows, totals = ar_aging(as_of)
    if request.GET.get('export') == 'csv':
        from financial.reports import csv_response
        header = ['العميل', 'الهاتف', 'حالي', '31-60', '61-90', '90+', 'الإجمالي']
        data = [[f"{r['customer'].first_name} {r['customer'].last_name}", r['customer'].phone or '',
                 r['current'], r['d30'], r['d60'], r['d90'], r['total']] for r in rows]
        return csv_response('ar_aging', header, data)
    return render(request, 'crm/ar_aging.html', {
        'title': 'أعمار الديون (المستحقات)',
        'rows': rows,
        'totals': totals,
        'as_of': raw or '',
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_permission('crm', 'view')
def customer_statement(request, pk):
    """كشف حساب العميل — chronological subledger with running balance (Phase 4.3)."""
    from datetime import datetime
    from .statements import build_customer_statement

    customer = get_object_or_404(Customer, pk=pk)

    def _parse(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None

    date_from = _parse(request.GET.get('date_from'))
    date_to = _parse(request.GET.get('date_to'))
    statement = build_customer_statement(customer, date_from, date_to)

    sys_settings = SystemSetting.objects.first()
    return render(request, 'crm/customer_statement.html', {
        'title': f"كشف حساب: {customer.first_name} {customer.last_name}",
        'customer': customer,
        'statement': statement,
        'sys_settings': sys_settings,
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_permission('crm', 'view')
def customer_report(request, pk):
    """عرض تقرير شامل عن العميل للطباعة"""
    customer = get_object_or_404(Customer, pk=pk)
    report_type = request.GET.get('type', 'a4') # 'thermal' or 'a4'
    
    # تجميع البيانات
    total_spent = customer.get_total_spent()
    balance = customer.get_balance()
    orders = customer.order_set.filter(items__isnull=False).distinct().order_by('-created_at')
    
    real_order_payments = customer.order_set.filter(items__isnull=False).distinct().aggregate(total=Sum('received_amount'))['total'] or Decimal('0.00')
    customer_payments_sum = customer.payments_log.filter(transaction_type='payment').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_payments = real_order_payments + customer_payments_sum
    
    total_manual_debt = customer.payments_log.filter(transaction_type='debt').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    if report_type == 'thermal':
        orders = orders[:5]
        
    most_purchased = customer.most_purchased_product()
    last_visit = customer.last_visit_date()
    avg_days = customer.average_purchase_frequency()
    
    suggestions = []
    if balance > 0:
        suggestions.append(("warning", "مديونية مستحقة", f"العميل عليه {balance} ج.م. يفضل التذكير بالسداد بلطف."))
    elif balance < 0:
        suggestions.append(("success", "رصيد دائن", f"العميل له رصيد {abs(balance)} ج.م. يمكن استخدامه في الشراء."))
        
    if last_visit:
        days_since_visit = (timezone.now() - last_visit).days
        if days_since_visit > 60:
            suggestions.append(("danger", "عميل خامل", f"لم يزورنا منذ {days_since_visit} يوم. اتصل به لعرض جديد."))
        elif days_since_visit > 30:
            suggestions.append(("info", "تنشيط", "مر شهر على آخر زيارة. أرسل له رسالة عروض."))
            
    if most_purchased != "لا يوجد":
        suggestions.append(("primary", "المنتجات المفضلة", f"يفضل شراء '{most_purchased}'. اعرض عليه المنتجات المكملة."))
        
    if customer.tier == 'gold':
        suggestions.append(("star", "عميل VIP", "عميل ذهبي. يستحق معاملة خاصة أو خصم ولاء."))
    elif customer.tier == 'silver' and total_spent > 8000:
        remaining = 10000 - total_spent
        suggestions.append(("trending_up", "تحفيز", f"باقي له {remaining} ج.م ويصبح ذهبي. شجعه للشراء."))

    sys_settings = SystemSetting.objects.first()
    if not sys_settings:
        sys_settings = SystemSetting.objects.create(shop_name="متجري")

    context = {
        'customer': customer,
        'report_type': report_type,
        'sys_settings': sys_settings,
        'total_spent': total_spent,
        'balance': balance,
        'total_payments': total_payments,
        'total_manual_debt': total_manual_debt,
        'orders': orders,
        'most_purchased': most_purchased,
        'last_visit': last_visit,
        'avg_days': avg_days,
        'suggestions': suggestions,
        'print_date': timezone.now(),
    }
    return render(request, 'crm/customer_report.html', context)

@login_required
@csrf_exempt
@require_permission('crm', 'edit')
def customer_pay_debt(request, pk):
    if request.method == 'POST':
        try:
            from sales.utils import get_or_create_active_shift
            from financial.models import Account, Transaction
            
            customer = get_object_or_404(Customer, pk=pk)
            data = json.loads(request.body)
            
            amount = Decimal(str(data.get('amount', 0)))
            notes = data.get('notes', '')
            action = data.get('action', 'payment') 
            payment_method = data.get('payment_method', 'cash').lower()
            
            if amount <= 0:
                return JsonResponse({'status': 'error', 'message': 'يجب أن يكون المبلغ أكبر من صفر'}, status=400)

            if action == 'payment':
                from sales.models import DocumentSequence
                payment = CustomerPayment.objects.create(
                    customer=customer,
                    user=request.user,
                    amount=amount,
                    transaction_type='payment',
                    payment_method=payment_method,
                    notes=notes,
                    voucher_number=DocumentSequence.next_number('RV'),  # سند قبض
                )

                # --- Create Financial Transaction ---
                current_shift, _ = get_or_create_active_shift(request.user)
                acc_type_map = {
                    'cash': ('CASH_DRAWER', 'درج الكاشير'),
                    'visa': ('BANK', 'حساب البنك'),
                    'bank': ('BANK', 'حساب البنك'),
                    'wallet': ('VODAFONE_CASH', 'محفظة فودافون كاش'),
                    'instapay': ('INSTAPAY', 'إنستا باي')
                }
                acc_type, acc_name = acc_type_map.get(payment_method, ('CASH_DRAWER', 'درج الكاشير'))
                
                account = Account.objects.filter(account_type=acc_type, is_active=True).first()
                if not account:
                    account = Account.objects.create(name=acc_name, account_type=acc_type, balance=0, is_active=True)
                
                Transaction.objects.create(
                    shift=current_shift,
                    account=account,
                    transaction_type='INCOME',
                    amount=amount,
                    description=f"سداد مديونية من العميل: {customer.first_name} {customer.last_name}",
                    created_by=request.user,
                    customer_payment=payment,  # Phase 1.2: link to source document
                )

                # Phase 8.2: apply this receipt to the customer's oldest open invoices.
                from .allocation import allocate_payment_fifo
                allocate_payment_fifo(payment)
            else:
                # Manual Debt (Sحب)
                payment = CustomerPayment.objects.create(
                    customer=customer,
                    user=request.user,
                    amount=amount,
                    transaction_type='debt',
                    payment_method=payment_method,
                    notes=notes
                )
            
            return JsonResponse({
                'status': 'success', 
                'message': 'تم تسجيل العملية بنجاح',
                'payment_id': payment.id
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=400)

@login_required
@require_permission('crm', 'view')
def payment_receipt(request, payment_id):
    payment = get_object_or_404(CustomerPayment, pk=payment_id)
    sys_settings = SystemSetting.objects.first()
    if not sys_settings:
        sys_settings = SystemSetting.objects.create(shop_name="اسم المحل الافتراضي")

    return render(request, 'crm/payment_receipt.html', {
        'payment': payment, 
        'customer': payment.customer,
        'sys_settings': sys_settings
    })

@login_required
@require_permission('crm', 'create')
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerForm()
    return render(request, 'crm/customer_form.html', {'form': form, 'title': 'إضافة عميل جديد'})

@login_required
@require_permission('crm', 'edit')
def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'crm/customer_form.html', {'form': form, 'title': 'تعديل بيانات عميل'})

@login_required
@require_permission('crm', 'delete')
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        return redirect('customer_list')
    return render(request, 'crm/customer_confirm_delete.html', {'object': customer, 'title': 'حذف عميل'})


# ─────────────────────────────────────────────
#  POS delivery-screen quick client lookup/create (phone-first flow)
# ─────────────────────────────────────────────

@login_required
@require_permission('crm', 'view')
def customer_lookup_by_phone(request):
    """Used by the POS delivery screen: enter a phone number, get back the matching
    customer (name/address/phone2) if one already exists, so the cashier never has to
    retype anything already on file."""
    phone = (request.GET.get('phone') or '').strip()
    if not phone:
        return JsonResponse({'found': False})
    customer = Customer.objects.filter(Q(phone=phone) | Q(phone2=phone)).first()
    if not customer:
        return JsonResponse({'found': False})
    return JsonResponse({
        'found': True,
        'customer': {
            'id': customer.id,
            'name': f"{customer.first_name} {customer.last_name}".strip(),
            'phone': customer.phone,
            'phone2': customer.phone2,
            'address': customer.address or '',
            'balance': float(customer.get_balance()),
            'credit_limit': float(customer.credit_limit or 0),
            'customer_type': customer.customer_type,
        },
    })


@login_required
@require_permission('crm', 'create')
@require_POST
def customer_quick_create(request):
    """Creates a customer from the delivery-screen "no existing client" form. The phone
    number was already entered on the lookup screen, so it's taken as-is here rather than
    asked for again."""
    try:
        payload = json.loads(request.body or b'{}')
    except ValueError:
        payload = request.POST

    phone = (payload.get('phone') or '').strip()
    name = (payload.get('name') or '').strip()
    address = (payload.get('address') or '').strip()
    phone2 = (payload.get('phone2') or '').strip()

    if not phone:
        return JsonResponse({'status': 'error', 'message': 'رقم الهاتف مطلوب'}, status=400)
    if not name:
        return JsonResponse({'status': 'error', 'message': 'اسم العميل مطلوب'}, status=400)
    if Customer.objects.filter(phone=phone).exists():
        return JsonResponse({'status': 'error', 'message': 'يوجد عميل بهذا الرقم بالفعل'}, status=400)

    parts = name.split(' ', 1)
    first_name = parts[0]
    last_name = parts[1] if len(parts) > 1 else ''

    customer = Customer.objects.create(
        first_name=first_name, last_name=last_name, phone=phone, phone2=phone2, address=address,
    )
    return JsonResponse({
        'status': 'ok',
        'customer': {
            'id': customer.id, 'name': name, 'phone': customer.phone,
            'phone2': customer.phone2, 'address': customer.address or '',
            'balance': 0.0, 'credit_limit': float(customer.credit_limit or 0),
            'customer_type': customer.customer_type,
        },
    })


@login_required
@require_permission('crm', 'edit')
@require_POST
def customer_quick_update(request, pk):
    """Inline "تعديل" from the delivery-screen lookup result — edits name/address/phone2
    without leaving the POS screen."""
    customer = get_object_or_404(Customer, pk=pk)
    try:
        payload = json.loads(request.body or b'{}')
    except ValueError:
        payload = request.POST

    name = (payload.get('name') or '').strip()
    if name:
        parts = name.split(' ', 1)
        customer.first_name = parts[0]
        customer.last_name = parts[1] if len(parts) > 1 else ''
    if 'address' in payload:
        customer.address = (payload.get('address') or '').strip()
    if 'phone2' in payload:
        customer.phone2 = (payload.get('phone2') or '').strip()
    customer.save(update_fields=['first_name', 'last_name', 'address', 'phone2'])

    return JsonResponse({
        'status': 'ok',
        'customer': {
            'id': customer.id, 'name': f"{customer.first_name} {customer.last_name}".strip(),
            'phone': customer.phone, 'phone2': customer.phone2, 'address': customer.address or '',
            'balance': float(customer.get_balance()), 'credit_limit': float(customer.credit_limit or 0),
            'customer_type': customer.customer_type,
        },
    })