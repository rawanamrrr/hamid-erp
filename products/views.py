from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count, F, Q
from decimal import Decimal, InvalidOperation
from io import BytesIO
import base64
from django.contrib.auth.decorators import login_required
from accounts.permissions import require_permission, require_granular_action
from django.http import JsonResponse , HttpResponse
from django.views.decorators.http import require_POST
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction as db_transaction
from django.contrib import messages
import json
import csv
from PIL import Image


from .models import (
    Product, Category, Kind, Size, UnitOfMeasure,
    Supplier, StockTransaction, Warehouse, WarehouseStock,
    PurchaseInvoice, PurchaseInvoiceItem, SupplierPayment, ProductImage,
    PurchaseOrder, PurchaseOrderItem, SupplierProduct, StockBatch,
    PurchaseReturn, PurchaseReturnItem
)
from django.views.decorators.csrf import csrf_exempt
from financial.models import Account, Transaction, DailyShift
from django.utils import timezone
from .forms import (
    ProductForm, CategoryForm, KindForm, SizeForm, UnitOfMeasureForm,
    StockTransactionForm, SupplierForm, WarehouseForm, StockTransferForm,
    PurchaseOrderForm, RawMaterialForm,
)
from settings.models import SystemSetting


def to_decimal(value, default=0):
    """
    Safely convert any numeric-like value to Decimal.
    In production we still rely on form validation; this is mainly
    to normalize types and avoid float/Decimal mixing.
    """
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(str(default))


def _next_sku(prefix=''):
    """Next free SKU for `prefix` (e.g. '' for plain products, 'RM-' for raw materials).

    Derives it from the highest existing numeric suffix under that prefix — NOT from
    whichever product happens to have been created last (Product.objects.order_by('id')
    .last()), which broke the moment a differently-prefixed SKU (e.g. a raw material's
    'RM-10004') became the most-recent row: its .isdigit() check failed, silently
    falling back to a hardcoded "10001" that already existed, so every new product save
    hit a duplicate-SKU validation error.
    """
    import re
    pattern = re.compile(r'^' + re.escape(prefix) + r'(\d+)$')
    max_n = 10000
    for sku in Product.objects.filter(sku__startswith=prefix).values_list('sku', flat=True):
        m = pattern.match(sku or '')
        if m:
            max_n = max(max_n, int(m.group(1)))
    return f"{prefix}{max_n + 1}"

def notify_price_changes(request, changes, price_type_label="أسعار المنتجات"):
    if not changes:
        return
        
    try:
        from django.contrib.auth.models import User
        from notifications.models import Notification
        from sales.email_utils import send_alert_email
        from django.template.loader import render_to_string
        from django.utils.timezone import localtime
        
        user_name = request.user.get_full_name() or request.user.username
        
        # 1. In-App Notification
        superusers = User.objects.filter(is_superuser=True)
        message_lines = [f"{c['name']}: {c['old']} -> {c['new']} ({c['type']})" for c in changes[:5]]
        if len(changes) > 5:
            message_lines.append("...والمزيد")
        msg_str = "\n".join(message_lines)
        
        for su in superusers:
            Notification.objects.create(
                recipient=su,
                title=f"تعديل {price_type_label}",
                message=f"قام {user_name} بتعديل الأسعار:\n{msg_str}",
                created_by=request.user
            )
            
        # 2. Email Alert
        html_body = render_to_string('emails/price_changed.html', {
            'user_name': user_name,
            'timestamp': localtime().strftime('%Y-%m-%d %H:%M:%S'),
            'price_type_label': price_type_label,
            'changes': changes
        })
        send_alert_email("تنبيه: تعديل أسعار منتجات", html_body)
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception("Failed to send price change notifications: %s", str(e))





def get_filtered_products(request):
    # Raw materials have their own dedicated screen (raw_material_list) — they're
    # ingredients, not sellable products, so they don't belong mixed into this list.
    queryset = Product.objects.filter(is_raw_material=False).order_by('-created_at')

    search_query = request.GET.get('search', '')
    barcode_search = request.GET.get('barcode_search', '')
    category_id = request.GET.get('category', '')
    supplier_id = request.GET.get('supplier', '')
    stock_status = request.GET.get('stock_status', '')
    is_active = request.GET.get('is_active', '')

    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) | 
            Q(sku__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
    if barcode_search:
        queryset = queryset.filter(barcode=barcode_search)
    if category_id:
        queryset = queryset.filter(category_id=category_id)
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)
    if is_active:
        queryset = queryset.filter(is_active=(is_active == '1'))

    if stock_status == 'low':
        queryset = queryset.filter(stock_quantity__lte=F('low_stock_threshold'), stock_quantity__gt=0)
    elif stock_status == 'out':
        queryset = queryset.filter(stock_quantity__lte=0)
    elif stock_status == 'ok':
        queryset = queryset.filter(stock_quantity__gt=F('low_stock_threshold'))

    return queryset


# ==========================================
# NEW FEATURE 1: Public Product Page
# ==========================================
def public_product_detail(request, sku):
    product = get_object_or_404(Product, sku=sku, is_active=True)
    # Get base URL to ensure QR points exactly here, but for rendering page we just need product
    return render(request, 'products/product_public.html', {
        'product': product, 
        'title': product.name,
        'public_url': request.build_absolute_uri()
    })


# ==========================================
# NEW FEATURE 2: Print QR Codes (A4)
# ==========================================
@login_required
@require_permission('products', 'view')
def print_product_qrs(request):
    # This uses the same filtering as the list, meaning it can print "by category"
    queryset = get_filtered_products(request)
    
    # If explicit IDs were selected (for "custom selection" printing)
    selected_ids = request.GET.get('ids', '')
    if selected_ids:
        id_list = [int(i) for i in selected_ids.split(',') if i.isdigit()]
        queryset = queryset.filter(id__in=id_list)

    # Get the base host URL to generate full QR code links
    base_url = request.build_absolute_uri('/')[:-1]

    return render(request, 'products/print_qrs.html', {
        'products': queryset, 
        'base_url': base_url,
        'title': 'طباعة QR للمنتجات'
    })


# ==========================================
# NEW FEATURE 3: Bulk Price Management
# ==========================================
@login_required
@require_permission('products', 'view')
def bulk_price_manage(request):
    categories = Category.objects.all()
    products = get_filtered_products(request)
    
    return render(request, 'products/bulk_price_manage.html', {
        'products': products,
        'categories': categories,
        'title': 'الإدارة المجمعة للأسعار',
        'query_params': request.GET.copy()
    })

@login_required
@require_permission('products', 'view')
@require_POST
def bulk_price_save_api(request):
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        
        if not items:
            return JsonResponse({'success': False, 'error': 'لا توجد بيانات للتحديث'})

        with db_transaction.atomic():
            changes = []
            for item in items:
                try:
                    product = Product.objects.get(id=item.get('id'))
                    old_r = Decimal(str(product.price_retail))
                    old_s = Decimal(str(product.price_semi_wholesale))
                    old_w = Decimal(str(product.price_wholesale))
                    
                    product.price_retail = Decimal(str(item.get('price_retail', product.price_retail)))
                    product.price_semi_wholesale = Decimal(str(item.get('price_semi_wholesale', product.price_semi_wholesale)))
                    product.price_wholesale = Decimal(str(item.get('price_wholesale', product.price_wholesale)))
                    
                    if old_r != product.price_retail:
                        changes.append({'name': product.name, 'type': 'قطاعي', 'old': old_r, 'new': product.price_retail})
                    if old_s != product.price_semi_wholesale:
                        changes.append({'name': product.name, 'type': 'نصف جملة', 'old': old_s, 'new': product.price_semi_wholesale})
                    if old_w != product.price_wholesale:
                        changes.append({'name': product.name, 'type': 'جملة', 'old': old_w, 'new': product.price_wholesale})
                        
                    product.save(update_fields=['price_retail', 'price_semi_wholesale', 'price_wholesale'])
                except (Product.DoesNotExist, ValueError, TypeError):
                    continue # Skip invalid items
            
            if changes:
                notify_price_changes(request, changes, 'أسعار البيع')

        return JsonResponse({'success': True, 'message': f'تم تحديث أسعار {len(items)} منتج بنجاح.'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==========================================
# NEW FEATURE 3.5: Bulk Cost Price Management
# ==========================================
@login_required
@require_permission('products', 'view')
def bulk_cost_price_manage(request):
    categories = Category.objects.all()
    products = get_filtered_products(request)
    
    return render(request, 'products/bulk_cost_price_manage.html', {
        'products': products,
        'categories': categories,
        'title': 'تعديل أسعار التكلفة',
        'query_params': request.GET.copy()
    })

@login_required
@require_permission('products', 'view')
@require_POST
def bulk_cost_price_save_api(request):
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        
        if not items:
            return JsonResponse({'success': False, 'error': 'لا توجد بيانات للتحديث'})

        with db_transaction.atomic():
            changes = []
            for item in items:
                try:
                    product = Product.objects.get(id=item.get('id'))
                    new_cost = Decimal(str(item.get('cost_price', product.cost_price)))
                    if new_cost < 0:
                        continue # Skip negative prices
                        
                    old_cost = Decimal(str(product.cost_price))
                    if old_cost != new_cost:
                        product.cost_price = new_cost
                        product.save(update_fields=['cost_price'])
                        changes.append({'name': product.name, 'type': 'تكلفة', 'old': old_cost, 'new': new_cost})
                        
                except (Product.DoesNotExist, ValueError, TypeError):
                    continue # Skip invalid items
                    
            if changes:
                notify_price_changes(request, changes, 'سعر التكلفة')

        return JsonResponse({'success': True, 'message': f'تم تحديث أسعار التكلفة لـ {len(items)} منتج بنجاح.'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# --- Sold Items List (Dashboard View) ---
@login_required
@require_granular_action('sales', 'sold_items', 'products', 'view')
def sold_items_list(request):
    queryset = StockTransaction.objects.filter(transaction_type='OUT').select_related('product')
    
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if search_query:
        queryset = queryset.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(reference_number__icontains=search_query)
        )
    
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)
        
    queryset = queryset.order_by('-created_at')
    
    total_quantity = 0
    total_revenue = 0
    for item in queryset:
        total_quantity += item.quantity
        total_revenue += item.total_price

    context = {
        'sold_items': queryset,
        'title': 'سجل مبيعات الأصناف',
        'total_quantity': total_quantity,
        'total_revenue': total_revenue,
        'filter_params': {
            'search': search_query,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    return render(request, 'products/sold_items_list.html', context)

# --- Sold Items Report ---
@login_required
@require_granular_action('sales', 'sold_items', 'products', 'view')
def sold_items_report(request):
    queryset = StockTransaction.objects.filter(transaction_type='OUT').select_related('product')
    
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if search_query:
        queryset = queryset.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(reference_number__icontains=search_query)
        )
    
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)
        
    queryset = queryset.order_by('-created_at')
    
    total_quantity = 0
    total_revenue = 0
    for item in queryset:
        total_quantity += item.quantity
        total_revenue += item.total_price

    sys_settings = SystemSetting.objects.first()

    context = {
        'sold_items': queryset,
        'title': 'تقرير مبيعات الأصناف',
        'total_quantity': total_quantity,
        'total_revenue': total_revenue,
        'sys_settings': sys_settings,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'products/sold_items_report.html', context)

@login_required
@require_permission('products', 'view')
def product_list(request):
    queryset = get_filtered_products(request)

    all_products = Product.objects.filter(is_raw_material=False)
    stats = {
        'total_count': all_products.count(),
        'low_stock_count': all_products.exclude(category__is_menu_category=True).filter(stock_quantity__lte=F('low_stock_threshold'), stock_quantity__gt=0).count(),
        'out_of_stock_count': all_products.exclude(category__is_menu_category=True).filter(stock_quantity__lte=0).count(),
        'total_value': all_products.aggregate(total=Sum(F('stock_quantity') * F('cost_price')))['total'] or 0,
    }

    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    categories = Category.objects.all()
    suppliers = Supplier.objects.all()
    sys_settings = SystemSetting.objects.first()

    context = {
        'products': page_obj,
        'stats': stats,
        'categories': categories,
        'suppliers': suppliers,
        'title': 'إدارة المنتجات',
        'query_params': request.GET.copy(),
        'sys_settings': sys_settings
    }
    return render(request, 'products/product_list.html', context)


# ==========================================
# PRODUCTS IMPORT / EXPORT — full round-trip
# ==========================================
# Column layout shared by xlsx/csv export, import template, and the importer.
# Built dynamically from the store's active market profile (settings.market_profiles)
# so a cafe never sees clothing/pharmacy-only columns (material, scientific_name, ...)
# and vice versa — same fields the product form itself shows for this market. (key, arabic_label).
def _get_export_headers():
    from settings.market_profiles import get_market_profile
    sys_settings = SystemSetting.objects.first()
    mp = get_market_profile(sys_settings.market_type if sys_settings else None)
    show = mp['show']
    terms = mp['terms']

    headers = [
        ('id',       'ID (لا تعدّل)'),
        ('sku',      'SKU (كود المنتج)'),
        ('name',     'اسم المنتج'),
        ('barcode',  'الباركود (EAN-13) — اتركه فارغاً للتوليد التلقائي'),
        ('category', 'القسم'),
        ('kind',     'النوع'),
        ('supplier', 'المورد'),
        ('sizes',    'المقاسات/الأحجام — بدون سعر مفصولة بفاصلة، أو "الاسم:السعر" لكل حجم بسعر خاص به'),
    ]
    if show.get('pharmacy_specs'):
        headers += [
            ('scientific_name', 'الاسم العلمي'),
            ('packaging_type',  'نوع التعبئة'),
            ('strips_per_box',  'عدد الأشرطة في العلبة'),
        ]
    if show.get('clothes_specs'):
        headers += [
            ('material', 'الخامة'),
            ('pattern',  'النقشة / المميزات'),
            ('color',    'اللون'),
        ]
    headers.append(('unit_measure', f"{terms['unit_label']} (PCS/KG/LTR...)"))
    headers.append(('cost_price', terms['cost_label']))
    headers.append(('price_retail', terms['retail_label']))
    if show.get('wholesale_price'):
        headers += [
            ('price_semi_wholesale', 'سعر نصف الجملة'),
            ('price_wholesale',      'سعر الجملة'),
        ]
    headers.append(('low_stock_threshold', 'حد التنبيه للنواقص'))
    if show.get('pieces_per_package'):
        headers.append(('pieces_per_package', terms['pieces_label']))
    headers.append(('is_active', 'نشط (1/0)'))
    headers.append(('stock_quantity', 'الرصيد الكلي (للقراءة فقط)'))
    return headers

# Arabic unit labels → model code, for tolerant parsing of unit_measure.
_UNIT_LABEL_TO_CODE = {label: code for code, label in Product.UNIT_CHOICES}
_UNIT_LABEL_TO_CODE.update({code: code for code, _ in Product.UNIT_CHOICES})


def _serialize_product_row(product):
    """Return a dict {key: cell_value} for one product, using the export schema."""
    if product.has_variants:
        # has_variants products (e.g. a cafe drink in "صغير/كبير") carry their sizes via
        # ProductVariant.size, each with its OWN price — read from there instead of the
        # plain product.sizes M2M, and encode as "name:price" pairs so the price round-trips
        # too (a bare name list would silently drop per-size pricing on re-import).
        seen = {}
        for v in product.variants.all():
            if v.size_id and v.size_id not in seen:
                seen[v.size_id] = f"{v.size.name}:{v.price}"
        sizes_str = ', '.join(seen.values())
    else:
        sizes_str = ', '.join(s.name for s in product.sizes.all())
    return {
        'id':                   product.id,
        'sku':                  product.sku or '',
        'name':                 product.name or '',
        'barcode':              product.barcode or '',
        'category':             product.category.name if product.category else '',
        'kind':                 product.kind.name if product.kind else '',
        'supplier':             product.supplier.name if product.supplier else '',
        'sizes':                sizes_str,
        'scientific_name':      product.scientific_name or '',
        'packaging_type':       product.packaging_type or '',
        'strips_per_box':       product.strips_per_box or 1,
        'material':             product.material or '',
        'pattern':              product.pattern or '',
        'color':                product.color or '',
        'unit_measure':         product.unit_measure or '',
        'cost_price':           product.cost_price,
        'price_retail':         product.price_retail,
        'price_semi_wholesale': product.price_semi_wholesale,
        'price_wholesale':      product.price_wholesale,
        'low_stock_threshold':  product.low_stock_threshold,
        'pieces_per_package':   product.pieces_per_package,
        'is_active':            1 if product.is_active else 0,
        'stock_quantity':       product.stock_quantity,
    }


def _to_decimal_or_none(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value).strip().replace(',', ''))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _to_bool(value, default=True):
    if value is None or value == '':
        return default
    s = str(value).strip().lower()
    if s in ('1', 'true', 'yes', 'y', 'نعم', 'نشط'):
        return True
    if s in ('0', 'false', 'no', 'n', 'لا', 'غير نشط'):
        return False
    return default


def _generate_ean13_from_sku(sku):
    if not sku:
        return ''
    digits = ''.join(filter(str.isdigit, sku)).zfill(12)[-12:]
    s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits))
    return digits + str((10 - (s % 10)) % 10)


def _next_auto_sku():
    """Mirror the next-SKU logic from product_create."""
    last_product = Product.objects.order_by('id').last()
    candidate = "10001"
    if last_product and last_product.sku and last_product.sku.isdigit():
        try:
            candidate = str(int(last_product.sku) + 1)
        except ValueError:
            pass
    while Product.objects.filter(sku=candidate).exists():
        try:
            candidate = str(int(candidate) + 1)
        except ValueError:
            candidate = candidate + '_1'
            break
    return candidate


_DEFAULT_COL_WIDTH = 18
_COL_WIDTH_OVERRIDES = {'id': 8, 'name': 30, 'sizes': 26, 'stock_quantity': 18}


def _write_xlsx_response(filename, products):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers = _get_export_headers()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (_, label) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, product in enumerate(products, start=2):
        data = _serialize_product_row(product)
        for col_idx, (key, _) in enumerate(headers, start=1):
            value = data[key]
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for i, (key, _) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTH_OVERRIDES.get(key, _DEFAULT_COL_WIDTH)
    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _write_csv_response(filename, products):
    headers = _get_export_headers()
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff'.encode('utf8'))
    writer = csv.writer(response)

    writer.writerow([label for _, label in headers])
    for product in products:
        data = _serialize_product_row(product)
        writer.writerow([data[key] for key, _ in headers])
    return response


@login_required
@require_permission('products', 'view')
def export_products_excel(request):
    """Export filtered products to Excel (.xlsx) or CSV based on ?format=."""
    fmt = (request.GET.get('format', 'xlsx') or 'xlsx').lower()
    queryset = get_filtered_products(request).prefetch_related('sizes', 'variants__size').select_related(
        'category', 'kind', 'supplier'
    )
    if fmt == 'csv':
        return _write_csv_response('products_export.csv', queryset)
    return _write_xlsx_response('products_export.xlsx', queryset)


@login_required
@require_permission('products', 'view')
def export_products_pdf(request):
    products = get_filtered_products(request).prefetch_related('warehouse_stocks')
    warehouses = Warehouse.objects.filter(is_active=True).order_by('id')
    
    # تحضير أرصدة كل منتج في كل مخزن على حدة
    for product in products:
        stock_dict = {ws.warehouse_id: ws.quantity for ws in product.warehouse_stocks.all()}
        # إنشاء قائمة بالكميات تترتب بنفس ترتيب المخازن
        product.warehouse_quantities = [stock_dict.get(wh.id, 0) for wh in warehouses]

    return render(request, 'products/print_pdf.html', {
        'products': products, 
        'warehouses': warehouses, 
        'title': 'طباعة المنتجات'
    })

@login_required
@require_permission('products', 'view')
def export_products_barcode(request):
    products = get_filtered_products(request)
    sys_settings = SystemSetting.objects.first()
    return render(request, 'products/print_barcodes.html', {
        'products': products,
        'shop_name': sys_settings.shop_name if sys_settings else '',
    })

import json

@login_required
@require_permission('products', 'view')
def barcode_generator_view(request):
    """
    Advanced Barcode Generator: Live preview, multiple formats (Thermal/A4), custom text.
    Accepts specific product IDs via '?ids=1,2,3' or falls back to filtered list.
    """
    ids_param = request.GET.get('ids', '')
    if ids_param:
        try:
            id_list = [int(i.strip()) for i in ids_param.split(',') if i.strip().isdigit()]
            products = Product.objects.filter(id__in=id_list, is_active=True).prefetch_related('sizes', 'category')
        except ValueError:
            products = Product.objects.none()
    else:
        products = get_filtered_products(request).prefetch_related('sizes', 'category')
    
    product_data = []
    for p in products:
        product_data.append({
            'id': p.id,
            'name': p.name,
            'sku': p.sku,
            'barcode': p.barcode or p.sku,
            'price': float(p.price_retail),
            'category': p.category.name if p.category else 'عام',
            'sizes': [{'id': s.id, 'name': s.name} for s in p.sizes.all()]
        })
        
    return render(request, 'products/barcode_generator.html', {
        'products_json': json.dumps(product_data),
        'title': 'مولد الباركود المتقدم'
    })

@login_required
@require_permission('products', 'view')
def download_import_template(request):
    """Download import template pre-filled with current products as editable example."""
    fmt = (request.GET.get('format', 'xlsx') or 'xlsx').lower()
    queryset = get_filtered_products(request).prefetch_related('sizes', 'variants__size').select_related(
        'category', 'kind', 'supplier'
    )
    if fmt == 'csv':
        return _write_csv_response('products_template.csv', queryset)
    return _write_xlsx_response('products_template.xlsx', queryset)


def _read_uploaded_rows(uploaded_file):
    """Yield list-of-strings rows from .csv/.xlsx/.xls file. First row = headers."""
    name = (uploaded_file.name or '').lower()

    if name.endswith('.csv'):
        decoded = uploaded_file.read().decode('utf-8-sig', errors='replace').splitlines()
        for row in csv.reader(decoded):
            yield row
        return

    import openpyxl
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        yield ['' if v is None else v for v in row]


def _build_header_index_map(header_row):
    """Map schema key → column index based on the uploaded header row.

    Tolerant: matches by Arabic label OR by the schema key itself, case-
    insensitive, ignoring everything after the first '(' or '—'.
    """
    def norm(s):
        s = str(s or '').strip().lower()
        for sep in ('(', '—', '-'):
            if sep in s:
                s = s.split(sep, 1)[0]
        return s.strip()

    label_lookup = {}
    for key, label in _get_export_headers():
        label_lookup[norm(label)] = key
        label_lookup[norm(key)] = key

    mapping = {}
    for idx, cell in enumerate(header_row):
        key = label_lookup.get(norm(cell))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _cell(row, header_map, key):
    idx = header_map.get(key)
    if idx is None or idx >= len(row):
        return ''
    val = row[idx]
    return '' if val is None else str(val).strip()


@login_required
@require_granular_action('products', 'import', 'products', 'view')
@require_POST
def import_products_excel(request):
    """
    Import products from xlsx/csv. Match rows by ID, then SKU.
    Empty both → create new product (auto-SKU + auto-EAN13).
    """
    uploaded = request.FILES.get('excel_file')
    if not uploaded:
        messages.error(request, 'الرجاء اختيار ملف')
        return redirect('product_list')

    try:
        rows_iter = _read_uploaded_rows(uploaded)
    except ImportError:
        messages.error(request, 'لم يتم العثور على مكتبة openpyxl لدعم ملفات Excel. يرجى استخدام صيغة CSV.')
        return redirect('product_list')
    except Exception as e:
        messages.error(request, f'تعذر قراءة الملف: {str(e)}')
        return redirect('product_list')

    rows = list(rows_iter)
    if not rows:
        messages.error(request, 'الملف فارغ.')
        return redirect('product_list')

    header_row = rows[0]
    header_map = _build_header_index_map(header_row)

    if 'sku' not in header_map and 'id' not in header_map and 'name' not in header_map:
        messages.error(
            request,
            'تعذر التعرف على أعمدة الملف. تأكد من استخدام القالب الجديد (الأعمدة: ID/SKU/الاسم...).'
        )
        return redirect('product_list')

    created = updated = skipped = 0
    errors = []

    try:
        with db_transaction.atomic():
            for line_no, row in enumerate(rows[1:], start=2):
                if not any(str(c).strip() for c in row if c is not None):
                    continue

                row_id = _cell(row, header_map, 'id')
                sku = _cell(row, header_map, 'sku')
                name = _cell(row, header_map, 'name')

                product = None
                if row_id and row_id.isdigit():
                    product = Product.objects.filter(id=int(row_id)).first()
                if product is None and sku:
                    product = Product.objects.filter(sku=sku).first()

                is_new = product is None
                if is_new:
                    if not name:
                        errors.append(f'سطر {line_no}: اسم المنتج مطلوب لإنشاء منتج جديد.')
                        skipped += 1
                        continue
                    product = Product(name=name)
                    product.sku = sku or _next_auto_sku()
                else:
                    if name:
                        product.name = name
                    if sku and sku != product.sku:
                        if Product.objects.filter(sku=sku).exclude(pk=product.pk).exists():
                            errors.append(
                                f'سطر {line_no}: SKU "{sku}" مستخدم لمنتج آخر — تم تخطي السطر.'
                            )
                            skipped += 1
                            continue
                        product.sku = sku

                barcode = _cell(row, header_map, 'barcode')
                if barcode:
                    product.barcode = barcode
                elif not product.barcode:
                    product.barcode = _generate_ean13_from_sku(product.sku)

                cat_name = _cell(row, header_map, 'category')
                if cat_name:
                    product.category, _c = Category.objects.get_or_create(name=cat_name)
                elif 'category' in header_map:
                    product.category = None

                kind_name = _cell(row, header_map, 'kind')
                if kind_name and product.category:
                    product.kind, _c = Kind.objects.get_or_create(
                        category=product.category, name=kind_name
                    )
                elif 'kind' in header_map and not kind_name:
                    product.kind = None

                supplier_name = _cell(row, header_map, 'supplier')
                if supplier_name:
                    product.supplier = Supplier.objects.filter(name=supplier_name).first()
                elif 'supplier' in header_map:
                    product.supplier = None

                if 'scientific_name' in header_map:
                    product.scientific_name = _cell(row, header_map, 'scientific_name')
                if 'packaging_type' in header_map:
                    product.packaging_type = _cell(row, header_map, 'packaging_type')
                if 'strips_per_box' in header_map:
                    strips_box_val = _cell(row, header_map, 'strips_per_box')
                    if strips_box_val and strips_box_val.isdigit():
                        product.strips_per_box = int(strips_box_val)
                if 'material' in header_map:
                    product.material = _cell(row, header_map, 'material')
                if 'pattern' in header_map:
                    product.pattern = _cell(row, header_map, 'pattern')
                if 'color' in header_map:
                    product.color = _cell(row, header_map, 'color')

                unit_raw = _cell(row, header_map, 'unit_measure')
                if unit_raw:
                    product.unit_measure = _UNIT_LABEL_TO_CODE.get(
                        unit_raw, _UNIT_LABEL_TO_CODE.get(unit_raw.upper(), 'MTR')
                    )
                elif is_new:
                    product.unit_measure = product.unit_measure or 'MTR'

                for field in ('cost_price', 'price_retail', 'price_semi_wholesale', 'price_wholesale'):
                    val = _to_decimal_or_none(_cell(row, header_map, field))
                    if val is not None:
                        setattr(product, field, val)
                    elif is_new:
                        setattr(product, field, Decimal('0.00'))

                threshold = _to_decimal_or_none(_cell(row, header_map, 'low_stock_threshold'))
                if threshold is not None:
                    product.low_stock_threshold = threshold
                elif is_new:
                    product.low_stock_threshold = Decimal('10.00')

                pieces_raw = _cell(row, header_map, 'pieces_per_package')
                if pieces_raw:
                    try:
                        product.pieces_per_package = max(1, int(Decimal(pieces_raw)))
                    except (InvalidOperation, ValueError):
                        pass

                if 'is_active' in header_map:
                    active_raw = _cell(row, header_map, 'is_active')
                    if active_raw != '' or is_new:
                        product.is_active = _to_bool(active_raw, default=True)

                product.save()

                sizes_raw = _cell(row, header_map, 'sizes')
                if 'sizes' in header_map:
                    tokens = [t.strip() for t in sizes_raw.replace('،', ',').split(',') if t.strip()] if sizes_raw else []
                    # "الاسم:السعر" tokens (as exported for a has_variants product) mean
                    # each size has its own price — build/update real ProductVariant rows
                    # (price_override) instead of the plain product.sizes M2M, which has
                    # no way to carry a per-size price.
                    priced_tokens = [t for t in tokens if ':' in t]
                    if priced_tokens:
                        from .models import ProductVariant
                        if not product.has_variants:
                            product.has_variants = True
                            product.save(update_fields=['has_variants'])
                        for token in tokens:
                            name_part, _, price_part = token.partition(':')
                            name_part = name_part.strip()
                            if not name_part:
                                continue
                            size_obj, _c = Size.objects.get_or_create(
                                name=name_part, defaults={'size_type': 'custom'}
                            )
                            price_val = _to_decimal_or_none(price_part.strip())
                            defaults = {'price_override': price_val} if price_val is not None else {}
                            ProductVariant.objects.update_or_create(
                                product=product, size=size_obj, color='', defaults=defaults
                            )
                    elif not product.has_variants:
                        # Plain size list (clothing-style, no per-size price) — round-trips
                        # via the product.sizes M2M as before.
                        if sizes_raw:
                            size_objs = []
                            for token in tokens:
                                size_obj, _c = Size.objects.get_or_create(
                                    name=token, defaults={'size_type': 'custom'}
                                )
                                size_objs.append(size_obj)
                            product.sizes.set(size_objs)
                        elif not is_new:
                            product.sizes.clear()
                    # else: has_variants product with a plain (unpriced) size list in this
                    # column — leave its ProductVariant rows alone rather than guessing.

                if is_new:
                    created += 1
                else:
                    updated += 1
    except Exception as e:
        messages.error(request, f'حدث خطأ أثناء الاستيراد، تم التراجع عن جميع التغييرات: {str(e)}')
        return redirect('product_list')

    summary = f'تم الاستيراد بنجاح — جديد: {created}، محدّث: {updated}، تم تخطيه: {skipped}.'
    if errors:
        summary += ' أول الأخطاء: ' + ' | '.join(errors[:10])
        if len(errors) > 10:
            summary += f' (+{len(errors) - 10} أخطاء أخرى)'
        messages.warning(request, summary)
    else:
        messages.success(request, summary)

    return redirect('product_list')



@login_required
@require_permission('products', 'view')
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    stocks = WarehouseStock.objects.filter(product=product)
    sys_settings = SystemSetting.objects.first()

    # Fetch all OrderItems for this product, with related Order and Customer
    from sales.models import OrderItem
    sale_items = (
        OrderItem.objects
        .filter(product=product)
        .select_related('order', 'order__customer')
        .order_by('-order__created_at')[:100]
    )

    # Compute totals for the footer
    total_sold_qty = sum(item.quantity for item in sale_items)
    total_sold_revenue = sum(item.subtotal for item in sale_items)

    return render(request, 'products/product_detail.html', {
        'product': product,
        'stocks': stocks,
        'sale_items': sale_items,
        'total_sold_qty': total_sold_qty,
        'total_sold_revenue': total_sold_revenue,
        'title': product.name,
        'sys_settings': sys_settings
    })

# ==========================================
# Product Report: PDF (print-ready HTML)
# ==========================================
@login_required
@require_permission('products', 'view')
def product_report_pdf(request, pk):
    product = get_object_or_404(Product, pk=pk)
    stocks = WarehouseStock.objects.filter(product=product).select_related('warehouse')

    from sales.models import OrderItem
    sale_items = (
        OrderItem.objects
        .filter(product=product)
        .select_related('order', 'order__customer')
        .order_by('-order__created_at')[:200]
    )

    total_sold_qty = sum(item.quantity for item in sale_items)
    total_sold_revenue = sum(item.subtotal for item in sale_items)
    sys_settings = SystemSetting.objects.first()

    return render(request, 'products/product_report.html', {
        'product': product,
        'stocks': stocks,
        'sale_items': sale_items,
        'total_sold_qty': total_sold_qty,
        'total_sold_revenue': total_sold_revenue,
        'sys_settings': sys_settings,
        'print_date': timezone.now(),
        'title': f'تقرير منتج - {product.name}',
    })


# ==========================================
# Product Report: Excel (.xlsx)
# ==========================================
@login_required
@require_permission('products', 'view')
def product_report_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    product = get_object_or_404(Product, pk=pk)
    stocks = WarehouseStock.objects.filter(product=product).select_related('warehouse')

    from sales.models import OrderItem
    sale_items = (
        OrderItem.objects
        .filter(product=product)
        .select_related('order', 'order__customer')
        .order_by('-order__created_at')[:500]
    )

    sys_settings = SystemSetting.objects.first()
    shop_name = sys_settings.shop_name if sys_settings else 'MR MEKAWY'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير المنتج'
    ws.sheet_view.rightToLeft = True  # RTL layout

    # Styles
    header_font = Font(name='Arial', bold=True, size=18, color='1F4E79')
    section_font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    section_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    label_font = Font(name='Arial', bold=True, size=11, color='555555')
    value_font = Font(name='Arial', size=11)
    money_font = Font(name='Arial', bold=True, size=12, color='047857')
    table_header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    table_header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    total_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    right_align = Alignment(horizontal='right', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')

    row = 1

    # --- Title ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=f'{shop_name} — تقرير منتج')
    title_cell.font = header_font
    title_cell.alignment = center_align
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    date_cell = ws.cell(row=row, column=1, value=f'تاريخ التقرير: {timezone.now().strftime("%Y-%m-%d %H:%M")}')
    date_cell.font = Font(name='Arial', size=10, color='888888')
    date_cell.alignment = center_align
    row += 2

    # Helper: write section header
    def write_section(title):
        nonlocal row
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=title).font = section_font
        ws.cell(row=row, column=1).alignment = right_align
        row += 1

    # Helper: write label-value pair
    def write_pair(label, value, col_label=1, col_value=2):
        nonlocal row
        ws.cell(row=row, column=col_label, value=label).font = label_font
        ws.cell(row=row, column=col_value, value=str(value)).font = value_font
        ws.cell(row=row, column=col_label).alignment = right_align
        ws.cell(row=row, column=col_value).alignment = right_align

    # --- Section: Product Info ---
    write_section('📋 بيانات المنتج')
    write_pair('اسم المنتج:', product.name)
    write_pair('كود المنتج (SKU):', product.sku, 4, 5)
    row += 1
    write_pair('القسم:', product.category.name if product.category else '—')
    write_pair('المورد:', product.supplier.name if product.supplier else '—', 4, 5)
    row += 1
    write_pair('وحدة القياس:', product.get_unit_measure_display())
    write_pair('الحالة:', 'نشط' if product.is_active else 'غير نشط', 4, 5)
    row += 1

    # Specs
    if product.material or product.pattern or product.color:
        write_pair('الخامة:', product.material or '—')
        write_pair('النقشة:', product.pattern or '—', 3, 4)
        write_pair('اللون:', product.color or '—', 5, 6)
        row += 1

    row += 1

    # --- Section: Pricing ---
    write_section('💰 الأسعار')
    for col_idx, (lbl, val) in enumerate([
        ('سعر التكلفة', float(product.cost_price)),
        ('سعر الجملة', float(product.price_wholesale)),
        ('نص جملة', float(product.price_semi_wholesale)),
        ('سعر قطاعي', float(product.price_retail)),
    ], start=1):
        cell_label = ws.cell(row=row, column=col_idx, value=lbl)
        cell_label.font = label_font
        cell_label.alignment = center_align
        cell_val = ws.cell(row=row + 1, column=col_idx, value=val)
        cell_val.font = money_font
        cell_val.alignment = center_align
        cell_val.number_format = '#,##0.00'
    row += 3

    # --- Section: Inventory ---
    write_section('📦 أرصدة المخزون')
    ws.cell(row=row, column=1, value='إجمالي المخزون الكلي:').font = label_font
    total_stock_cell = ws.cell(row=row, column=2, value=float(product.stock_quantity))
    total_stock_cell.font = Font(name='Arial', bold=True, size=14, color='047857')
    total_stock_cell.number_format = '#,##0.##'
    ws.cell(row=row, column=3, value=product.get_unit_measure_display()).font = value_font
    row += 1

    if stocks:
        for stock in stocks:
            ws.cell(row=row, column=1, value=f'🏭 {stock.warehouse.name}').font = value_font
            qty_cell = ws.cell(row=row, column=2, value=float(stock.quantity))
            qty_cell.font = Font(name='Arial', bold=True, size=11)
            qty_cell.number_format = '#,##0.##'
            row += 1
    row += 1

    # --- Section: Sales History ---
    write_section('📊 سجل المبيعات')
    if sale_items:
        # Table Header
        headers = ['رقم الفاتورة', 'التاريخ', 'العميل', 'الكمية', 'سعر الوحدة', 'الإجمالي']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        # Table Data
        for item in sale_items:
            customer_name = 'عميل نقدي'
            if item.order.customer:
                customer_name = f'{item.order.customer.first_name} {item.order.customer.last_name}'.strip()

            data = [
                f'#{item.order.id}',
                item.order.created_at.strftime('%Y-%m-%d'),
                customer_name,
                float(item.quantity),
                float(item.price),
                float(item.subtotal),
            ]
            for col_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = value_font
                cell.alignment = center_align
                cell.border = thin_border
                if col_idx in (4, 5, 6):
                    cell.number_format = '#,##0.00'
            row += 1

        # Totals
        total_sold_qty = sum(float(item.quantity) for item in sale_items)
        total_sold_revenue = sum(float(item.subtotal) for item in sale_items)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        total_label = ws.cell(row=row, column=1, value='الإجمالي')
        total_label.font = Font(name='Arial', bold=True, size=11)
        total_label.alignment = center_align
        total_label.fill = total_fill
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = total_fill
            ws.cell(row=row, column=col).border = thin_border

        qty_total = ws.cell(row=row, column=4, value=total_sold_qty)
        qty_total.font = Font(name='Arial', bold=True, size=11, color='1F4E79')
        qty_total.number_format = '#,##0.##'
        qty_total.alignment = center_align

        rev_total = ws.cell(row=row, column=6, value=total_sold_revenue)
        rev_total.font = Font(name='Arial', bold=True, size=12, color='047857')
        rev_total.number_format = '#,##0.00'
        rev_total.alignment = center_align
    else:
        ws.cell(row=row, column=1, value='لا توجد مبيعات مسجلة لهذا المنتج').font = Font(name='Arial', color='999999')
    row += 2

    # Column widths
    col_widths = [20, 18, 25, 14, 16, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Write to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = product.sku or product.name[:20]
    response['Content-Disposition'] = f'attachment; filename="product_report_{safe_name}.xlsx"'
    wb.save(response)
    return response


def _save_size_quantities(product, request, selected_size_ids=None):
    """Per-size stock counts entered on the add/edit product form (clothes market).

    Creates/updates a plain ProductVariant per size (color left blank) — the same model
    already used by the dedicated "إدارة الخيارات (مقاس × لون)" page, so stock entered
    here is immediately visible/editable there too, and vice versa.

    A ProductVariant row (defaulting to 0 stock) is ensured for EVERY selected size, not
    just ones with a quantity typed in — otherwise picking sizes without immediately
    entering counts (e.g. adding the dress now, receiving stock via purchase invoice
    later) leaves has_variants False and the POS size picker never appears at all.
    """
    from .models import ProductVariant
    raw = request.POST.get('size_quantities', '{}')
    try:
        qty_map = json.loads(raw) or {}
    except (ValueError, TypeError):
        qty_map = {}

    selected_size_ids = set(selected_size_ids or [])
    for size_id_str, qty_str in qty_map.items():
        if str(size_id_str).isdigit():
            selected_size_ids.add(int(size_id_str))

    for size_id in selected_size_ids:
        qty_str = qty_map.get(str(size_id))
        if qty_str not in (None, ''):
            try:
                qty = Decimal(str(qty_str))
            except Exception:
                qty = None
        else:
            qty = None
        if qty is not None:
            ProductVariant.objects.update_or_create(
                product=product, size_id=size_id, color='',
                defaults={'stock_quantity': qty},
            )
        else:
            # No quantity entered for this size yet — ensure the variant exists
            # (0 stock) so it shows up in the POS picker/variants page immediately.
            ProductVariant.objects.get_or_create(
                product=product, size_id=size_id, color='',
                defaults={'stock_quantity': Decimal('0')},
            )

    if selected_size_ids and not product.has_variants:
        product.has_variants = True
        product.save(update_fields=['has_variants'])


def _save_product_sizes(product, request):
    """Optional sizes (كبير/وسط...), each with its OWN full price (not a delta over a
    base price). Reuses ProductVariant with color='' — the exact size×color mechanism
    the POS already drives for fashion products: a has_variants product already opens
    "اختر المقاس" BEFORE it's added to the cart, uses the variant's own price
    (price_override) as the line price, and the invoice already prints the variant's
    label. This function is the only new code needed — POS/invoice already handle it.
    """
    from .models import ProductVariant

    has_sizes = request.POST.get('has_sizes') == 'on'
    if not has_sizes:
        product.variants.filter(color='').delete()
        if not product.variants.exists() and product.has_variants:
            product.has_variants = False
            product.save(update_fields=['has_variants'])
        return

    size_ids = request.POST.getlist('size_id')
    prices = request.POST.getlist('size_price')
    kept_ids = []
    for size_id, price in zip(size_ids, prices):
        if not size_id:
            continue
        try:
            price_val = Decimal(str(price or 0))
        except (InvalidOperation, TypeError):
            continue
        variant, _ = ProductVariant.objects.update_or_create(
            product=product, size_id=size_id, color='',
            defaults={'price_override': price_val},
        )
        kept_ids.append(variant.id)

    # Drop sizes that were removed on this save (kept_ids is empty when has_sizes was
    # checked but every row was left blank — clears all of them, same as unchecking).
    product.variants.filter(color='').exclude(id__in=kept_ids).delete()

    update_fields = []
    if kept_ids and not product.has_variants:
        product.has_variants = True
        update_fields.append('has_variants')
    elif not kept_ids and product.has_variants and not product.variants.exists():
        product.has_variants = False
        update_fields.append('has_variants')

    # price_retail has no real meaning for a multi-size product (each size prices itself) —
    # backfill it to the lowest size's price so reports/dashboard code that reads
    # price_retail directly (instead of variant prices) still shows a sane "from" number.
    if kept_ids:
        prices_kept = [Decimal(str(p)) for size_id, p in zip(size_ids, prices) if size_id]
        if prices_kept:
            product.price_retail = min(prices_kept)
            update_fields.append('price_retail')

    if update_fields:
        product.save(update_fields=update_fields)


@login_required
@require_granular_action('products', 'create', 'products', 'view')
def product_create(request):
    # --- 1. Calculate Next SKU Logic ---
    next_sku = _next_sku()

    all_sizes = list(Size.objects.filter(is_active=True).values('id', 'name', 'sort_order'))

    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            if not product.sku:
                product.sku = next_sku
                while Product.objects.filter(sku=product.sku).exists():
                    try:
                        product.sku = str(int(product.sku) + 1)
                    except ValueError:
                        break
            # Auto-generate barcode from SKU if not provided
            if not product.barcode and product.sku:
                digits = ''.join(filter(str.isdigit, product.sku)).zfill(12)[-12:]
                s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits))
                product.barcode = digits + str((10 - (s % 10)) % 10)
            # price_retail is optional on the form (a multi-size product prices per size
            # instead) but the column itself isn't nullable — 0 here, backfilled below
            # once _save_product_sizes knows the sizes' prices.
            if product.price_retail is None:
                product.price_retail = Decimal('0')
            product.save()

            # Save sizes (from chip selection)
            sizes_str = request.POST.get('sizes_selected', '')
            size_ids = [int(x) for x in sizes_str.split(',') if x.strip().isdigit()] if sizes_str else []
            if size_ids:
                product.sizes.set(size_ids)

            _save_size_quantities(product, request, size_ids)
            _save_product_sizes(product, request)

            # Save images (base64)
            import json as _json
            images_raw = request.POST.get('images_b64', '[]')
            try:
                images_list = _json.loads(images_raw)
                for i, img_b64 in enumerate(images_list[:5]):
                    if img_b64 and img_b64.startswith('data:image'):
                        ProductImage.objects.create(product=product, image_data=img_b64, order=i)
            except Exception:
                pass

            messages.success(request, f'تم إضافة المنتج "{product.name}" بنجاح')
            return redirect('product_list')
    else:
        form = ProductForm()

    sys_settings_obj = SystemSetting.objects.first()
    market_type = sys_settings_obj.market_type if sys_settings_obj else 'clothes'
    return render(request, 'products/product_form.html', {
        'form': form,
        'title': 'إضافة منتج جديد',
        'next_sku': next_sku,
        'all_sizes_json': json.dumps(all_sizes),
        'preselected_sizes_json': '[]',
        'market_type': market_type,
        'global_sizes': all_sizes,
        'has_size_group': False,
        'size_rows': [],
    })

@login_required
@require_permission('products', 'view')
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    all_sizes = list(Size.objects.filter(is_active=True).values('id', 'name', 'sort_order'))

    if request.method == 'POST':
        old_retail = Decimal(str(product.price_retail))
        old_semi = Decimal(str(product.price_semi_wholesale))
        old_wholesale = Decimal(str(product.price_wholesale))
        old_cost = Decimal(str(product.cost_price))

        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            saved_product = form.save(commit=False)
            # Auto-generate barcode from SKU if missing
            if not saved_product.barcode and saved_product.sku:
                digits = ''.join(filter(str.isdigit, saved_product.sku)).zfill(12)[-12:]
                s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits))
                saved_product.barcode = digits + str((10 - (s % 10)) % 10)
            # price_retail is optional (a multi-size product prices per size instead) but
            # the column isn't nullable — 0 here, backfilled by _save_product_sizes below.
            if saved_product.price_retail is None:
                saved_product.price_retail = Decimal('0')
            saved_product.save()

            # Update sizes
            sizes_str = request.POST.get('sizes_selected', '')
            size_ids = [int(x) for x in sizes_str.split(',') if x.strip().isdigit()] if sizes_str else []
            if size_ids:
                saved_product.sizes.set(size_ids)
            else:
                saved_product.sizes.clear()

            _save_size_quantities(saved_product, request, size_ids)
            _save_product_sizes(saved_product, request)

            # Handle images
            import json as _json
            images_raw = request.POST.get('images_b64', '[]')
            try:
                images_list = _json.loads(images_raw)
                if images_list:
                    saved_product.images.all().delete()
                    for i, img_b64 in enumerate(images_list[:5]):
                        if img_b64 and img_b64.startswith('data:image'):
                            ProductImage.objects.create(product=saved_product, image_data=img_b64, order=i)
            except Exception:
                pass

            changes = []
            if old_retail != Decimal(str(saved_product.price_retail)):
                changes.append({'name': saved_product.name, 'type': 'قطاعي', 'old': old_retail, 'new': saved_product.price_retail})
            if old_semi != Decimal(str(saved_product.price_semi_wholesale)):
                changes.append({'name': saved_product.name, 'type': 'نصف جملة', 'old': old_semi, 'new': saved_product.price_semi_wholesale})
            if old_wholesale != Decimal(str(saved_product.price_wholesale)):
                changes.append({'name': saved_product.name, 'type': 'جملة', 'old': old_wholesale, 'new': saved_product.price_wholesale})
            if old_cost != Decimal(str(saved_product.cost_price)):
                changes.append({'name': saved_product.name, 'type': 'تكلفة', 'old': old_cost, 'new': saved_product.cost_price})

            if changes:
                notify_price_changes(request, changes, 'السعر / التكلفة (تعديل فردي)')

            messages.success(request, f'تم تعديل المنتج "{saved_product.name}"')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    sys_settings_obj = SystemSetting.objects.first()
    market_type = sys_settings_obj.market_type if sys_settings_obj else 'clothes'

    size_variants = list(product.variants.filter(color='').select_related('size').order_by('size__sort_order'))

    return render(request, 'products/product_form.html', {
        'form': form, 'title': 'تعديل منتج',
        'next_sku': product.sku,
        'all_sizes_json': json.dumps(all_sizes),
        'preselected_sizes_json': json.dumps(list(product.sizes.values_list('id', flat=True))),
        'market_type': market_type,
        'product': product,
        'global_sizes': all_sizes,
        'has_size_group': bool(size_variants),
        'size_rows': [{'size_id': v.size_id, 'price_override': v.price_override} for v in size_variants],
    })

@login_required
@require_permission('products', 'view')
def product_delete(request, pk):
    from django.db.models.deletion import ProtectedError

    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        try:
            product.delete()
        except ProtectedError:
            # Any product that's already appeared on a purchase invoice, sale, etc. is kept
            # around by PROTECT so those historical documents never lose their line items —
            # deleting it would silently corrupt past invoices/reports. Deactivating instead
            # keeps the history intact while removing it from POS/waiter (both check is_active).
            product.is_active = False
            product.save(update_fields=['is_active'])
            messages.warning(
                request,
                f'لا يمكن حذف "{product.name}" لأنه مرتبط بمستندات سابقة (فواتير شراء/بيع). '
                'تم إلغاء تفعيله بدلاً من ذلك — لن يظهر في الكاشير أو شاشة الويتر بعد الآن.'
            )
            return redirect('product_detail', pk=product.pk)
        return redirect('product_list')
    return render(request, 'products/product_confirm_delete.html', {'object': product, 'title': 'حذف منتج'})


# --- Raw materials (ingredients) — separate from "إضافة منتج" which is for sellable
# final products. A raw material is never rung up at POS; it only exists to be
# consumed via a Recipe (restaurant.Recipe/RecipeItem) when a finished menu item sells. ---

@login_required
@require_permission('products', 'view')
def raw_material_list(request):
    materials = Product.objects.filter(is_raw_material=True).order_by('name')
    return render(request, 'products/raw_material_list.html', {
        'materials': materials, 'title': 'المواد الخام',
    })


@login_required
@require_permission('products', 'edit')
def raw_material_create(request):
    if request.method == 'POST':
        form = RawMaterialForm(request.POST)
        if form.is_valid():
            material = form.save(commit=False)
            material.is_raw_material = True
            # price_retail/barcode are required columns on Product but meaningless for a
            # raw material (never sold directly) — default them instead of asking for them.
            material.price_retail = Decimal('0')
            # cost_price is optional here — usually set later from the purchase invoice
            # when the material is actually bought in.
            if material.cost_price is None:
                material.cost_price = Decimal('0')
            if not material.sku:
                material.sku = _next_sku(prefix='RM-')
            material.save()
            messages.success(request, f'تمت إضافة المادة الخام "{material.name}" بنجاح.')
            return redirect('raw_material_list')
    else:
        form = RawMaterialForm()

    return render(request, 'products/raw_material_form.html', {
        'form': form, 'title': 'إضافة مادة خام',
    })


@login_required
@require_permission('products', 'edit')
def raw_material_update(request, pk):
    material = get_object_or_404(Product, pk=pk, is_raw_material=True)
    if request.method == 'POST':
        original_cost_price = material.cost_price
        form = RawMaterialForm(request.POST, instance=material)
        if form.is_valid():
            updated = form.save(commit=False)
            # Left blank on edit → keep whatever cost the last purchase invoice set,
            # instead of wiping it out to 0.
            if updated.cost_price is None:
                updated.cost_price = original_cost_price
            updated.save()
            messages.success(request, 'تم حفظ التعديلات.')
            return redirect('raw_material_list')
    else:
        form = RawMaterialForm(instance=material)

    return render(request, 'products/raw_material_form.html', {
        'form': form, 'title': f'تعديل مادة خام: {material.name}', 'material': material,
    })


# --- Warehouses CRUD ---

@login_required
@require_granular_action('inventory', 'warehouses', 'products', 'view')
def warehouse_list(request):
    warehouses = Warehouse.objects.all()
    # Calculate total value/items per warehouse for display
    for wh in warehouses:
        wh.total_items = WarehouseStock.objects.filter(warehouse=wh).aggregate(Sum('quantity'))['quantity__sum'] or 0
    
    return render(request, 'products/warehouse_list.html', {'warehouses': warehouses, 'title': 'المخازن'})

@login_required
@require_permission('products', 'view')
def warehouse_detail(request, pk):
    from django.db.models import Count, Q, Sum

    warehouse = get_object_or_404(Warehouse, pk=pk)
    stocks = WarehouseStock.objects.filter(warehouse=warehouse).select_related('product').order_by('product__name')
    
    # Search functionality within warehouse detail
    search_query = request.GET.get('search', '')
    if search_query:
        stocks = stocks.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query)
        )
        
    transactions = StockTransaction.objects.filter(warehouse=warehouse).select_related('product').order_by('-created_at')[:50]
    
    top_products = Product.objects.filter(
        transactions__warehouse=warehouse
    ).annotate(
        trans_count=Count('transactions', filter=Q(transactions__warehouse=warehouse))
    ).order_by('-trans_count')[:5]

    top_sold_items = Product.objects.filter(
        orderitem__order__warehouse=warehouse
    ).annotate(
        sold_qty=Sum('orderitem__quantity')
    ).exclude(sold_qty=None).order_by('-sold_qty')[:5]

    return render(request, 'products/warehouse_detail.html', {
        'warehouse': warehouse, 
        'stocks': stocks,
        'transactions': transactions,
        'top_products': top_products,
        'top_sold_items': top_sold_items,
        'title': f'تفاصيل مخزن: {warehouse.name}'
    })

@login_required
@require_granular_action('inventory', 'warehouses', 'products', 'view')
def warehouse_create(request):
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('warehouse_list')
    else:
        form = WarehouseForm()
    return render(request, 'products/warehouse_form.html', {'form': form, 'title': 'إضافة مخزن جديد'})

@login_required
@require_permission('products', 'view')
def warehouse_update(request, pk):
    wh = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=wh)
        if form.is_valid():
            form.save()
            return redirect('warehouse_list')
    else:
        form = WarehouseForm(instance=wh)
    return render(request, 'products/warehouse_form.html', {'form': form, 'title': 'تعديل مخزن'})

@login_required
@require_permission('products', 'view')
def warehouse_delete(request, pk):
    wh = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        wh.delete()
        return redirect('warehouse_list')
    return render(request, 'products/product_confirm_delete.html', {'object': wh, 'title': 'حذف مخزن'})

# --- Stock Transfer (Single - Old View) ---

@login_required
@require_granular_action('inventory', 'transfer', 'products', 'view')
def stock_transfer(request):
    if request.method == 'POST':
        form = StockTransferForm(request.POST)
        if form.is_valid():
            from_wh = form.cleaned_data['from_warehouse']
            to_wh = form.cleaned_data['to_warehouse']
            product = form.cleaned_data['product']
            qty = to_decimal(form.cleaned_data['quantity'])
            note = form.cleaned_data['note']

            with db_transaction.atomic():
                src_stock = WarehouseStock.objects.select_for_update().get(warehouse=from_wh, product=product)
                src_stock.quantity = to_decimal(src_stock.quantity)
                src_stock.quantity -= qty
                src_stock.save()

                dest_stock, created = WarehouseStock.objects.select_for_update().get_or_create(
                    warehouse=to_wh,
                    product=product,
                    defaults={'quantity': Decimal('0')}
                )
                dest_stock.quantity = to_decimal(dest_stock.quantity)
                dest_stock.quantity += qty
                dest_stock.save()

                StockTransaction.objects.create(
                    product=product, warehouse=from_wh, transaction_type='TRN', 
                    quantity=qty, note=f"تحويل صادر إلى {to_wh.name} - {note}"
                )
                StockTransaction.objects.create(
                    product=product, warehouse=to_wh, transaction_type='TRN', 
                    quantity=qty, note=f"تحويل وارد من {from_wh.name} - {note}"
                )
            
            return redirect('warehouse_list')
    else:
        form = StockTransferForm()
    return render(request, 'products/stock_transfer.html', {'form': form, 'title': 'تحويل مخزون'})

# --- NEW: Bulk Stock Transfer ---

@login_required
@require_permission('products', 'view')
def bulk_transfer_view(request):
    warehouses = Warehouse.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True).select_related('category')
    return render(request, 'products/bulk_stock_transfer.html', {
        'warehouses': warehouses,
        'products': products,
        'title': 'تحويل مخزون مجمع'
    })

@login_required
@require_permission('products', 'view')
@require_POST
def bulk_transfer_save_api(request):
    try:
        data = json.loads(request.body)
        from_wh_id = data.get('from_warehouse_id')
        to_wh_id = data.get('to_warehouse_id')
        items = data.get('items', [])
        note_global = data.get('note', '')

        if not from_wh_id or not to_wh_id:
            return JsonResponse({'success': False, 'error': 'يجب تحديد المخزن المحول منه والمحول إليه'})
        
        if from_wh_id == to_wh_id:
            return JsonResponse({'success': False, 'error': 'لا يمكن التحويل لنفس المخزن'})

        if not items:
            return JsonResponse({'success': False, 'error': 'لا توجد أصناف للتحويل'})

        with db_transaction.atomic():
            from_wh = Warehouse.objects.get(id=from_wh_id)
            to_wh = Warehouse.objects.get(id=to_wh_id)

            for item in items:
                product_id = item.get('product_id')
                qty = to_decimal(item.get('quantity', 0))
                
                if not product_id:
                    raise ValueError('بيانات صنف غير مكتملة')
                if qty <= 0:
                    raise ValueError('الكمية يجب أن تكون أكبر من صفر')

                product = Product.objects.get(id=product_id)
                
                # Check Stock
                src_stock = WarehouseStock.objects.filter(warehouse=from_wh, product=product).select_for_update().first()
                current_qty = to_decimal(src_stock.quantity if src_stock else Decimal('0'))

                if current_qty < qty:
                    raise ValueError(f"الرصيد غير كافي للصنف {product.name}. المتاح: {current_qty}")

                # Update Stocks
                if src_stock:
                    src_stock.quantity = to_decimal(src_stock.quantity)
                    src_stock.quantity -= qty
                    src_stock.save()

                dest_stock, _ = WarehouseStock.objects.select_for_update().get_or_create(
                    warehouse=to_wh,
                    product=product,
                    defaults={'quantity': Decimal('0')}
                )
                dest_stock.quantity = to_decimal(dest_stock.quantity)
                dest_stock.quantity += qty
                dest_stock.save()

                # Create Transactions
                StockTransaction.objects.create(
                    product=product, warehouse=from_wh, transaction_type='TRN', 
                    quantity=qty, note=f"تحويل صادر إلى {to_wh.name} - {note_global}",
                    unit_price=product.cost_price 
                )
                StockTransaction.objects.create(
                    product=product, warehouse=to_wh, transaction_type='TRN', 
                    quantity=qty, note=f"تحويل وارد من {from_wh.name} - {note_global}",
                    unit_price=product.cost_price
                )

        return JsonResponse({'success': True, 'message': 'تم التحويل بنجاح'})

    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f"خطأ غير متوقع: {str(e)}"})

@login_required
@require_granular_action('pos', 'transfer', 'products', 'view')
@require_POST
def pos_bulk_transfer_api(request):
    """
    Handles bulk transfers from POS with per-item source and destination warehouses.
    Expects JSON: { items: [{product_id, from_wh_id, to_wh_id, quantity, note}, ...] }
    """
    try:
        data = json.loads(request.body)
        items = data.get('items', [])

        if not items:
            return JsonResponse({'success': False, 'error': 'لا توجد بيانات للتحويل'})

        with db_transaction.atomic():
            for entry in items:
                try:
                    product_id = entry.get('product_id')
                    from_wh_id = entry.get('from_warehouse_id')
                    to_wh_id = entry.get('to_warehouse_id')
                    qty = to_decimal(entry.get('quantity', 0))
                    note = entry.get('note', '')

                    if not product_id or not from_wh_id or not to_wh_id:
                        raise ValueError("بيانات التحويل غير مكتملة في أحد الأسطر")
                    
                    if from_wh_id == to_wh_id:
                        raise ValueError(f"لا يمكن التحويل لنفس المخزن في أحد الأسطر")

                    if qty <= 0:
                        raise ValueError("الكمية يجب أن تكون أكبر من صفر")

                    product = Product.objects.get(id=product_id)
                    from_wh = Warehouse.objects.get(id=from_wh_id)
                    to_wh = Warehouse.objects.get(id=to_wh_id)

                    # Check Stock
                    src_stock = WarehouseStock.objects.filter(warehouse=from_wh, product=product).select_for_update().first()
                    current_qty = to_decimal(src_stock.quantity if src_stock else Decimal('0'))

                    if current_qty < qty:
                        raise ValueError(f"الرصيد غير كافي للصنف {product.name} في مخزن {from_wh.name}")

                    # Update Stocks
                    src_stock.quantity = to_decimal(src_stock.quantity)
                    src_stock.quantity -= qty
                    src_stock.save()

                    dest_stock, _ = WarehouseStock.objects.select_for_update().get_or_create(
                        warehouse=to_wh,
                        product=product,
                        defaults={'quantity': Decimal('0')}
                    )
                    dest_stock.quantity = to_decimal(dest_stock.quantity)
                    dest_stock.quantity += qty
                    dest_stock.save()

                    # Create Transactions
                    StockTransaction.objects.create(
                        product=product, warehouse=from_wh, transaction_type='TRN', 
                        quantity=qty, note=f"تحويل صادر إلى {to_wh.name} (من POS) - {note}",
                        unit_price=product.cost_price 
                    )
                    StockTransaction.objects.create(
                        product=product, warehouse=to_wh, transaction_type='TRN', 
                        quantity=qty, note=f"تحويل وارد من {from_wh.name} (من POS) - {note}",
                        unit_price=product.cost_price
                    )
                except (Product.DoesNotExist, Warehouse.DoesNotExist):
                    raise ValueError(f"أحد المنتجات أو المخازن غير موجود")

        return JsonResponse({'success': True, 'message': 'تم تنفيذ جميع عمليات التحويل بنجاح'})

    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f"خطأ غير متوقع: {str(e)}"})

@login_required
@require_permission('products', 'view')
@require_POST
def pos_stock_movement_api(request):
    """
    Handles single-item stock movement (IN/OUT) from POS.
    Expects JSON: { product_id, warehouse_id, transaction_type, quantity, note }
    """
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        warehouse_id = data.get('warehouse_id')
        transaction_type = data.get('transaction_type') # 'IN' or 'OUT'
        qty = Decimal(str(data.get('quantity', 0)))
        note = data.get('note', 'حركة سريعة من الموبايل')

        if not product_id or not warehouse_id or not transaction_type or qty <= 0:
            return JsonResponse({'success': False, 'error': 'بيانات غير مكتملة'})

        if transaction_type not in ['IN', 'OUT']:
            return JsonResponse({'success': False, 'error': 'نوع حركة غير صالح'})

        product = Product.objects.get(id=product_id)
        warehouse = Warehouse.objects.get(id=warehouse_id)

        with db_transaction.atomic():
            ws, created = WarehouseStock.objects.get_or_create(warehouse=warehouse, product=product)
            
            if transaction_type == 'OUT' and ws.quantity < qty:
                return JsonResponse({'success': False, 'error': f"الرصيد غير كافي. المتوفر: {ws.quantity}"})

            if transaction_type == 'IN':
                ws.quantity += qty
            else:
                ws.quantity -= qty
            
            ws.save()

            StockTransaction.objects.create(
                product=product,
                warehouse=warehouse,
                transaction_type=transaction_type,
                quantity=qty,
                note=note,
                unit_price=product.cost_price
            )

        return JsonResponse({'success': True, 'message': 'تم تسجيل الحركة بنجاح'})

    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'المنتج غير موجود'})
    except Warehouse.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'المخزن غير موجود'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# --- Suppliers CRUD ---
@login_required
@require_granular_action('inventory', 'suppliers', 'products', 'view')
def supplier_list(request):
    from django.db.models import Q
    from decimal import Decimal
    suppliers = Supplier.objects.all()
    
    # 1. Search Filter
    search = request.GET.get('search', '').strip()
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(contact_name__icontains=search)
        )
        
    # 2. Type Filter
    sup_type = request.GET.get('type', '').strip()
    if sup_type:
        suppliers = suppliers.filter(supplier_type=sup_type)
        
    # 3. Calculate Summary Stats (Overall stats for all active suppliers)
    total_suppliers = Supplier.objects.count()
    active_suppliers = Supplier.objects.filter(is_active=True).count()
    
    total_outstanding = Decimal('0.00')
    for s in Supplier.objects.all():
        bal = s.outstanding_balance
        if bal > 0:
            total_outstanding += bal

    context = {
        'suppliers': suppliers,
        'title': 'الموردين',
        'total_suppliers': total_suppliers,
        'active_suppliers': active_suppliers,
        'total_outstanding': total_outstanding,
    }
    return render(request, 'products/supplier_list.html', context)

@login_required
@require_granular_action('inventory', 'suppliers', 'products', 'view')
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'products/supplier_form.html', {'form': form, 'title': 'إضافة مورد'})

@login_required
@require_permission('products', 'view')
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'products/supplier_form.html', {'form': form, 'title': 'تعديل مورد'})

@login_required
@require_permission('products', 'view')
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        return redirect('supplier_list')
    return render(request, 'products/product_confirm_delete.html', {'object': supplier, 'title': 'حذف مورد'})

# --- Categories CRUD ---
@login_required
@require_granular_action('master_data', 'categories', 'products', 'view')
def category_list(request):
    categories = Category.objects.all().order_by('-created_at')
    return render(request, 'products/category_list.html', {'categories': categories, 'title': 'الأقسام'})

@login_required
@require_permission('products', 'view')
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'products/category_form.html', {'form': form, 'title': 'إضافة قسم جديد'})

@login_required
@require_permission('products', 'view')
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'products/category_form.html', {'form': form, 'title': 'تعديل قسم'})

@login_required
@require_permission('products', 'view')
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'products/product_confirm_delete.html', {'object': category, 'title': 'حذف قسم'})

# --- Stock Transactions CRUD ---

@login_required
@require_granular_action('inventory', 'transactions', 'products', 'view')
def transaction_list(request):
    queryset = StockTransaction.objects.select_related('product', 'warehouse').all().order_by('-created_at')

    search_query = request.GET.get('search', '')
    trans_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    warehouse_id = request.GET.get('warehouse', '')

    if search_query:
        queryset = queryset.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(note__icontains=search_query)
        )
    
    if trans_type:
        queryset = queryset.filter(transaction_type=trans_type)

    if warehouse_id:
        queryset = queryset.filter(warehouse_id=warehouse_id)
        
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    warehouses = Warehouse.objects.all()

    # Resolve the name of the selected warehouse for the print header
    selected_warehouse_name = ''
    if warehouse_id:
        wh_obj = Warehouse.objects.filter(id=warehouse_id).first()
        if wh_obj:
            selected_warehouse_name = wh_obj.name

    return render(request, 'products/transaction_list.html', {
        'transactions': page_obj,
        'title': 'سجل حركات المخزن',
        'warehouses': warehouses,
        'selected_warehouse_name': selected_warehouse_name,
    })

@login_required
@require_granular_action('inventory', 'transactions', 'products', 'view')
def transaction_create(request):
    if request.method == 'POST':
        form = StockTransactionForm(request.POST)
        if form.is_valid():
            with db_transaction.atomic():
                trans = form.save(commit=False)

                # Normalize quantity to Decimal to avoid float/Decimal mixing
                qty = to_decimal(trans.quantity, default=0)

                if qty <= 0:
                    form.add_error('quantity', 'يجب أن تكون الكمية أكبر من صفر')
                elif not trans.warehouse:
                    form.add_error('warehouse', 'يجب اختيار المخزن')
                else:
                    stock, created = WarehouseStock.objects.select_for_update().get_or_create(
                        warehouse=trans.warehouse,
                        product=trans.product,
                        defaults={'quantity': Decimal('0.00')},
                    )

                    stock_qty = to_decimal(stock.quantity, default=0)

                    if trans.transaction_type in ['IN', 'RET']:
                        stock_qty += qty
                    elif trans.transaction_type == 'OUT':
                        if stock_qty < qty:
                            form.add_error(
                                'quantity',
                                f'الكمية المطلوبة ({qty}) أكبر من المتاح في المخزن ({stock_qty})'
                            )
                        else:
                            stock_qty -= qty

                    if not form.errors:
                        stock.quantity = stock_qty
                        trans.quantity = qty
                        stock.save()
                        trans.save()

                        # Use a fresh product instance to ensure calculate_total_stock uses up-to-date DB values
                        prod = Product.objects.get(pk=trans.product.pk)
                        prod.calculate_total_stock()

                        messages.success(request, 'تم تسجيل الحركة المخزنية بنجاح.')
                        return redirect('transaction_list')
    else:
        form = StockTransactionForm()

    products_data = list(Product.objects.filter(is_active=True).values('id', 'name', 'sku', 'stock_quantity'))
    avail_warehouses = Warehouse.objects.filter(is_active=True)

    return render(request, 'products/transaction_form.html', {
        'form': form,
        'title': 'تسجيل حركة مخزنية',
        'products_data': products_data,
        'avail_warehouses': avail_warehouses,
    })

# --- Bulk Audit System ---

@login_required
@require_permission('products', 'view')
def bulk_audit_view(request):
    # Fetch all active products
    products = Product.objects.filter(is_active=True).order_by('category', 'name')
    warehouses = Warehouse.objects.filter(is_active=True)
    return render(request, 'products/bulk_audit.html', {'products': products, 'warehouses': warehouses})

@login_required
@require_permission('products', 'view')
def get_warehouse_stocks_ajax(request, pk):
    """API to fetch current system stock for all products in a specific warehouse"""
    try:
        # Check if warehouse exists to avoid 500 error if ID is bad
        if not Warehouse.objects.filter(pk=pk).exists():
             return JsonResponse({'success': False, 'error': 'المخزن غير موجود'})

        stocks = WarehouseStock.objects.filter(warehouse_id=pk).values('product_id', 'quantity')
        # Convert to dictionary for fast lookup: {product_id: quantity}
        stock_map = {item['product_id']: float(item['quantity']) for item in stocks}
        return JsonResponse({'success': True, 'stocks': stock_map})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_permission('products', 'view')
@require_POST
def update_product_stock_ajax(request):
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        warehouse_id = data.get('warehouse_id')
        actual_qty_str = data.get('actual_quantity')
        note_text = data.get('note', '')

        if not product_id or not warehouse_id or actual_qty_str is None:
            return JsonResponse({'success': False, 'error': 'Missing data'})

        product = Product.objects.get(id=product_id)
        warehouse = Warehouse.objects.get(id=warehouse_id)
        
        wh_stock, created = WarehouseStock.objects.get_or_create(warehouse=warehouse, product=product)
        system_qty_float = wh_stock.quantity if wh_stock.quantity is not None else 0
        system_qty = Decimal(str(system_qty_float))
        try:
            actual_qty = Decimal(str(actual_qty_str))
        except (ValueError, TypeError, InvalidOperation):
            return JsonResponse({'success': False, 'error': 'Invalid quantity format'})
        
        diff = actual_qty - system_qty

        if abs(diff) > Decimal('0.001'):
            abs_diff = abs(diff)

            StockTransaction.objects.create(
                product=product,
                warehouse=warehouse,
                transaction_type='ADJ',
                quantity=abs_diff,
                unit_price=product.cost_price, 
                note=f"جرد فوري ({warehouse.name}): {note_text}" if note_text else f"جرد فوري ({warehouse.name}) - تسوية تلقائية"
            )

            wh_stock.quantity = actual_qty
            wh_stock.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'تم التحديث', 
                'new_stock': str(actual_qty),
                'diff': diff
            })
        else:
            return JsonResponse({'success': True, 'message': 'لا يوجد تغيير في الكمية'})

    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'المنتج غير موجود'})
    except Warehouse.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'المخزن غير موجود'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'قيمة رقمية غير صحيحة'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# --- NEW: Bulk Product Add System ---

@login_required
@require_permission('products', 'view')
def bulk_product_add_view(request):
    """View to render the bulk add page with necessary context"""
    categories = Category.objects.all()
    kinds = Kind.objects.select_related('category').order_by('category__name', 'name')
    sizes = Size.objects.filter(is_active=True).order_by('sort_order', 'name')
    suppliers = Supplier.objects.all()
    warehouses = Warehouse.objects.filter(is_active=True)
    sys_settings = SystemSetting.objects.first()
    
    # Calculate suggested start SKU
    last_product = Product.objects.order_by('id').last()
    start_sku = "10001"
    if last_product and last_product.sku and last_product.sku.isdigit():
        try:
            start_sku = str(int(last_product.sku) + 1)
        except ValueError:
            pass

    context = {
        'categories': categories,
        'kinds': kinds,
        'sizes': sizes,
        'suppliers': suppliers,
        'warehouses': warehouses,
        'start_sku': start_sku,
        'title': 'إضافة منتجات متعددة (Bulk Add)',
        'sys_settings': sys_settings
    }
    return render(request, 'products/bulk_product_add.html', context)

def _compress_uploaded_image_to_base64(uploaded_file, max_width=1280, quality=70):
    """
    Compress uploaded image server-side and return data:image/jpeg;base64,...
    """
    with Image.open(uploaded_file) as img:
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        width, height = img.size
        if width > max_width:
            new_height = int((max_width / float(width)) * height)
            img = img.resize((max_width, new_height), Image.LANCZOS)

        out = BytesIO()
        img.save(out, format='JPEG', quality=quality, optimize=True)
        encoded = base64.b64encode(out.getvalue()).decode('utf-8')
        return f'data:image/jpeg;base64,{encoded}'

@login_required
@require_permission('products', 'view')
@require_POST
def bulk_product_save_ajax(request):
    """API Endpoint to save a single product from the bulk table"""
    try:
        data = request.POST.dict() if request.POST else json.loads(request.body)
        
        # 1. Extract Data
        sku = (data.get('sku') or '').strip()
        name = (data.get('name') or '').strip()
        barcode = (data.get('barcode') or '').strip()
        category_id = data.get('category')
        kind_id = data.get('kind')
        supplier_id = data.get('supplier')
        material = (data.get('material') or '').strip()
        pattern = (data.get('pattern') or '').strip()
        color = (data.get('color') or '').strip()
        scientific_name = (data.get('scientific_name') or '').strip()
        packaging_type = (data.get('packaging_type') or '').strip()
        strips_per_box = int(data.get('strips_per_box') or 1)
        unit_measure = (data.get('unit_measure') or 'PCS').strip()
        sizes_selected = (data.get('sizes_selected') or '').strip()

        cost_price = to_decimal(data.get('cost_price', 0), 0)
        retail_price = to_decimal(data.get('price_retail', data.get('sell_price', 0)), 0)
        # Left at 0 (not sold at these tiers) if the row doesn't specify them — do not
        # auto-copy retail_price; price_retail is the safe fallback at the point of sale
        # (see products.pricing.tier_price) if a wholesale-tier sale is ever attempted.
        semi_wholesale_price = to_decimal(data.get('price_semi_wholesale', 0), 0)
        wholesale_price = to_decimal(data.get('price_wholesale', 0), 0)

        quantity = to_decimal(data.get('quantity', 0), 0)
        low_stock_threshold = to_decimal(data.get('low_stock_threshold', 10), 10)
        try:
            pieces_per_package = int(data.get('pieces_per_package', 48) or 48)
        except (TypeError, ValueError):
            pieces_per_package = 48
        is_active = str(data.get('is_active', '1')).strip().lower() in ('1', 'true', 'yes', 'on', 'نعم')
        warehouse_id = data.get('warehouse_id')

        # 2. Validation
        if not sku:
            return JsonResponse({'success': False, 'error': 'كود المنتج (SKU) مطلوب'})
        if Product.objects.filter(sku=sku).exists():
            return JsonResponse({'success': False, 'error': f'كود المنتج {sku} موجود مسبقاً!'})
        
        if not name:
             return JsonResponse({'success': False, 'error': 'اسم المنتج مطلوب'})

        if not barcode and sku:
            barcode = _generate_ean13_from_sku(sku)

        # 3. Create Product
        product = Product(
            name=name,
            sku=sku,
            barcode=barcode,
            material=material,
            pattern=pattern,
            color=color,
            scientific_name=scientific_name,
            packaging_type=packaging_type,
            strips_per_box=strips_per_box,
            cost_price=cost_price,
            price_retail=retail_price,
            price_semi_wholesale=semi_wholesale_price,
            price_wholesale=wholesale_price,
            stock_quantity=0,               # Will be updated by signal/logic later
            unit_measure=unit_measure if unit_measure in dict(Product.UNIT_CHOICES) else 'PCS',
            low_stock_threshold=low_stock_threshold,
            pieces_per_package=max(1, pieces_per_package),
            is_active=is_active
        )
        
        if category_id:
            product.category_id = category_id
        if kind_id:
            product.kind_id = kind_id
        if supplier_id:
            product.supplier_id = supplier_id
            
        product.save()

        # 3.1 Handle sizes many-to-many
        if sizes_selected:
            size_ids = [int(x) for x in sizes_selected.split(',') if x.strip().isdigit()]
            if size_ids:
                product.sizes.set(size_ids)

        # 3.2 Handle image uploads (up to 5 images), compress on backend
        uploaded_images = request.FILES.getlist('images')
        for idx, img_file in enumerate(uploaded_images[:5]):
            try:
                img_b64 = _compress_uploaded_image_to_base64(img_file)
                ProductImage.objects.create(
                    product=product,
                    image_data=img_b64,
                    order=idx
                )
            except Exception:
                continue

        # 4. Handle Initial Stock (if quantity > 0 and warehouse selected)
        if quantity > 0 and warehouse_id:
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
                StockTransaction.objects.create(
                    product=product,
                    warehouse=warehouse,
                    transaction_type='IN',
                    quantity=quantity,
                    unit_price=cost_price,
                    note="رصيد افتتاحي (إضافة مجمعة)"
                )
                
                # Update Warehouse Stock manually to be safe immediately
                stock, created = WarehouseStock.objects.get_or_create(warehouse=warehouse, product=product)
                stock.quantity += quantity
                stock.save()
                
            except Warehouse.DoesNotExist:
                pass # Should not happen if UI is correct

        return JsonResponse({'success': True, 'message': 'تم الحفظ بنجاح', 'id': product.id})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_permission('products', 'create')
@require_POST
def bulk_quick_add_category_api(request):
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'اسم القسم مطلوب'})
    obj, created = Category.objects.get_or_create(name=name, defaults={'is_active': True})
    return JsonResponse({'success': True, 'id': obj.id, 'name': obj.name, 'created': created})

@login_required
@require_permission('products', 'create')
@require_POST
def bulk_quick_add_kind_api(request):
    name = (request.POST.get('name') or '').strip()
    category_id = request.POST.get('category_id')
    if not name or not category_id:
        return JsonResponse({'success': False, 'error': 'اسم النوع والقسم مطلوبان'})
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'القسم غير موجود'})
    obj, created = Kind.objects.get_or_create(category=category, name=name, defaults={'is_active': True})
    return JsonResponse({
        'success': True, 'id': obj.id, 'name': obj.name,
        'category_id': category.id, 'created': created
    })

@login_required
@require_permission('products', 'create')
@require_POST
def bulk_quick_add_size_api(request):
    name = (request.POST.get('name') or '').strip()
    size_type = (request.POST.get('size_type') or 'custom').strip()
    sort_order = request.POST.get('sort_order')
    if not name:
        return JsonResponse({'success': False, 'error': 'اسم المقاس مطلوب'})
    try:
        sort_order = int(sort_order) if sort_order not in (None, '') else 0
    except ValueError:
        sort_order = 0
    if size_type not in dict(Size.SIZE_TYPES):
        size_type = 'custom'
    obj, created = Size.objects.get_or_create(
        name=name,
        defaults={'size_type': size_type, 'sort_order': sort_order, 'is_active': True}
    )
    return JsonResponse({'success': True, 'id': obj.id, 'name': obj.name, 'created': created})

# --- Supplier Financials ---
@login_required
def api_products_search(request):
    """
    JSON API for product search (used by Purchase Invoice / Purchase Return item pickers).
    """
    query = request.GET.get('q', '').strip()

    # A supplier invoice buys raw materials (or other stocked retail goods) — never a
    # kitchen-routed menu item, which is a prepared dish, not something you receive from
    # a supplier. Recipe ingredients are the thing purchase invoices should restock.
    #
    # Raw materials are always exempt from that exclusion though: they have no category
    # field on their own add form, so they fall back to the generic "بدون قسم" category
    # (is_menu_category=True, used to keep an uncategorized item from being wrongly
    # stock-gated in the POS/waiter grids) — which would otherwise make every raw
    # material invisible here too, even though they're exactly what a purchase invoice
    # is meant to restock.
    products = Product.objects.filter(is_active=True).filter(
        Q(is_raw_material=True) | ~Q(category__is_menu_category=True)
    )
    if query:
        products = products.filter(
            Q(name__icontains=query) | 
            Q(sku__icontains=query) |
            Q(barcode__icontains=query)
        )
    
    products = products.prefetch_related('sizes')[:20]

    results = []
    for p in products:
        results.append({
            'id': p.id,
            'name': p.name,
            'sku': p.sku,
            'cost_price': float(p.cost_price),
            'stock': float(p.stock_quantity),
            'unit': p.get_unit_measure_display(),
            'sizes': [{'id': s.id, 'name': s.name} for s in p.sizes.all()],
        })

    return JsonResponse(results, safe=False)

@login_required
@require_granular_action('inventory', 'purchase_invoices', 'products', 'view')
def purchase_invoice_create(request, pk=None):
    """
    Advanced Purchase Invoice creation and editing page (Purchase POS).
    """
    from django.contrib import messages
    from django.shortcuts import redirect
    import json
    
    invoice = None
    invoice_items_json = '[]'
    po = None
    prefill_supplier_id = None
    prefill_warehouse_id = None

    if pk:
        invoice = get_object_or_404(PurchaseInvoice, id=pk)
        if invoice.status != 'DRAFT':
            messages.error(request, "لا يمكن تعديل الفاتورة المؤكدة أو الملغية.")
            return redirect('purchase_invoice_detail', pk=pk)

        # Serialize existing items
        items_list = []
        for item in invoice.items.select_related('product').prefetch_related('product__sizes').all():
            items_list.append({
                'product_id': item.product.id,
                'name': item.product.name,
                'sku': item.product.sku,
                'quantity': float(item.quantity),
                'bonus_quantity': float(item.bonus_quantity),
                'unit_price': float(item.unit_price),
                'discount': float(item.discount),
                'tax_rate': float(item.tax_rate),
                'batch_number': item.batch_number or '',
                'expiry_date': item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else '',
                'sizes': [{'id': s.id, 'name': s.name} for s in item.product.sizes.all()],
                'size_breakdown': item.size_breakdown or {},
            })
        invoice_items_json = json.dumps(items_list)
    elif request.GET.get('po'):
        # Phase 7.1: convert a Purchase Order into a receiving invoice — prefill the lines
        # with each PO item's REMAINING quantity (ordered − already received). The user can
        # adjust for over/under-receipt before confirming; submit links back and updates the PO.
        po = get_object_or_404(PurchaseOrder, id=request.GET.get('po'))
        prefill_supplier_id = po.supplier_id
        prefill_warehouse_id = po.destination_warehouse_id
        items_list = []
        for poi in po.items.select_related('product'):
            remaining = poi.ordered_quantity - poi.received_quantity
            if remaining <= 0:
                continue
            items_list.append({
                'product_id': poi.product.id,
                'name': poi.product.name,
                'sku': poi.product.sku,
                'quantity': float(remaining),
                'bonus_quantity': 0,
                'unit_price': float(poi.unit_price),
                'discount': float(poi.discount),
                'tax_rate': 0,
                'batch_number': '',
                'expiry_date': '',
            })
        invoice_items_json = json.dumps(items_list)

    suppliers = Supplier.objects.filter(is_active=True)
    if invoice and invoice.supplier not in suppliers:
        suppliers = list(suppliers) + [invoice.supplier]

    warehouses = Warehouse.objects.filter(is_active=True)
    if invoice and invoice.warehouse not in warehouses:
        warehouses = list(warehouses) + [invoice.warehouse]

    try:
        from financial.models import Account
        # Only real cash/bank/wallet accounts — not the nominal chart-of-accounts rows
        # (AR/AP/COGS/Revenue/...), which exist purely for the internal double-entry
        # journal and would silently corrupt those balances if hand-picked here.
        accounts = Account.exclude_dead_duplicates(Account.objects.filter(
            is_active=True,
            account_type__in=['CASH_DRAWER', 'SAFE', 'BANK', 'VODAFONE_CASH', 'INSTAPAY'],
        ))
    except ImportError:
        accounts = []
        
    sys_settings = SystemSetting.objects.first()
        
    return render(request, 'products/purchase_invoice_create.html', {
        'suppliers': suppliers,
        'warehouses': warehouses,
        'accounts': accounts,
        'invoice': invoice,
        'invoice_items_json': invoice_items_json,
        'po': po,
        'po_id': po.id if po else None,
        'prefill_supplier_id': prefill_supplier_id,
        'prefill_warehouse_id': prefill_warehouse_id,
        'title': f'تعديل فاتورة مشتريات #{invoice.id}' if invoice else (f'استلام أمر شراء {po.po_number}' if po else 'فاتورة مشتريات جديدة'),
        'sys_settings': sys_settings
    })

@csrf_exempt
@login_required
@require_POST
def api_purchase_invoice_submit(request):
    """
    AJAX API to submit a complete purchase invoice.
    Handles: Items, Stock Batches, Weighted Average Cost, Supplier Balance, and Financial Transactions.
    """
    try:
        data = json.loads(request.body)
        invoice_id = data.get('invoice_id')
        supplier_id = data.get('supplier_id')
        warehouse_id = data.get('warehouse_id')
        account_id = data.get('account_id')
        payment_method = data.get('payment_method', 'cash')
        items = data.get('items', [])
        discount = Decimal(str(data.get('discount', 0)))
        landed_cost = Decimal(str(data.get('landed_cost', 0)))  # Phase 7.2
        paid_amount = Decimal(str(data.get('paid_amount', 0)))
        notes = data.get('notes', '')
        invoice_number = data.get('invoice_number', '')

        po_id = data.get('po_id')  # Phase 7.1: receiving against a Purchase Order

        invoice_status = data.get('status', 'CONFIRMED')
        if invoice_status not in ['DRAFT', 'CONFIRMED']:
            invoice_status = 'CONFIRMED'

        if not items:
            return JsonResponse({'status': 'error', 'message': 'لا يمكن حفظ فاتورة فارغة'}, status=400)

        supplier = get_object_or_404(Supplier, id=supplier_id)
        warehouse = get_object_or_404(Warehouse, id=warehouse_id)
        
        with db_transaction.atomic():
            # 1. Create/Fetch Invoice Header
            total_amount = Decimal('0.00')
            for item in items:
                qty = Decimal(str(item['quantity']))
                price = Decimal(str(item['unit_price']))
                disc = Decimal(str(item.get('discount', 0)))
                tax_r = Decimal(str(item.get('tax_rate', 0)))
                taxable = max(Decimal('0.00'), (qty * price) - disc)
                tax_amt = taxable * (tax_r / Decimal('100.00'))
                total_amount += taxable + tax_amt
            
            if invoice_id:
                invoice = get_object_or_404(PurchaseInvoice, id=invoice_id)
                if invoice.status != 'DRAFT':
                    return JsonResponse({'status': 'error', 'message': 'لا يمكن تعديل فاتورة غير مسودة'}, status=400)
                
                # Check for duplicate paper invoice number excluding this invoice itself
                if invoice_number and PurchaseInvoice.objects.filter(invoice_number=invoice_number, supplier=supplier).exclude(id=invoice_id).exists():
                    return JsonResponse({'status': 'error', 'message': f'رقم الفاتورة {invoice_number} مسجل مسبقاً لهذا المورد'}, status=400)
                
                # Clear existing items for draft update
                invoice.items.all().delete()
                
                invoice.supplier = supplier
                invoice.warehouse = warehouse
                invoice.invoice_number = invoice_number
                invoice.total_amount = total_amount
                invoice.discount = discount
                invoice.landed_cost = landed_cost
                invoice.payment_method = payment_method
                invoice.paid_amount = paid_amount
                invoice.notes = notes
                invoice.status = invoice_status
                invoice.save()
            else:
                if invoice_number and PurchaseInvoice.objects.filter(invoice_number=invoice_number, supplier=supplier).exists():
                     return JsonResponse({'status': 'error', 'message': f'رقم الفاتورة {invoice_number} مسجل مسبقاً لهذا المورد'}, status=400)

                invoice = PurchaseInvoice.objects.create(
                    supplier=supplier,
                    user=request.user,
                    warehouse=warehouse,
                    invoice_number=invoice_number,
                    total_amount=total_amount,
                    discount=discount,
                    landed_cost=landed_cost,
                    payment_method=payment_method,
                    paid_amount=paid_amount,
                    notes=notes,
                    status=invoice_status,
                    is_stock_applied=False # will be set to True if confirmed below
                )
            
            if paid_amount > 0 and account_id:
                from financial.models import Account
                invoice.account = get_object_or_404(Account, id=account_id)
                invoice.save(update_fields=['account'])

            # 2. Process Items
            for item_data in items:
                product = get_object_or_404(Product, id=item_data['product_id'])
                qty = Decimal(str(item_data['quantity']))
                bonus_qty = Decimal(str(item_data.get('bonus_quantity', 0)))
                unit_price = Decimal(str(item_data['unit_price']))
                batch_no = item_data.get('batch_number', '')
                exp_date = item_data.get('expiry_date') or None
                
                if qty <= 0:
                    raise ValueError(f"الكمية للمنتج {product.name} يجب أن تكون أكبر من صفر")

                size_breakdown = item_data.get('size_breakdown') or {}

                PurchaseInvoiceItem.objects.create(
                    invoice=invoice,
                    product=product,
                    quantity=qty,
                    bonus_quantity=bonus_qty,
                    unit_price=unit_price,
                    batch_number=batch_no,
                    expiry_date=exp_date,
                    discount=Decimal(str(item_data.get('discount', 0))),
                    tax_rate=Decimal(str(item_data.get('tax_rate', 0))),
                    size_breakdown=size_breakdown,
                )

            # 3. If CONFIRMED, apply stock & handle financials
            if invoice_status == 'CONFIRMED':
                from products.inventory_services import apply_purchase_invoice_stock
                # Price suggestions (Layer 2 'purchases.update_sale_price_on_receipt') are
                # only surfaced on the AJAX confirm flow (api_purchase_invoice_confirm) below,
                # where a JSON round-trip to a modal is straightforward.
                apply_purchase_invoice_stock(invoice, user=request.user)

                from .models import ProductVariant

                # Clothes market: receive stock into per-size variants too (in addition to
                # the product-level batch/warehouse stock apply_purchase_invoice_stock just
                # applied). Increments existing variant stock rather than overwriting it —
                # receiving is additive, unlike setting counts from the product form.
                for inv_item in invoice.items.select_related('product').all():
                    if not inv_item.size_breakdown:
                        continue
                    for size_id_str, size_qty_str in inv_item.size_breakdown.items():
                        if not str(size_id_str).isdigit():
                            continue
                        try:
                            size_qty = Decimal(str(size_qty_str))
                        except Exception:
                            continue
                        if size_qty <= 0:
                            continue
                        variant, _ = ProductVariant.objects.get_or_create(
                            product=inv_item.product, size_id=int(size_id_str), color='',
                            defaults={'stock_quantity': Decimal('0')},
                        )
                        variant.stock_quantity = (variant.stock_quantity or Decimal('0')) + size_qty
                        variant.save(update_fields=['stock_quantity'])
                    if not inv_item.product.has_variants:
                        inv_item.product.has_variants = True
                        inv_item.product.save(update_fields=['has_variants'])

                if paid_amount > 0 and account_id:
                    from financial.models import Account, Transaction, DailyShift
                    account = Account.objects.get(id=account_id)
                    shift = DailyShift.objects.filter(is_closed=False).last()

                    if not shift:
                        raise ValueError("يجب فتح وردية (شيفت) أولاً لتتمكن من تسجيل مبالغ مدفوعة")

                    if account.balance < paid_amount:
                        raise ValueError(
                            f"رصيد حساب {account.name} ({account.balance} ج.م) لا يكفي لسداد {paid_amount} ج.م"
                        )

                    Transaction.objects.create(
                        shift=shift,
                        account=account,
                        transaction_type='SUPPLIER_PAYMENT',
                        amount=paid_amount,
                        description=f"سداد (جزء من) فاتورة مشتريات #{invoice.id} للمورد: {supplier.name}",
                        created_by=request.user
                    )

                # Phase 7.1: if this invoice was received against a PO, update the PO's
                # received quantities (over-receipt allowed) and recompute its status.
                if po_id:
                    po = PurchaseOrder.objects.filter(id=po_id).first()
                    if po:
                        from collections import defaultdict
                        received_by_product = defaultdict(Decimal)
                        for it in items:
                            received_by_product[int(it['product_id'])] += Decimal(str(it['quantity']))
                        for poi in po.items.all():
                            got = received_by_product.get(poi.product_id, Decimal('0'))
                            if got > 0:
                                poi.received_quantity += got
                                poi.save(update_fields=['received_quantity'])
                        if all(i.received_quantity >= i.ordered_quantity for i in po.items.all()):
                            po.status = 'RECEIVED'
                        else:
                            po.status = 'PARTIAL'
                        po.save(update_fields=['status'])
                        invoice.notes = (invoice.notes + f' | من أمر شراء {po.po_number}').strip(' |')
                        invoice.save(update_fields=['notes'])

        return JsonResponse({'status': 'success', 'message': 'تم حفظ الفاتورة بنجاح', 'invoice_id': invoice.id})

    except ValueError as ve:
        return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'خطأ في النظام: {str(e)}'}, status=500)

@login_required
@require_granular_action('inventory', 'purchase_invoices', 'products', 'view')
def purchase_invoice_list(request):
    """قائمة فواتير المشتريات"""
    invoices = PurchaseInvoice.objects.all().order_by('-created_at')
    return render(request, 'products/purchase_invoice_list.html', {
        'invoices': invoices,
        'title': 'سجل فواتير المشتريات'
    })

@login_required
@require_permission('products', 'view')
def purchase_invoice_detail(request, pk):
    """تفاصيل فاتورة المشتريات"""
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    has_discount = any(item.discount > 0 for item in invoice.items.all())
    has_tax = any(item.tax_rate > 0 for item in invoice.items.all())
    
    remaining_amount = invoice.net_amount - invoice.paid_amount
    payment_percentage = 0
    if invoice.net_amount > 0:
        payment_percentage = (invoice.paid_amount / invoice.net_amount) * 100
        if payment_percentage > 100:
            payment_percentage = 100
        elif payment_percentage < 0:
            payment_percentage = 0

    sys_settings = SystemSetting.objects.first()
    return render(request, 'products/purchase_invoice_detail.html', {
        'invoice': invoice,
        'has_discount': has_discount,
        'has_tax': has_tax,
        'remaining_amount': remaining_amount,
        'payment_percentage': payment_percentage,
        'title': f'تفاصيل فاتورة شراء #{invoice.id}',
        'sys_settings': sys_settings
    })

@login_required
@require_permission('products', 'view')
def purchase_invoice_print_barcodes(request, pk):
    """طباعة باركود لأصناف الفاتورة"""
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    products = []
    import copy
    for item in invoice.items.all():
        if item.size_breakdown:
            # Clothes: one barcode per unit, each labeled with its OWN size — not
            # every size lumped onto every label (e.g. S=1/M=1/L=1 → 3 separate
            # labels, one per size, instead of 3 identical labels all reading "S، M، L").
            size_names = {s.id: s.name for s in item.product.sizes.all()}
            for size_id_str, size_qty_str in item.size_breakdown.items():
                if not str(size_id_str).isdigit():
                    continue
                try:
                    size_qty = int(float(size_qty_str))
                except (ValueError, TypeError):
                    continue
                size_qty = min(size_qty, 100)  # safety limit per size
                size_name = size_names.get(int(size_id_str), '')
                for _ in range(size_qty):
                    p_copy = copy.copy(item.product)
                    p_copy.expiry_date_str = item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else ""
                    p_copy.batch_number_str = item.batch_number or ""
                    p_copy.forced_size = size_name
                    products.append(p_copy)
            continue

        # Print quantity of barcodes for each product
        qty = int(item.quantity)
        if qty > 100: qty = 100 # Safety limit
        for _ in range(qty):
            prod = item.product
            p_copy = copy.copy(prod)
            p_copy.expiry_date_str = item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else ""
            p_copy.batch_number_str = item.batch_number or ""
            products.append(p_copy)

    sys_settings = SystemSetting.objects.first()
    return render(request, 'products/print_barcodes.html', {
        'products': products,
        'title': f'باركود فاتورة #{invoice.id}',
        'shop_name': sys_settings.shop_name if sys_settings else '',
    })

@login_required
@require_granular_action('inventory', 'purchase_return', 'products', 'view')
def purchase_return_create(request, pk=None):
    """واجهة إنشاء مرتجع مشتريات"""
    invoice = None
    if pk:
        invoice = get_object_or_404(PurchaseInvoice, pk=pk)
        
    suppliers = Supplier.objects.filter(is_active=True)
    warehouses = Warehouse.objects.filter(is_active=True)
    try:
        from financial.models import Account
        # Only real cash/bank/wallet accounts — not the nominal chart-of-accounts rows
        # (AR/AP/COGS/Revenue/...), which exist purely for the internal double-entry
        # journal and would silently corrupt those balances if hand-picked here.
        accounts = Account.exclude_dead_duplicates(Account.objects.filter(
            is_active=True,
            account_type__in=['CASH_DRAWER', 'SAFE', 'BANK', 'VODAFONE_CASH', 'INSTAPAY'],
        ))
    except ImportError:
        accounts = []

    return render(request, 'products/purchase_return_create.html', {
        'invoice': invoice,
        'suppliers': suppliers,
        'warehouses': warehouses,
        'accounts': accounts,
        'title': 'إضافة مرتجع مشتريات'
    })

@csrf_exempt
@login_required
@require_permission('products', 'view')
@require_POST
def api_purchase_return_submit(request):
    """AJAX API لإرسال مرتجع المشتريات"""
    try:
        data = json.loads(request.body)
        supplier_id = data.get('supplier_id')
        warehouse_id = data.get('warehouse_id')
        account_id = data.get('account_id')
        items = data.get('items', [])
        notes = data.get('notes', '')
        invoice_id = data.get('invoice_id') or None

        if not supplier_id:
            return JsonResponse({'status': 'error', 'message': 'يجب اختيار المورد'}, status=400)
        if not warehouse_id:
            return JsonResponse({'status': 'error', 'message': 'يجب اختيار المخزن'}, status=400)
        if not items:
            return JsonResponse({'status': 'error', 'message': 'لا توجد أصناف للمرتجع — أدخل كميات أولاً'}, status=400)

        supplier = get_object_or_404(Supplier, id=supplier_id)
        warehouse = get_object_or_404(Warehouse, id=warehouse_id)

        with db_transaction.atomic():
            total_amount = Decimal('0.00')
            for item in items:
                qty = Decimal(str(item['quantity']))
                price = Decimal(str(item['unit_price']))
                total_amount += qty * price

            p_return = PurchaseReturn.objects.create(
                supplier=supplier,
                user=request.user,
                warehouse=warehouse,
                original_invoice_id=invoice_id if invoice_id else None,
                total_amount=total_amount,
                refund_method='cash' if account_id else 'debt',
                notes=notes
            )

            if account_id:
                from financial.models import Account
                acc = get_object_or_404(Account, id=account_id)
                p_return.account = acc
                p_return.save(update_fields=['account'])

            for item_data in items:
                product = get_object_or_404(Product, id=item_data['product_id'])
                qty = Decimal(str(item_data['quantity']))
                unit_price = Decimal(str(item_data['unit_price']))
                line_subtotal = qty * unit_price

                if qty <= 0:
                    continue

                PurchaseReturnItem.objects.create(
                    purchase_return=p_return,
                    product=product,
                    quantity=qty,
                    unit_price=unit_price,
                    subtotal=line_subtotal,
                )

                # Reduce from batches (newest first for returns — LIFO), then resync
                # WarehouseStock FROM the batches (not a separate F() decrement) so the
                # two ledgers can never disagree. Previously this blindly subtracted the
                # full qty from WarehouseStock while only deducting from batches up to
                # whatever was actually available there, silently dropping any shortfall —
                # a return submitted for more than the batches hold used to desync
                # WarehouseStock from Σ StockBatch.current_quantity with no error raised.
                batches = (
                    StockBatch.objects
                    .select_for_update()
                    .filter(product=product, warehouse=warehouse, current_quantity__gt=0)
                    .order_by('-created_at')
                )
                remaining_to_return = qty
                for batch in batches:
                    if remaining_to_return <= 0:
                        break
                    deduct = min(batch.current_quantity, remaining_to_return)
                    batch.current_quantity -= deduct
                    remaining_to_return -= deduct
                    batch.save()

                if remaining_to_return > 0:
                    raise ValueError(
                        f"لا يمكن إرجاع {qty} من {product.name} — المتاح فعلياً في المخازن "
                        f"{qty - remaining_to_return} فقط."
                    )

                from products.inventory_services import _resync_warehouse_stock
                _resync_warehouse_stock(product, warehouse)

                # Record Stock Transaction
                StockTransaction.objects.create(
                    product=product,
                    warehouse=warehouse,
                    transaction_type='RET_OUT',
                    quantity=qty,
                    unit_price=unit_price,
                    reference_number=f"مرتجع #{p_return.id}",
                    note=f"مرتجع مشتريات للمورد: {supplier.name}",
                    created_by=request.user,
                )

                product.calculate_total_stock()
                product.update_cost_price()

            # 3. Update Original Invoice Status if fully returned
            if invoice_id:
                invoice = PurchaseInvoice.objects.get(id=invoice_id)
                # We can mark it as RETURNED. 
                # Ideally we check if items returned match invoice items, 
                # but for simplicity and based on user request "make it مرتجع":
                invoice.status = 'RETURNED'
                invoice.save(update_fields=['status'])

            # Financial Transaction: if cash refund, deposit money back into account
            if account_id and total_amount > 0:
                from financial.models import Account, Transaction, DailyShift
                account = Account.objects.get(id=account_id)
                shift = DailyShift.objects.filter(is_closed=False).last()

                if not shift:
                    raise ValueError("يجب فتح وردية (شيفت) أولاً لتتمكن من تسجيل مبالغ مستردة نقدياً")

                Transaction.objects.create(
                    shift=shift,
                    account=account,
                    transaction_type='INCOME',
                    amount=total_amount,
                    description=f"استرداد نقدي — مرتجع مشتريات #{p_return.id} من المورد: {supplier.name}",
                    created_by=request.user,
                )

        return JsonResponse({
            'status': 'success',
            'message': f'تم تسجيل مرتجع بإجمالي {float(total_amount):.2f} ج.م وتحديث المخازن بنجاح',
            'return_id': p_return.id,
        })

    except ValueError as ve:
        return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('api_purchase_return_submit error')
        return JsonResponse({'status': 'error', 'message': f'خطأ في النظام: {str(e)}'}, status=500)

@login_required
@require_permission('products', 'view')
def supplier_statement(request, pk):
    """كشف حساب المورد — chronological payables subledger with running balance (Phase 4.3)."""
    from datetime import datetime
    from .statements import build_supplier_statement

    supplier = get_object_or_404(Supplier, pk=pk)

    def _parse(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None

    date_from = _parse(request.GET.get('date_from'))
    date_to = _parse(request.GET.get('date_to'))
    statement = build_supplier_statement(supplier, date_from, date_to)

    from settings.models import SystemSetting
    return render(request, 'products/supplier_statement.html', {
        'title': f"كشف حساب المورد: {supplier.name}",
        'supplier': supplier,
        'statement': statement,
        'sys_settings': SystemSetting.objects.first(),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_permission('products', 'view')
def supplier_profile(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    invoices = supplier.invoices.all().order_by('-created_at')
    payments = supplier.payments.all().order_by('-date')
    returns = supplier.returns.all().order_by('-created_at')
    
    try:
        from financial.models import Account
        # Only real cash/bank/wallet accounts — not the nominal chart-of-accounts rows
        # (AR/AP/COGS/Revenue/...), which exist purely for the internal double-entry
        # journal and would silently corrupt those balances if hand-picked here.
        accounts = Account.exclude_dead_duplicates(Account.objects.filter(
            is_active=True,
            account_type__in=['CASH_DRAWER', 'SAFE', 'BANK', 'VODAFONE_CASH', 'INSTAPAY'],
        ))
    except ImportError:
        accounts = []
        
    return render(request, 'products/supplier_profile.html', {
        'supplier': supplier,
        'invoices': invoices,
        'payments': payments,
        'returns': returns,
        'accounts': accounts,
        'title': f'كشف حساب المورد: {supplier.name}'
    })

@login_required
@require_permission('products', 'view')
@require_POST
def supplier_add_purchase(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    invoice_number = request.POST.get('invoice_number', '')
    try:
        total_amount = Decimal(str(request.POST.get('total_amount') or 0))
        discount = Decimal(str(request.POST.get('discount') or 0))
        paid_amount = Decimal(str(request.POST.get('paid_amount') or 0))
    except (InvalidOperation, ValueError):
        messages.error(request, 'قيمة رقمية غير صحيحة.')
        return redirect('supplier_profile', pk=supplier.id)
    notes = request.POST.get('notes', '')

    PurchaseInvoice.objects.create(
        supplier=supplier,
        user=request.user,
        invoice_number=invoice_number,
        total_amount=total_amount,
        discount=discount,
        paid_amount=paid_amount,
        notes=notes
    )
    messages.success(request, 'تمت إضافة الفاتورة بنجاح')
    return redirect('supplier_profile', pk=supplier.id)

@login_required
@require_permission('products', 'view')
@require_POST
def supplier_add_payment(request, pk):
    """
    Settle a supplier debt.
    Supports AJAX (X-Requested-With: XMLHttpRequest) and plain POST.
    Always deducts from the chosen account atomically.
    """
    from django.db import transaction as db_tx

    supplier   = get_object_or_404(Supplier, pk=pk)
    is_ajax    = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    amount_raw     = request.POST.get('amount', '0').strip()
    account_id     = request.POST.get('account_id', '').strip()
    notes          = request.POST.get('notes', '').strip()
    payment_method = request.POST.get('payment_method', 'cash').strip()

    def _err(msg, status=400):
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': msg}, status=status)
        messages.error(request, msg)
        return redirect('supplier_profile', pk=supplier.id)

    # ── Validate amount ──────────────────────────────────────────────────
    try:
        amount = Decimal(amount_raw)
        if amount <= 0:
            return _err('المبلغ يجب أن يكون أكبر من صفر')
    except Exception:
        return _err('المبلغ المُدخل غير صحيح')

    if not account_id:
        return _err('يجب اختيار حساب الدفع أولاً')

    try:
        from financial.models import Account, Transaction, DailyShift
        account = Account.objects.filter(id=account_id, is_active=True).first()
        if not account:
            return _err('الحساب المحدد غير موجود أو غير نشط')

        # ── Balance guard ────────────────────────────────────────────────
        if account.balance < amount:
            return _err(
                'رصيد الحساب "{name}" ({bal:.2f} ج.م) غير كافٍ لسداد {amt:.2f} ج.م'.format(
                    name=account.name, bal=float(account.balance), amt=float(amount)
                )
            )

        # ── Atomic: SupplierPayment + financial Transaction ──────────────
        with db_tx.atomic():
            from sales.models import DocumentSequence
            payment = SupplierPayment.objects.create(
                supplier=supplier,
                user=request.user,
                amount=amount,
                payment_method=payment_method,
                notes=notes,
                voucher_number=DocumentSequence.next_number('PV'),  # سند صرف
            )
            shift = DailyShift.objects.filter(is_closed=False).last()
            extra = ' — ' + notes if notes else ''
            txn = Transaction.objects.create(
                shift=shift,
                account=account,
                transaction_type='SUPPLIER_PAYMENT',
                amount=amount,
                description='سداد دين مورد: {name} | سند {vno}{extra}'.format(
                    name=supplier.name, vno=payment.voucher_number, extra=extra
                ),
                created_by=request.user,
                supplier_payment=payment,  # Phase 1.2: FK link
            )

        # Refresh to pick up updated balance from F() expression
        account.refresh_from_db()
        supplier.refresh_from_db()

        success_msg = 'تم سداد {amt:.2f} ج.م وخُصمت من "{name}" بنجاح'.format(
            amt=float(amount), name=account.name
        )

        if is_ajax:
            return JsonResponse({
                'status':              'ok',
                'message':             success_msg,
                'new_account_balance': float(account.balance),
                'new_outstanding':     float(supplier.outstanding_balance),
                'payment_id':          payment.id,
                'payment_date':        payment.date.strftime('%Y-%m-%d %H:%M'),
                'amount':              float(amount),
                'account_name':        account.name,
                'notes':               notes,
            })

        messages.success(request, success_msg)

    except Exception as exc:
        import logging
        logging.getLogger(__name__).exception('supplier_add_payment error')
        return _err('خطأ غير متوقع: ' + str(exc), status=500)

    return redirect('supplier_profile', pk=supplier.id)
from django.views.decorators.csrf import csrf_exempt

@login_required
def costing_view(request):
    """
    View for the secure Costing Tool.
    """
    return render(request, 'products/costing.html', {
        'title': 'حساب التكاليف',
        'is_costing_page': True,
        'categories': Category.objects.filter(is_active=True).order_by('name'),
    })

@csrf_exempt
@login_required
def api_costing_unlock(request):
    """
    Verifies the submitted password against the master (owner) account's real
    password, rather than a hardcoded string baked into the template's JS.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

    password = json.loads(request.body or '{}').get('password', '')
    if not password:
        return JsonResponse({'status': 'error', 'ok': False})

    from django.contrib.auth import authenticate
    from accounts.models import UserProfile

    master_profiles = UserProfile.objects.filter(is_master=True).select_related('user')
    ok = any(
        authenticate(username=p.user.username, password=password) is not None
        for p in master_profiles if p.user and p.user.is_active
    )
    return JsonResponse({'status': 'success', 'ok': ok})

@csrf_exempt
@login_required
def api_update_product_cost(request):
    """
    AJAX API to update a product's cost_price.
    """
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        new_cost = request.POST.get('new_cost')
        
        try:
            product = Product.objects.get(id=product_id)
            old_cost = Decimal(str(product.cost_price))
            new_cost_decimal = Decimal(str(new_cost))
            
            if old_cost != new_cost_decimal:
                product.cost_price = new_cost_decimal
                product.save()
                
                changes = [{'name': product.name, 'type': 'تكلفة (سريع)', 'old': old_cost, 'new': new_cost_decimal}]
                notify_price_changes(request, changes, 'سعر التكلفة')
                
            return JsonResponse({'status': 'success', 'message': 'تم تحديث سعر التكلفة بنجاح'})
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'المنتج غير موجود'}, status=404)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
@login_required
@require_permission('products', 'edit')
def update_price_api(request):
    """
    AJAX API to update a product's retail price.
    """
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        new_price = request.POST.get('new_price')
        
        try:
            product = Product.objects.get(id=product_id)
            old_price = Decimal(str(product.price_retail))
            new_price_decimal = Decimal(str(new_price))
            
            if old_price != new_price_decimal:
                product.price_retail = new_price_decimal
                # Also adjust wholesale prices if needed or just retail? User only specified current -> suggested.
                product.save(update_fields=['price_retail', 'updated_at'])
                
                # Optional notification if you have it
                try:
                    changes = [{'name': product.name, 'type': 'سعر القطاعي', 'old': old_price, 'new': new_price_decimal}]
                    notify_price_changes(request, changes, 'تحديث سعر البيع')
                except Exception:
                    pass
                
            return JsonResponse({'status': 'success', 'message': 'تم تحديث السعر بنجاح'})
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'المنتج غير موجود'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
@login_required
@require_permission('products', 'edit')
def quick_update_product_ajax(request):
    """
    AJAX API to update a single field of a product.
    Supports: price_retail, price_wholesale, stock, name, sku, category_id, is_active.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('product_id')
            field = data.get('field')
            value = data.get('value')
            
            product = Product.objects.get(id=product_id)
            
            if field == 'price_retail':
                product.price_retail = Decimal(str(value))
                product.save(update_fields=['price_retail', 'updated_at'])
            elif field == 'price_wholesale':
                product.price_wholesale = Decimal(str(value))
                product.save(update_fields=['price_wholesale', 'updated_at'])
            elif field == 'stock':
                product.stock_quantity = Decimal(str(value))
                product.save(update_fields=['stock_quantity', 'updated_at'])
            elif field == 'name':
                value = str(value).strip()
                if not value:
                    return JsonResponse({'status': 'error', 'message': 'اسم المنتج لا يمكن أن يكون فارغاً'}, status=400)
                product.name = value
                product.save(update_fields=['name', 'updated_at'])
            elif field == 'sku':
                value = str(value).strip()
                if not value:
                    return JsonResponse({'status': 'error', 'message': 'كود المنتج لا يمكن أن يكون فارغاً'}, status=400)
                if Product.objects.filter(sku=value).exclude(id=product_id).exists():
                    return JsonResponse({'status': 'error', 'message': 'هذا الكود (SKU) مستخدم بالفعل مع منتج آخر!'}, status=400)
                product.sku = value
                product.save(update_fields=['sku', 'updated_at'])
            elif field == 'category_id':
                if value == '' or value is None:
                    product.category = None
                else:
                    product.category_id = int(value)
                product.save(update_fields=['category', 'updated_at'])
            elif field == 'is_active':
                product.is_active = value in (True, 'true', '1', 1)
                product.save(update_fields=['is_active', 'updated_at'])
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid field'}, status=400)
                
            return JsonResponse({'status': 'success', 'message': 'تم التحديث بنجاح'})
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'المنتج غير موجود'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required
@require_permission('products', 'view')
def quick_edit_products_page(request):
    """
    Dedicated page (تعديل المنتجات) to search and quick-edit all main fields of products in a grid.
    """
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    categories = Category.objects.all().order_by('name')
    products_list = Product.objects.all().select_related('category', 'kind').order_by('-created_at')
    
    # Search
    search = request.GET.get('search', '').strip()
    if search:
        products_list = products_list.filter(
            Q(name__icontains=search) |
            Q(sku__icontains=search)
        )
        
    # Category filter
    category_id = request.GET.get('category', '').strip()
    if category_id:
        products_list = products_list.filter(category_id=category_id)
        
    # Paginate (50 per page)
    paginator = Paginator(products_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Store query params to persist filters in pagination
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
        
    return render(request, 'products/quick_edit_page.html', {
        'products': page_obj,
        'categories': categories,
        'title': 'تعديل المنتجات',
        'query_params': query_params,
    })


@login_required
def api_stock_alerts_count(request):
    """
    AJAX API to get the count of products with low stock.
    """
    from settings.policies import get_policy
    if not get_policy('inventory.warn_low_stock'):
        return JsonResponse({'count': 0})
    count = (Product.objects.filter(is_active=True, stock_quantity__lte=F('low_stock_threshold'))
             .exclude(category__is_menu_category=True, is_raw_material=False).count())
    return JsonResponse({'count': count})

@login_required
def api_search_products(request):
    """
    AJAX API to search products for the costing tool.

    Plain search (default): any product, used to pick an existing menu item whose
    recipe/cost we want to load or sync.
    `?raw_only=1`: restricted to is_raw_material=True products, used by the "build a
    new recipe" material picker — each result carries its unit_measure and the units
    a recipe line may be entered in (see restaurant.models.compatible_units), so the
    picker can offer e.g. grams for an ingredient stocked in KG.
    """
    from restaurant.models import Recipe, compatible_units

    q = request.GET.get('q', '')
    raw_only = request.GET.get('raw_only') == '1'

    if raw_only:
        # Raw-material picker: with no query yet, browse the full raw-material list
        # instead of showing nothing — most cafes only have a handful, so there's no
        # need to force typing before anything appears.
        qs = Product.objects.filter(is_raw_material=True)
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(sku__icontains=q))
        qs = qs.order_by('name')
    else:
        if len(q) < 2:
            return JsonResponse({'products': []})
        qs = Product.objects.filter(Q(name__icontains=q) | Q(sku__icontains=q))

    results = []
    for p in qs.prefetch_related('variants__size')[:20]:
        row = {
            'id': p.id,
            'name': p.name,
            'sku': p.sku,
            'cost_price': float(p.cost_price),
            'unit_measure': p.unit_measure,
        }
        if raw_only:
            row['compatible_units'] = compatible_units(p.unit_measure)
        else:
            sizes = []
            if p.has_variants:
                seen = set()
                for v in p.variants.all():
                    if v.size_id and v.size_id not in seen:
                        seen.add(v.size_id)
                        sizes.append({'id': v.size_id, 'name': v.size.name})
            row['has_variants'] = p.has_variants
            row['sizes'] = sizes
            row['has_recipe'] = Recipe.objects.filter(product=p, is_active=True).exists()
        results.append(row)

    return JsonResponse({'products': results})


@login_required
def api_recipe_cost(request):
    """
    Costing tool "Option A" (منتج له وصفة بالفعل): given a product (and, for a
    has_variants product, an optional size), expands its active Recipe — direct
    ingredient lines plus any sub_recipe lines' own ingredients — into one row per
    ingredient (same expansion sales.services.preview_recipe_shortages uses), pricing
    each ingredient's line at its current Product.cost_price per its own tracked unit.
    """
    from restaurant.models import Recipe

    product_id = request.GET.get('product_id')
    size_id = request.GET.get('size_id') or None
    product = get_object_or_404(Product, pk=product_id)

    recipe = Recipe.for_product(product, size_id=size_id)
    if not recipe:
        return JsonResponse({'found': False})

    rows = {}
    for ri in recipe.items.select_related('ingredient', 'sub_recipe').all():
        if ri.ingredient_id:
            ing = ri.ingredient
            entry = rows.setdefault(ing.id, {'name': ing.name, 'unit': ing.unit_measure,
                                              'qty': Decimal('0'), 'unit_cost': ing.cost_price})
            entry['qty'] += ri.base_quantity
        elif ri.sub_recipe_id:
            for si in ri.sub_recipe.items.select_related('ingredient').all():
                ing = si.ingredient
                entry = rows.setdefault(ing.id, {'name': ing.name, 'unit': ing.unit_measure,
                                                  'qty': Decimal('0'), 'unit_cost': ing.cost_price})
                entry['qty'] += si.base_quantity * ri.quantity

    rows_list = []
    total = Decimal('0')
    for r in rows.values():
        line_total = r['qty'] * r['unit_cost']
        total += line_total
        rows_list.append({
            'name': r['name'],
            'qty': float(r['qty']),
            'unit': r['unit'],
            'unit_cost': float(r['unit_cost']),
            'line_total': float(line_total),
        })

    return JsonResponse({'found': True, 'rows': rows_list, 'total': float(total), 'notes': recipe.notes})


@csrf_exempt
@login_required
@require_POST
def api_create_recipe_product(request):
    """
    Costing tool "Option B" (إنشاء وصفة جديدة من المواد الخام): creates a brand-new
    sellable menu-item Product plus a real restaurant.Recipe/RecipeItem set from the
    raw materials the admin picked, so it works immediately in POS/KDS with real stock
    deduction — not just a cost estimate. Recomputes the material cost server-side from
    each ingredient's live cost_price (never trusts the client-side total).
    """
    from restaurant.models import Recipe, RecipeItem, convert_quantity

    try:
        data = json.loads(request.body)
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'status': 'error', 'message': 'بيانات غير صالحة'}, status=400)

    name = (data.get('name') or '').strip()
    items = data.get('items') or []
    if not name:
        return JsonResponse({'status': 'error', 'message': 'اسم المنتج مطلوب'}, status=400)
    if not items:
        return JsonResponse({'status': 'error', 'message': 'أضف مكوّناً واحداً على الأقل من المواد الخام'}, status=400)

    category_id = data.get('category_id') or None
    extra_cost = _to_decimal_or_none(data.get('extra_cost')) or Decimal('0.00')
    selling_price = _to_decimal_or_none(data.get('selling_price'))

    try:
        with db_transaction.atomic():
            materials = []
            material_cost = Decimal('0.00')
            for idx, item in enumerate(items, start=1):
                product_id = item.get('product_id')
                qty = _to_decimal_or_none(item.get('quantity'))
                unit = (item.get('unit') or '').strip()
                if not product_id or qty is None or qty <= 0:
                    raise ValueError(f'سطر {idx}: بيانات المكوّن غير مكتملة')
                material = Product.objects.filter(id=product_id, is_raw_material=True).first()
                if not material:
                    raise ValueError(f'سطر {idx}: الخامة غير موجودة')
                base_qty = convert_quantity(qty, unit, material.unit_measure)
                material_cost += base_qty * material.cost_price
                materials.append((material, qty, unit))

            total_cost = material_cost + extra_cost
            if selling_price is None:
                profit_pct = _to_decimal_or_none(data.get('profit_pct')) or Decimal('0')
                pct = profit_pct / Decimal('100')
                selling_price = total_cost / (1 - pct) if pct < 1 else total_cost * (1 + pct)

            product = Product(name=name, is_active=True, unit_measure='PCS')
            product.sku = _next_auto_sku()
            product.barcode = _generate_ean13_from_sku(product.sku)
            if category_id:
                product.category = Category.objects.filter(id=category_id).first()
            product.cost_price = total_cost.quantize(Decimal('0.01'))
            product.price_retail = selling_price.quantize(Decimal('0.01'))
            product.low_stock_threshold = Decimal('0.00')
            product.save()

            recipe = Recipe.objects.create(product=product, size=None)
            for material, qty, unit in materials:
                RecipeItem.objects.create(recipe=recipe, ingredient=material, quantity=qty, unit=unit)

    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'خطأ غير متوقع: {str(e)}'}, status=500)

    return JsonResponse({
        'status': 'success',
        'id': product.id,
        'cost_price': float(product.cost_price),
        'price_retail': float(product.price_retail),
    })

import json
from .models import ProductCosting

@login_required
def costing_list_view(request):
    """
    List all saved costings.
    """
    query = request.GET.get('q', '')
    costings = ProductCosting.objects.all()
    if query:
        costings = costings.filter(name__icontains=query)
    
    return render(request, 'products/costing_list.html', {
        'costings': costings,
        'query': query,
        'title': 'سجل حساب التكاليف'
    })

@csrf_exempt
@login_required
def api_save_costing(request):
    """
    AJAX API to save a costing configuration.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            costing = ProductCosting.objects.create(
                name=data.get('name', 'بدون اسم'),
                linked_product_id=data.get('product_id'),
                mode=data.get('mode'),
                total_cost=data.get('total_cost'),
                selling_price=data.get('selling_price'),
                config_json=data.get('config'),
                notes=data.get('notes', '')
            )
            return JsonResponse({'status': 'success', 'id': costing.id, 'message': 'تم حفظ الحسبة بنجاح'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
def costing_print_view(request, pk):
    """
    View to render a clean print layout for a saved costing.
    """
    costing = get_object_or_404(ProductCosting, pk=pk)
    print_mode = request.GET.get('mode', 'a4') # a4 or thermal
    
    return render(request, 'products/costing_print.html', {
        'costing': costing,
        'mode': print_mode,
        'title': f'طباعة - {costing.name}'
    })

@login_required
@require_permission('products', 'view')
def bulk_transaction_create(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            items = data.get('items', [])
            
            if not items:
                return JsonResponse({'success': False, 'error': 'لا توجد حركات للتسجيل'})

            with db_transaction.atomic():
                for index, item in enumerate(items):
                    row_num = index + 1
                    warehouse_id = item.get('warehouse_id')
                    product_id = item.get('product_id')
                    type_code = item.get('transaction_type')
                    qty = item.get('quantity')
                    note = item.get('note', '')

                    # Basic Validation
                    if not warehouse_id:
                        raise ValueError(f"المخزن مطلوب في السطر {row_num}")
                    if not product_id:
                        raise ValueError(f"المنتج مطلوب في السطر {row_num}")
                    if not type_code:
                        raise ValueError(f"نوع الحركة مطلوب في السطر {row_num}")
                    
                    try:
                        qty = Decimal(str(qty))
                    except (ValueError, TypeError, InvalidOperation):
                        raise ValueError(f"الكمية يجب أن تكون رقماً في السطر {row_num}")
                    
                    if qty <= 0:
                        raise ValueError(f"الكمية يجب أن تكون أكبر من صفر في السطر {row_num}")

                    # Fetch Objects
                    try:
                        warehouse = Warehouse.objects.get(id=warehouse_id)
                        product = Product.objects.get(id=product_id)
                    except (Warehouse.DoesNotExist, Product.DoesNotExist):
                        raise ValueError(f"بيانات غيرة صحيحة في السطر {row_num}")

                    # Stock Handling
                    stock, created = WarehouseStock.objects.select_for_update().get_or_create(
                        warehouse=warehouse,
                        product=product,
                        defaults={'quantity': Decimal('0.00')},
                    )
                    stock_qty = Decimal(str(stock.quantity or '0'))

                    if type_code in ['IN', 'RET']:
                        stock_qty += qty
                    elif type_code == 'OUT':
                        if stock_qty < qty:
                            raise ValueError(f"الرصيد غير كافي للصنف {product.name} في المخزن {warehouse.name}. المتاح: {stock_qty}")
                        stock_qty -= qty

                    stock.quantity = stock_qty
                    
                    stock.save()

                    # Create Transaction
                    StockTransaction.objects.create(
                        product=product,
                        warehouse=warehouse,
                        transaction_type=type_code,
                        quantity=qty,
                        note=note or "تسجيل بالجملة"
                    )

                    # Update Global Stock
                    product.calculate_total_stock()

            return JsonResponse({'success': True, 'message': 'تم تسجيل الحركات بنجاح'})

        except ValueError as e:
            return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f"خطأ غير متوقع: {str(e)}"})

    # GET Request
    products_data = list(Product.objects.filter(is_active=True).values('id', 'name', 'sku', 'stock_quantity'))
    avail_warehouses = Warehouse.objects.filter(is_active=True)

    return render(request, 'products/bulk_transaction_form.html', {
        'title': 'تسجيل حركات مخزنية بالجملة',
        'products_data': products_data,
        'avail_warehouses': avail_warehouses,
    })

@login_required
@require_permission('products', 'view')
def warehouse_print(request, pk, print_format):
    from django.db.models import Sum
    from products.models import Warehouse, WarehouseStock, Product
    from settings.models import SystemSetting

    warehouse = get_object_or_404(Warehouse, pk=pk)
    stocks = WarehouseStock.objects.filter(warehouse=warehouse).select_related('product', 'product__category').order_by('product__name')

    top_sold_items = Product.objects.filter(
        orderitem__order__warehouse=warehouse
    ).annotate(
        sold_qty=Sum('orderitem__quantity')
    ).exclude(sold_qty=None).order_by('-sold_qty')[:5]

    sys_settings = SystemSetting.objects.first()

    context = {
        'warehouse': warehouse,
        'stocks': stocks,
        'top_sold_items': top_sold_items,
        'sys_settings': sys_settings,
        'title': f'طباعة رصيد المخزن ({warehouse.name})',
        'print_date': timezone.now()
    }
    
    if print_format == 'a4':
        return render(request, 'products/warehouse_print_a4.html', context)
    elif print_format == 'a5':
        return render(request, 'products/warehouse_print_a5.html', context)
    elif print_format == 'thermal':
        return render(request, 'products/warehouse_print_thermal.html', context)
    else:
        return redirect('warehouse_detail', pk=pk)

@login_required
@require_permission('products', 'view')
def warehouse_export_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum
    from settings.models import SystemSetting

    warehouse = get_object_or_404(Warehouse, pk=pk)
    stocks = WarehouseStock.objects.filter(warehouse=warehouse).select_related('product', 'product__category').order_by('product__name')
    
    top_sold_items = Product.objects.filter(
        orderitem__order__warehouse=warehouse
    ).annotate(
        sold_qty=Sum('orderitem__quantity')
    ).exclude(sold_qty=None).order_by('-sold_qty')[:5]

    sys_settings = SystemSetting.objects.first()
    shop_name = sys_settings.shop_name if sys_settings else 'MR MEKAWY'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'رصيد {warehouse.name}'
    ws.sheet_view.rightToLeft = True

    # Styles
    header_font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    bold_font = Font(name='Arial', bold=True, size=12)
    normal_font = Font(name='Arial', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    row = 1
    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 1, f"تقرير رصيد المخزن: {warehouse.name}").font = bold_font
    ws.cell(row, 1).alignment = center_align
    row += 2

    # Top Sold Items
    if top_sold_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1, "أكثر 5 أصناف مبيعاً").font = Font(name='Arial', bold=True, size=13, color='047857')
        row += 1
        headers_top = ['م', 'كود الصنف', 'اسم الصنف', 'الكمية المباعة']
        for col, h in enumerate(headers_top, 1):
            cell = ws.cell(row, col, h)
            cell.font = bold_font
            cell.alignment = center_align
        row += 1
        for i, item in enumerate(top_sold_items, 1):
            ws.cell(row, 1, i).alignment = center_align
            ws.cell(row, 2, item.sku).font = normal_font
            ws.cell(row, 2).alignment = center_align
            ws.cell(row, 3, item.name).font = normal_font
            ws.cell(row, 3).alignment = center_align
            ws.cell(row, 4, float(item.sold_qty)).font = bold_font
            ws.cell(row, 4).alignment = center_align
            row += 1
        row += 2
    
    # Stock
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 1, "الرصيد الحالي").font = Font(name='Arial', bold=True, size=13, color='1F4E79')
    row += 1
    headers_stock = ['م', 'كود الصنف', 'اسم الصنف', 'الرصيد']
    for col, h in enumerate(headers_stock, 1):
        cell = ws.cell(row, col, h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
    row += 1
    for i, stock in enumerate(stocks, 1):
        ws.cell(row, 1, i).alignment = center_align
        ws.cell(row, 2, stock.product.sku).font = normal_font
        ws.cell(row, 2).alignment = center_align
        ws.cell(row, 3, stock.product.name).font = normal_font
        ws.cell(row, 3).alignment = center_align
        ws.cell(row, 4, float(stock.quantity)).font = bold_font
        ws.cell(row, 4).alignment = center_align
        row+=1

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions['C'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="warehouse_{warehouse.id}_stock.xlsx"'
    wb.save(response)
    return response


# ══════════════════════════════════════════════
# MODULE 3 — MASTER DATA VIEWS
# ══════════════════════════════════════════════

# ── KINDS ──
@login_required
@require_granular_action('master_data', 'kinds', 'products', 'view')
def kind_list(request):
    kinds = Kind.objects.select_related('category').order_by('category__name', 'name')
    categories = Category.objects.filter(is_active=True)
    return render(request, 'products/kind_list.html', {
        'kinds': kinds, 'categories': categories, 'title': 'الأنواع'
    })

@login_required
@require_permission('products', 'view')
def kind_create(request):
    if request.method == 'POST':
        form = KindForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم إضافة النوع بنجاح')
            return redirect('kind_list')
    else:
        form = KindForm()
    return render(request, 'products/kind_form.html', {'form': form, 'title': 'إضافة نوع'})

@login_required
@require_permission('products', 'view')
def kind_update(request, pk):
    kind = get_object_or_404(Kind, pk=pk)
    if request.method == 'POST':
        form = KindForm(request.POST, instance=kind)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم التعديل بنجاح')
            return redirect('kind_list')
    else:
        form = KindForm(instance=kind)
    return render(request, 'products/kind_form.html', {'form': form, 'title': 'تعديل نوع', 'kind': kind})

@login_required
@require_permission('products', 'view')
@require_POST
def kind_delete(request, pk):
    kind = get_object_or_404(Kind, pk=pk)
    kind.delete()
    messages.success(request, 'تم الحذف')
    return redirect('kind_list')

# API: kinds filtered by category
@login_required
def api_kinds_by_category(request):
    cat_id = request.GET.get('category_id')
    qs = Kind.objects.filter(is_active=True)
    if cat_id:
        qs = qs.filter(category_id=cat_id)
    data = list(qs.values('id', 'name'))
    return JsonResponse({'kinds': data})

# ── SIZES ──
@login_required
@require_granular_action('master_data', 'sizes', 'products', 'view')
def size_list(request):
    sizes = Size.objects.all()
    return render(request, 'products/size_list.html', {'sizes': sizes, 'title': 'المقاسات'})

@login_required
@require_permission('products', 'view')
@require_POST
def size_create(request):
    name = request.POST.get('name', '').strip()
    size_type = request.POST.get('size_type', 'custom')
    if not name:
        messages.error(request, 'اسم المقاس مطلوب')
        return redirect('size_list')
    Size.objects.get_or_create(name=name, defaults={'size_type': size_type, 'sort_order': Size.objects.count()})
    messages.success(request, f'تم إضافة المقاس "{name}"')
    return redirect('size_list')

@login_required
@require_permission('products', 'view')
@require_POST
def size_reorder(request):
    try:
        data = json.loads(request.body)
        for item in data.get('order', []):
            Size.objects.filter(pk=item['id']).update(sort_order=item['order'])
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_permission('products', 'view')
@require_POST
def size_delete(request, pk):
    size = get_object_or_404(Size, pk=pk)
    size.delete()
    messages.success(request, 'تم الحذف')
    return redirect('size_list')

# ── UNITS ──
@login_required
@require_granular_action('master_data', 'units', 'products', 'view')
def unit_list(request):
    units = UnitOfMeasure.objects.all()
    form = UnitOfMeasureForm()
    return render(request, 'products/unit_list.html', {'units': units, 'form': form, 'title': 'وحدات القياس'})

@login_required
@require_permission('products', 'view')
@require_POST
def unit_create(request):
    form = UnitOfMeasureForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'تمت الإضافة بنجاح')
    return redirect('unit_list')

@login_required
@require_POST
def quick_create_unit_ajax(request):
    """AJAX endpoint to quickly create a new unit of measure"""
    try:
        import json
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
            
        name = data.get('name', '').strip()
        abbreviation = data.get('abbreviation', '').strip()
        
        if not name:
            return JsonResponse({'success': False, 'error': 'اسم الوحدة مطلوب'})
            
        if not abbreviation:
            # Auto-generate abbreviation (first 10 chars of name)
            abbreviation = name[:10].strip().upper()
            
        # Check if already exists
        if UnitOfMeasure.objects.filter(abbreviation=abbreviation).exists():
            existing = UnitOfMeasure.objects.filter(abbreviation=abbreviation).first()
            return JsonResponse({
                'success': True,
                'id': existing.id,
                'name': existing.name,
                'abbreviation': existing.abbreviation,
                'message': 'الوحدة موجودة بالفعل'
            })
            
        # Create unit
        unit = UnitOfMeasure.objects.create(
            name=name,
            abbreviation=abbreviation,
            is_active=True
        )
        
        # Clear cache
        from django.core.cache import cache
        cache.delete('db_unit_measure_choices')
        
        return JsonResponse({
            'success': True,
            'id': unit.id,
            'name': unit.name,
            'abbreviation': unit.abbreviation
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_permission('products', 'view')
@require_POST
def unit_delete(request, pk):
    unit = get_object_or_404(UnitOfMeasure, pk=pk)
    unit.delete()
    messages.success(request, 'تم الحذف')
    return redirect('unit_list')

# ── CATEGORIES — enhanced ──
@login_required
@require_granular_action('master_data', 'categories', 'products', 'view')
def category_list(request):
    cats = Category.objects.all().order_by('name')
    return render(request, 'products/category_list.html', {'categories': cats, 'title': 'الأقسام'})

@login_required
@require_permission('products', 'view')
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تمت إضافة القسم')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'products/category_form.html', {'form': form, 'title': 'إضافة قسم'})

@login_required
@require_permission('products', 'view')
def category_update(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=cat)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم التعديل')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=cat)
    return render(request, 'products/category_form.html', {'form': form, 'title': 'تعديل قسم', 'cat': cat})

@login_required
@require_permission('products', 'view')
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        if cat.products.exists():
            messages.error(request, 'لا يمكن الحذف: يوجد منتجات مرتبطة بهذا القسم')
            return redirect('category_list')
        cat.delete()
        messages.success(request, 'تم الحذف')
        return redirect('category_list')
    return render(request, 'products/product_confirm_delete.html', {'object': cat, 'title': 'حذف قسم'})


# ══════════════════════════════════════════════
# MODULE 4 — PRODUCT IMAGES API
# ══════════════════════════════════════════════

@login_required
@require_POST
def product_image_upload(request, pk):
    """Accepts base64-encoded images (up to 5) for a product"""
    product = get_object_or_404(Product, pk=pk)
    try:
        data = json.loads(request.body)
        images_data = data.get('images', [])
        # Remove old images if replacing all
        if data.get('replace_all', False):
            product.images.all().delete()
        # Enforce max 5
        current_count = product.images.count()
        for i, img_b64 in enumerate(images_data[:max(0, 5 - current_count)]):
            ProductImage.objects.create(
                product=product,
                image_data=img_b64,
                order=current_count + i
            )
        return JsonResponse({'success': True, 'count': product.images.count()})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def product_image_delete(request, pk, img_pk):
    img = get_object_or_404(ProductImage, pk=img_pk, product_id=pk)
    img.delete()
    return JsonResponse({'success': True})

@login_required
def product_images_list(request, pk):
    product = get_object_or_404(Product, pk=pk)
    imgs = list(product.images.values('id', 'order', 'image_data'))
    return JsonResponse({'images': imgs})


# ══════════════════════════════════════════════
# MODULE 4 — STOCK ALERTS
# ══════════════════════════════════════════════

@login_required
@require_granular_action('inventory', 'stock_alerts', 'products', 'view')
def stock_alerts(request):
    # Kitchen-routed menu items are prepared on demand and never carry a meaningful
    # stock count — excluding them keeps this screen focused on actual stocked goods
    # (retail products and raw materials). Raw materials are kept even if they fall
    # back to the generic "بدون قسم" category (is_menu_category=True) since they have
    # no category field of their own on the raw-material add form.
    stock_tracked = Product.objects.exclude(category__is_menu_category=True, is_raw_material=False)

    low_stock = stock_tracked.filter(
        is_active=True,
        stock_quantity__gt=0,
        stock_quantity__lte=F('low_stock_threshold')
    ).select_related('category', 'supplier').order_by('stock_quantity')

    out_of_stock = stock_tracked.filter(
        is_active=True,
        stock_quantity__lte=0
    ).select_related('category', 'supplier').order_by('name')

    return render(request, 'products/stock_alerts.html', {
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
        'title': 'تنبيهات المخزون',
        'low_count': low_stock.count(),
        'out_count': out_of_stock.count(),
    })


# ══════════════════════════════════════════════
# MODULE 5 — MANUFACTURING TRANSACTIONS
# ══════════════════════════════════════════════

@login_required
@require_permission('products', 'view')
@require_POST
def create_manufacturing_transaction(request):
    try:
        data = json.loads(request.body)
        raw_product = get_object_or_404(Product, pk=data['raw_product_id'])
        raw_warehouse = get_object_or_404(Warehouse, pk=data['raw_warehouse_id'])
        finished_product = get_object_or_404(Product, pk=data['finished_product_id'])
        finished_warehouse = get_object_or_404(Warehouse, pk=data['finished_warehouse_id'])
        raw_qty = Decimal(str(data['raw_qty']))
        finished_qty = Decimal(str(data['finished_qty']))

        with db_transaction.atomic():
            debit = StockTransaction.objects.create(
                product=raw_product, warehouse=raw_warehouse,
                transaction_type='MFG', quantity=raw_qty,
                note='تصنيع — خصم خامات', unit_price=raw_product.cost_price
            )
            credit = StockTransaction.objects.create(
                product=finished_product, warehouse=finished_warehouse,
                transaction_type='MFG_OUT', quantity=finished_qty,
                note='تصنيع — إضافة منتج نهائي', linked_transaction=debit,
                unit_price=finished_product.cost_price
            )
            debit.linked_transaction = credit
            debit.save(update_fields=['linked_transaction'])

            raw_stock = WarehouseStock.objects.get(product=raw_product, warehouse=raw_warehouse)
            raw_stock.quantity = max(raw_stock.quantity - raw_qty, Decimal('0'))
            raw_stock.save()

            finished_stock, _ = WarehouseStock.objects.get_or_create(
                product=finished_product, warehouse=finished_warehouse,
                defaults={'quantity': Decimal('0')}
            )
            finished_stock.quantity += finished_qty
            finished_stock.save()

            raw_product.calculate_total_stock()
            finished_product.calculate_total_stock()

        return JsonResponse({'success': True, 'message': 'تم تسجيل عملية التصنيع بنجاح'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ══════════════════════════════════════════════
# MODULE 6 — PURCHASE ORDERS
# ══════════════════════════════════════════════

@login_required
@require_granular_action('inventory', 'purchase_orders', 'products', 'view')
def purchase_order_list(request):
    orders = PurchaseOrder.objects.select_related('supplier', 'destination_warehouse').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(status=status_filter)
    supplier_filter = request.GET.get('supplier', '')
    if supplier_filter:
        orders = orders.filter(supplier_id=supplier_filter)
    paginator = Paginator(orders, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    suppliers = Supplier.objects.filter(is_active=True)
    return render(request, 'products/purchase_order_list.html', {
        'orders': page_obj, 'suppliers': suppliers,
        'status_choices': PurchaseOrder.STATUS_CHOICES,
        'current_status': status_filter, 'title': 'أوامر الشراء',
    })

@login_required
@require_granular_action('inventory', 'purchase_orders', 'products', 'view')
def purchase_order_create(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        if form.is_valid():
            po = form.save(commit=False)
            po.created_by = request.user
            po.save()
            # Save line items from JSON
            items_json = request.POST.get('items_json', '[]')
            try:
                items = json.loads(items_json)
                total = Decimal('0')
                for item in items:
                    if not item.get('product_id'):
                        continue
                    qty = Decimal(str(item.get('quantity', 0)))
                    price = Decimal(str(item.get('unit_price', 0)))
                    disc = Decimal(str(item.get('discount', 0)))
                    poi = PurchaseOrderItem.objects.create(
                        purchase_order=po,
                        product_id=item['product_id'],
                        size_id=item.get('size_id') or None,
                        ordered_quantity=qty,
                        unit_price=price,
                        discount=disc,
                    )
                    total += poi.line_total
                po.total_amount = total
                po.save(update_fields=['total_amount'])
            except Exception:
                pass
            messages.success(request, f'تم إنشاء أمر الشراء {po.po_number}')
            return redirect('purchase_order_detail', pk=po.pk)
    else:
        form = PurchaseOrderForm()
    # A PO orders from a supplier — a menu item (prepared in-house) is never something a
    # supplier ships, only its raw materials are (same restriction as purchase invoices).
    products = (Product.objects.filter(is_active=True)
                .exclude(category__is_menu_category=True, is_raw_material=False)
                .values('id', 'name', 'sku', 'cost_price'))
    sizes = Size.objects.filter(is_active=True).values('id', 'name')
    return render(request, 'products/purchase_order_form.html', {
        'form': form, 'title': 'إنشاء أمر شراء جديد',
        'products_json': json.dumps(list(products), default=str),
        'sizes_json': json.dumps(list(sizes)),
    })

@login_required
@require_permission('products', 'view')
def purchase_order_detail(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    items = po.items.select_related('product', 'size')
    return render(request, 'products/purchase_order_detail.html', {
        'po': po, 'items': items, 'title': f'أمر الشراء {po.po_number}'
    })

@login_required
@require_permission('products', 'view')
@require_POST
def purchase_order_confirm(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if po.status == 'DRAFT':
        po.status = 'CONFIRMED'
        po.save(update_fields=['status'])
        messages.success(request, 'تم تأكيد أمر الشراء')
    return redirect('purchase_order_detail', pk=pk)

@login_required
@require_permission('products', 'view')
@require_POST
def purchase_order_cancel(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if po.status in ('DRAFT', 'CONFIRMED'):
        po.status = 'CANCELLED'
        po.save(update_fields=['status'])
        messages.success(request, 'تم إلغاء أمر الشراء')
    return redirect('purchase_order_detail', pk=pk)

@login_required
@require_permission('products', 'view')
def purchase_order_receive(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    if po.status not in ('CONFIRMED', 'PARTIAL'):
        messages.error(request, 'لا يمكن الاستلام بهذه الحالة')
        return redirect('purchase_order_detail', pk=pk)

    if request.method == 'POST':
        try:
            with db_transaction.atomic():
                items_data = json.loads(request.POST.get('items_json', '[]'))
                for item_data in items_data:
                    item = get_object_or_404(PurchaseOrderItem, pk=item_data['id'], purchase_order=po)
                    qty = Decimal(str(item_data.get('received_qty', 0)))
                    if qty <= 0:
                        continue
                    remaining = item.ordered_quantity - item.received_quantity
                    qty = min(qty, remaining)
                    item.received_quantity += qty
                    item.save(update_fields=['received_quantity'])

                    StockTransaction.objects.create(
                        product=item.product,
                        warehouse=po.destination_warehouse,
                        transaction_type='IN', quantity=qty,
                        unit_price=item.unit_price,
                        reference_number=po.po_number,
                        note=f'استلام من أمر شراء {po.po_number}'
                    )
                    stock, _ = WarehouseStock.objects.get_or_create(
                        product=item.product, warehouse=po.destination_warehouse,
                        defaults={'quantity': Decimal('0')}
                    )
                    stock.quantity += qty
                    stock.save()

                    # Receiving via a Purchase Order used to skip StockBatch entirely (unlike
                    # the Purchase Invoice receiving path), so Product.update_cost_price() —
                    # which only reads from batches — had nothing to recompute from. That made
                    # 'purchases.update_cost_on_receipt' silently do nothing here regardless of
                    # the toggle. Create the batch so this path matches invoice-based receiving.
                    StockBatch.objects.create(
                        product=item.product,
                        warehouse=po.destination_warehouse,
                        supplier=po.supplier,
                        purchase_price=item.unit_price,
                        initial_quantity=qty,
                        current_quantity=qty,
                    )

                    item.product.calculate_total_stock()
                    from settings.policies import get_policy
                    if get_policy('purchases.update_cost_on_receipt'):
                        item.product.update_cost_price()

                all_items = po.items.all()
                if all(i.is_fully_received for i in all_items):
                    po.status = 'RECEIVED'
                else:
                    po.status = 'PARTIAL'
                po.save(update_fields=['status'])

                PurchaseInvoice.objects.create(
                    supplier=po.supplier,
                    user=request.user,
                    total_amount=po.total_amount,
                    net_amount=po.total_amount,
                    notes=f'من أمر شراء {po.po_number}'
                )
            messages.success(request, 'تم تسجيل الاستلام بنجاح')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
        return redirect('purchase_order_detail', pk=pk)

    items = po.items.select_related('product', 'size')
    return render(request, 'products/purchase_order_receive.html', {
        'po': po, 'items': items, 'title': f'استلام بضاعة — {po.po_number}'
    })

@login_required
@require_permission('products', 'view')
def purchase_order_print(request, pk):
    po = get_object_or_404(PurchaseOrder, pk=pk)
    sys_settings = SystemSetting.objects.first()
    return render(request, 'products/purchase_order_print.html', {
        'po': po, 'items': po.items.select_related('product', 'size'),
        'sys_settings': sys_settings, 'print_date': timezone.now(),
    })


# ══════════════════════════════════════════════
# MODULE 6 — SUPPLIER ACCOUNT & STATEMENT
# ══════════════════════════════════════════════

@login_required
@require_permission('products', 'view')
def supplier_account(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    invoices_qs = supplier.invoices.filter(status='CONFIRMED')
    payments_qs = supplier.payments.all()
    if date_from:
        invoices_qs = invoices_qs.filter(created_at__date__gte=date_from)
        payments_qs = payments_qs.filter(date__date__gte=date_from)
    if date_to:
        invoices_qs = invoices_qs.filter(created_at__date__lte=date_to)
        payments_qs = payments_qs.filter(date__date__lte=date_to)

    entries = []

    # Opening balance row (always shown even if date filter active)
    if supplier.opening_balance and supplier.opening_balance != 0:
        entries.append({
            'date': supplier.created_at if hasattr(supplier, 'created_at') else timezone.now(),
            'ref': 'رصيد افتتاحي',
            'desc': 'رصيد افتتاحي عند بداية التشغيل',
            'debit': supplier.opening_balance if supplier.opening_balance > 0 else Decimal('0'),
            'credit': abs(supplier.opening_balance) if supplier.opening_balance < 0 else Decimal('0'),
            'type': 'opening',
        })

    for inv in invoices_qs:
        entries.append({
            'date': inv.created_at, 'ref': f'فاتورة #{inv.id}',
            'desc': inv.notes or 'فاتورة مشتريات',
            'debit': inv.net_amount, 'credit': Decimal('0'),
            'type': 'invoice',
        })
        # Any amount paid directly at invoice time (paid_amount) reduces the payable
        # right away — omitting this credit leg is exactly what made this page show a
        # supplier as owed the full invoice total forever, even after it was paid in
        # full at receipt (see products/statements.py::build_supplier_statement, which
        # already accounts for this correctly and was used to catch this bug).
        inv_paid = inv.paid_amount or Decimal('0')
        if inv_paid > 0:
            entries.append({
                'date': inv.created_at, 'ref': f'فاتورة #{inv.id}',
                'desc': 'مدفوع مع الفاتورة',
                'debit': Decimal('0'), 'credit': inv_paid,
                'type': 'invoice_payment',
            })
    # Supplier returns credited to the account reduce what we owe (same as invoices —
    # this view previously omitted returns entirely).
    returns_qs = supplier.returns.filter(refund_method='debt')
    if date_from:
        returns_qs = returns_qs.filter(created_at__date__gte=date_from)
    if date_to:
        returns_qs = returns_qs.filter(created_at__date__lte=date_to)
    for ret in returns_qs:
        entries.append({
            'date': ret.created_at, 'ref': f'مرتجع #{ret.id}',
            'desc': 'مرتجع مشتريات (رصيد)',
            'debit': Decimal('0'), 'credit': ret.total_amount or Decimal('0'),
            'type': 'return',
        })
    for pay in payments_qs:
        entries.append({
            'date': pay.date, 'ref': f'سند #{pay.id}',
            'desc': pay.notes or 'سند دفع',
            'debit': Decimal('0'), 'credit': pay.amount,
            'type': 'payment',
        })
    # Sort but keep opening first
    other = [e for e in entries if e.get('type') != 'opening']
    other.sort(key=lambda x: x['date'])
    opening = [e for e in entries if e.get('type') == 'opening']
    entries = opening + other

    balance = Decimal('0')
    for e in entries:
        balance += e['debit'] - e['credit']
        e['balance'] = balance

    total_debit = sum(e['debit'] for e in entries)
    total_credit = sum(e['credit'] for e in entries)

    return render(request, 'products/supplier_account.html', {
        'supplier': supplier, 'entries': entries,
        'total_debit': total_debit, 'total_credit': total_credit,
        'closing_balance': balance,
        'date_from': date_from, 'date_to': date_to,
        'title': f'حساب المورد — {supplier.name}',
    })

@login_required
@require_permission('products', 'view')
def supplier_statement_pdf(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    invoices_qs = supplier.invoices.filter(status='CONFIRMED')
    payments_qs = supplier.payments.all()
    if date_from:
        invoices_qs = invoices_qs.filter(created_at__date__gte=date_from)
        payments_qs = payments_qs.filter(date__date__gte=date_from)
    if date_to:
        invoices_qs = invoices_qs.filter(created_at__date__lte=date_to)
        payments_qs = payments_qs.filter(date__date__lte=date_to)

    entries = []
    for inv in invoices_qs:
        entries.append({'date': inv.created_at, 'ref': f'فاتورة #{inv.id}',
            'desc': inv.notes or 'فاتورة مشتريات', 'debit': inv.net_amount, 'credit': Decimal('0')})
    for pay in payments_qs:
        entries.append({'date': pay.date, 'ref': f'سند #{pay.id}',
            'desc': pay.notes or 'سند دفع', 'debit': Decimal('0'), 'credit': pay.amount})
    entries.sort(key=lambda x: x['date'])

    balance = Decimal('0')
    for e in entries:
        balance += e['debit'] - e['credit']
        e['balance'] = balance

    sys_settings = SystemSetting.objects.first()
    from django.template.loader import render_to_string
    try:
        import weasyprint
        html = render_to_string('products/supplier_statement_pdf.html', {
            'supplier': supplier, 'entries': entries,
            'closing_balance': balance, 'sys_settings': sys_settings,
            'date_from': date_from, 'date_to': date_to,
            'print_date': timezone.now(),
        })
        pdf = weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="supplier_statement_{supplier.id}.pdf"'
        return response
    except Exception as e:
        # Fallback to HTML view if PDF generation fails (e.g. missing GTK)
        html = render_to_string('products/supplier_statement_pdf.html', {
            'supplier': supplier, 'entries': entries,
            'closing_balance': balance, 'sys_settings': sys_settings,
            'date_from': date_from, 'date_to': date_to,
            'print_date': timezone.now(),
            'is_print_preview': True,
        })
        return HttpResponse(html)

@login_required
@require_permission('products', 'view')
def bulk_edit_products(request):
    """Bulk edit selected products"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        field = data.get('field', '')
        value = data.get('value')
        allowed_fields = ['category_id', 'is_active', 'supplier_id', 'kind_id']
        if field not in allowed_fields:
            return JsonResponse({'success': False, 'error': 'حقل غير مسموح'})
        if field == 'is_active':
            value = value in (True, 'true', '1', 1)
        count = Product.objects.filter(id__in=ids).update(**{field: value})
        return JsonResponse({'success': True, 'message': f'تم تعديل {count} منتج'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


import urllib.request
import urllib.parse
import re

@login_required
@require_permission('products', 'view')
def fetch_product_image_api(request):
    """
    Auto-fetches an image thumbnail from Bing Images based on query.
    Used in Product Create/Edit forms to automatically suggest images.
    """
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'success': False, 'error': 'No query'})
    
    # Inject smart store-type context via the MarketProfile engine (single source of truth).
    try:
        from settings.models import SystemSetting
        from settings.market_profiles import get_market_profile
        s = SystemSetting.objects.first()
        prefix = get_market_profile(s.market_type if s else 'general').get('image_prefix', '')
    except Exception:
        prefix = ""

    if prefix and not query.startswith(prefix):
        # Only add prefix if it's not already there and if the query is simple
        if len(query.split()) <= 2:
            query = prefix + query
    
    # Switch to Bing native search
    try:
        import urllib.request, urllib.parse, re, json, base64
        # Clean query: Remove common non-essential words for better image matching
        clean_query = re.sub(r'\b(advance|new|latest|original)\b', '', query, flags=re.IGNORECASE).strip()
        
        # Add "product white background" to ensure clean, isolated product shots
        search_query = f"{clean_query} product isolated on white background"
        encoded_query = urllib.parse.quote(search_query)
        
        # Use Bing with filters for square photos and high-quality photography
        url = f"https://www.bing.com/images/search?q={encoded_query}&qft=+filterui:aspect-square+filterui:photo-photo+filterui:color2-color"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
        
        req = urllib.request.Request(url, headers=headers)
        html = urllib.request.urlopen(req, timeout=5).read().decode('utf-8', errors='ignore')
        
        # Bing embeds image info in m="..." fields
        matches = re.findall(r'm="([^"]+)"', html)
        images = []
        
        # Log the search attempt for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Image search for: {query} - Found {len(matches)} matches")

        for m in matches: # Check candidates
            try:
                # Clean and parse JSON
                item_json = m.replace('&quot;', '"')
                # Remove any potential trailing garbage before parsing
                item_json = item_json.split('"}')[0] + '"}' if '"}' in item_json else item_json
                
                item = json.loads(item_json)
                img_url = item.get('murl') or item.get('turl') # Prefer original URL for better quality if possible
                if not img_url: continue
                
                # Verify it's a real image URL
                if not img_url.startswith('http'): continue

                img_req = urllib.request.Request(img_url, headers=headers)
                img_data = urllib.request.urlopen(img_req, timeout=3).read()
                
                # Check if we got an actual image (at least 1KB)
                if len(img_data) < 1024: continue

                b64 = "data:image/jpeg;base64," + base64.b64encode(img_data).decode('utf-8')
                images.append(b64)
                if len(images) >= 5: break
            except Exception as e:
                logger.debug(f"Failed to process image match: {str(e)}")
                continue
        
        if images:
            return JsonResponse({'success': True, 'images': images})
            
        return JsonResponse({'success': False, 'error': 'لم يتم العثور على صور مناسبة. حاول تغيير اسم المنتج قليلاً.'})
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Image fetch failed: %s", str(e))
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@login_required
@require_POST
def api_purchase_invoice_confirm(request):
    try:
        data = json.loads(request.body)
        invoice_id = data.get('invoice_id')
        invoice = get_object_or_404(PurchaseInvoice, id=invoice_id)
        
        with db_transaction.atomic():
            from products.inventory_services import apply_purchase_invoice_stock
            applied, price_suggestions = apply_purchase_invoice_stock(invoice, user=request.user)
            if applied:
                # Deduct money and create payment transaction if paid_amount > 0 and account is set
                if invoice.paid_amount > 0 and invoice.account:
                    from financial.models import Transaction, DailyShift
                    shift = DailyShift.objects.filter(is_closed=False).last()
                    if not shift:
                        raise ValueError("يجب فتح وردية (شيفت) أولاً لتتمكن من تسجيل مبالغ مدفوعة")

                    if invoice.account.balance < invoice.paid_amount:
                        raise ValueError(
                            f"رصيد حساب {invoice.account.name} ({invoice.account.balance} ج.م) لا يكفي لسداد {invoice.paid_amount} ج.م"
                        )

                    Transaction.objects.create(
                        shift=shift,
                        account=invoice.account,
                        transaction_type='SUPPLIER_PAYMENT',
                        amount=invoice.paid_amount,
                        description=f"سداد (جزء من) فاتورة مشتريات #{invoice.id} للمورد: {invoice.supplier.name} (عند تأكيد المسودة)",
                        created_by=request.user
                    )
                return JsonResponse({
                    'status': 'success', 'message': 'تم تأكيد الفاتورة وتطبيق الكميات والمدفوعات بنجاح',
                    'price_suggestions': price_suggestions,
                })
            else:
                return JsonResponse({'status': 'error', 'message': 'المخزون مطبق بالفعل لهذه الفاتورة'})
    except ValueError as ve:
        return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
@login_required
@require_POST
def api_purchase_invoice_cancel(request):
    try:
        data = json.loads(request.body)
        invoice_id = data.get('invoice_id')
        invoice = get_object_or_404(PurchaseInvoice, id=invoice_id)
        
        with db_transaction.atomic():
            was_confirmed = invoice.status == 'CONFIRMED' and invoice.is_stock_applied
            
            from products.inventory_services import reverse_purchase_invoice_stock
            reversed_ok = reverse_purchase_invoice_stock(invoice, user=request.user)
            if reversed_ok:
                # If invoice had a paid amount, refund/reverse the financial transaction
                if was_confirmed and invoice.paid_amount > 0 and invoice.account:
                    from financial.models import Transaction, DailyShift
                    shift = DailyShift.objects.filter(is_closed=False).last()
                    if not shift:
                        raise ValueError("يجب فتح وردية (شيفت) أولاً لتتمكن من إلغاء الفاتورة واسترداد المبالغ")
                    
                    Transaction.objects.create(
                        shift=shift,
                        account=invoice.account,
                        transaction_type='INCOME',
                        amount=invoice.paid_amount,
                        description=f"استرداد مبلغ سداد فاتورة مشتريات ملغاة #{invoice.id} للمورد: {invoice.supplier.name}",
                        created_by=request.user
                    )
                return JsonResponse({'status': 'success', 'message': 'تم إلغاء الفاتورة وعكس حركة المخزون واسترداد المدفوعات بنجاح'})
            else:
                return JsonResponse({'status': 'error', 'message': 'لا يمكن إلغاء الفاتورة لأنها ملغاة بالفعل أو لم يتم تطبيق المخزون عليها'})
    except ValueError as ve:
        return JsonResponse({'status': 'error', 'message': str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
@login_required
@require_POST
def api_quick_create_product(request):
    try:
        data = json.loads(request.body)
        name = data.get('name')
        sku = data.get('sku')
        price_retail = data.get('price_retail')
        category_id = data.get('category_id')
        kind_id = data.get('kind_id')
        cost_price = data.get('cost_price') or '0.00'
        unit_measure = data.get('unit_measure') or 'PCS'

        # Market type fields
        scientific_name = data.get('scientific_name', '')
        packaging_type = data.get('packaging_type', '')
        strips_per_box = data.get('strips_per_box', 1)
        material = data.get('material', '')
        pattern = data.get('pattern', '')
        color = data.get('color', '')

        # Multi-Unit fields
        has_sub_unit = str(data.get('has_sub_unit', 'false')).lower() == 'true'
        sub_unit_id = data.get('sub_unit_id')
        sub_units_per_main_unit = data.get('sub_units_per_main_unit') or 1
        sub_unit_price = data.get('sub_unit_price') or 0.00

        if not name or not price_retail:
            return JsonResponse({'status': 'error', 'message': 'اسم المنتج والسعر القطاعي مطلوبان'}, status=400)

        with db_transaction.atomic():
            # Generate SKU if not provided
            if not sku:
                last_product = Product.objects.order_by('id').last()
                next_sku = "10001"
                if last_product and last_product.sku and last_product.sku.isdigit():
                    try:
                        next_sku = str(int(last_product.sku) + 1)
                    except ValueError:
                        pass
                sku = next_sku
                while Product.objects.filter(sku=sku).exists():
                    try:
                        sku = str(int(sku) + 1)
                    except ValueError:
                        break

            # Generate barcode from SKU
            digits = ''.join(filter(str.isdigit, sku)).zfill(12)[-12:]
            s = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits))
            barcode = digits + str((10 - (s % 10)) % 10)

            # Resolve category and kind
            category = None
            if category_id:
                category = get_object_or_404(Category, id=category_id)
            kind = None
            if kind_id:
                kind = get_object_or_404(Kind, id=kind_id)

            # Resolve sub_unit
            sub_unit_obj = None
            if sub_unit_id:
                sub_unit_obj = UnitOfMeasure.objects.filter(id=sub_unit_id).first()

            # Create product
            product = Product.objects.create(
                name=name,
                sku=sku,
                barcode=barcode,
                price_retail=Decimal(str(price_retail)),
                # Left at 0 (not sold at these tiers) — this quick-add modal only ever takes
                # a retail price; price_retail is the safe fallback at the point of sale
                # (see products.pricing.tier_price) if a wholesale-tier sale is attempted.
                cost_price=Decimal(str(cost_price)),
                unit_measure=unit_measure,
                category=category,
                kind=kind,
                is_active=True,
                scientific_name=scientific_name,
                packaging_type=packaging_type,
                strips_per_box=int(strips_per_box),
                material=material,
                pattern=pattern,
                color=color,
                has_sub_unit=has_sub_unit,
                sub_unit=sub_unit_obj,
                sub_units_per_main_unit=Decimal(str(sub_units_per_main_unit)),
                sub_unit_price=Decimal(str(sub_unit_price))
            )

            # If supplier is default, we can link it
            supplier_id = data.get('supplier_id')
            if supplier_id:
                supplier = Supplier.objects.filter(id=supplier_id).first()
                if supplier:
                    product.supplier = supplier
                    product.save(update_fields=['supplier'])

            return JsonResponse({
                'status': 'success',
                'product': {
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'cost_price': str(product.cost_price),
                    'unit': product.get_unit_measure_display()
                }
            })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@require_permission('products', 'view')
def low_stock_report(request):
    """Reorder / shortage report (Phase 5.7): products at or below their low-stock
    threshold, grouped by supplier, with a suggested order quantity (restock to ~2×
    the threshold) and estimated purchase cost."""
    import math
    from collections import OrderedDict

    only_out = request.GET.get('out') == '1'
    supplier_id = request.GET.get('supplier')

    # Kitchen-routed menu items are prepared on demand — never reorder stock for them.
    qs = (Product.objects.filter(is_active=True)
          .exclude(category__is_menu_category=True, is_raw_material=False)
          .select_related('supplier'))
    if only_out:
        qs = qs.filter(stock_quantity__lte=0)
    else:
        qs = qs.filter(stock_quantity__lte=F('low_stock_threshold'), low_stock_threshold__gt=0)
    if supplier_id:
        qs = qs.filter(supplier_id=supplier_id)
    qs = qs.order_by('supplier__name', 'name')

    groups = OrderedDict()
    grand_cost = Decimal('0.00')
    total_items = 0
    for p in qs:
        threshold = p.low_stock_threshold or Decimal('0')
        current = p.stock_quantity or Decimal('0')
        target = threshold * 2 if threshold > 0 else Decimal('1')
        suggested = target - current
        if suggested < 1:
            suggested = (threshold - current) if (threshold - current) > 0 else Decimal('1')
        suggested = Decimal(str(int(math.ceil(float(suggested)))))
        line_cost = suggested * (p.cost_price or Decimal('0'))
        grand_cost += line_cost
        total_items += 1

        key = p.supplier_id or 0
        if key not in groups:
            groups[key] = {'supplier': p.supplier, 'rows': [], 'subtotal': Decimal('0.00')}
        groups[key]['rows'].append({
            'product': p, 'current': current, 'threshold': threshold,
            'suggested': suggested, 'line_cost': line_cost,
        })
        groups[key]['subtotal'] += line_cost

    from .models import Supplier
    return render(request, 'products/low_stock_report.html', {
        'title': 'تقرير النواقص وإعادة الطلب',
        'groups': list(groups.values()),
        'grand_cost': grand_cost,
        'total_items': total_items,
        'only_out': only_out,
        'suppliers': Supplier.objects.all().order_by('name'),
        'supplier_id': supplier_id or '',
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_permission('products', 'view')
def item_movement_card(request, pk):
    """Item movement card / حركة الصنف (Phase 5.2): chronological stock ledger for a
    product with signed quantities and a running balance."""
    from datetime import datetime
    from .models import StockTransaction, Warehouse

    product = get_object_or_404(Product, pk=pk)

    wh_id = request.GET.get('warehouse')
    wh_id = int(wh_id) if (wh_id and wh_id.isdigit()) else None

    def _parse(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None
    date_from = _parse(request.GET.get('date_from'))
    date_to = _parse(request.GET.get('date_to'))

    INC = {'IN', 'RET_IN', 'RET', 'MFG_OUT'}
    DEC = {'OUT', 'RET_OUT', 'MFG'}

    def signed(t):
        q = t.quantity or Decimal('0')
        tt = t.transaction_type
        if tt in INC:
            return q
        if tt in DEC:
            return -q
        if tt == 'TRN':
            if wh_id is None:
                return Decimal('0')  # internal transfer nets out across all warehouses
            if t.warehouse_id == wh_id:
                return -q
            if t.destination_warehouse_id == wh_id:
                return q
            return Decimal('0')
        return q  # ADJ — quantity stored signed

    base = StockTransaction.objects.filter(product=product)
    if wh_id:
        base = base.filter(Q(warehouse_id=wh_id) | Q(destination_warehouse_id=wh_id))

    # Opening balance = signed sum of everything before the start date.
    opening = Decimal('0.00')
    if date_from:
        for t in base.filter(created_at__date__lt=date_from):
            opening += signed(t)

    in_range = base.order_by('created_at', 'id')
    if date_from:
        in_range = in_range.filter(created_at__date__gte=date_from)
    if date_to:
        in_range = in_range.filter(created_at__date__lte=date_to)

    rows = []
    running = opening
    total_in = Decimal('0.00')
    total_out = Decimal('0.00')
    for t in in_range.select_related('warehouse', 'destination_warehouse', 'created_by'):
        s = signed(t)
        running += s
        if s > 0:
            total_in += s
        else:
            total_out += -s
        rows.append({'t': t, 'signed': s, 'balance': running, 'is_in': s >= 0})

    return render(request, 'products/item_movement_card.html', {
        'title': f'حركة الصنف: {product.name}',
        'product': product,
        'rows': rows,
        'opening': opening,
        'closing': running,
        'total_in': total_in,
        'total_out': total_out,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'wh_id': wh_id or '',
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_granular_action('inventory', 'valuation', 'products', 'view')
def stock_valuation(request):
    """Inventory valuation (Phase 5.11): current stock x batch cost, by category."""
    from collections import OrderedDict
    from .models import StockBatch
    from django.db.models import Sum, F, DecimalField, ExpressionWrapper

    cat_id = request.GET.get('category')
    batches = StockBatch.objects.filter(current_quantity__gt=0).select_related('product', 'product__category')
    if cat_id and cat_id.isdigit():
        batches = batches.filter(product__category_id=int(cat_id))

    groups = OrderedDict()
    grand_qty = Decimal('0'); grand_val = Decimal('0.00')
    per_product = {}
    for b in batches:
        p = b.product
        if not p:
            continue
        val = (b.current_quantity or Decimal('0')) * (b.purchase_price or Decimal('0'))
        rec = per_product.setdefault(p.id, {'product': p, 'qty': Decimal('0'), 'value': Decimal('0.00')})
        rec['qty'] += b.current_quantity or Decimal('0')
        rec['value'] += val

    for rec in per_product.values():
        cat = rec['product'].category.name if rec['product'].category else 'بدون قسم'
        g = groups.setdefault(cat, {'name': cat, 'rows': [], 'qty': Decimal('0'), 'value': Decimal('0.00')})
        g['rows'].append(rec)
        g['qty'] += rec['qty']; g['value'] += rec['value']
        grand_qty += rec['qty']; grand_val += rec['value']

    from .models import Category
    return render(request, 'products/stock_valuation.html', {
        'title': 'تقييم المخزون', 'groups': list(groups.values()),
        'grand_qty': grand_qty, 'grand_val': grand_val,
        'categories': Category.objects.all(), 'cat_id': cat_id or '',
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_granular_action('inventory', 'expiry', 'products', 'view')
def expiry_report(request):
    """Near-expiry & expired stock (Phase 5.6)."""
    from datetime import timedelta
    from django.utils import timezone
    from .models import StockBatch

    today = timezone.localdate()
    qs = (StockBatch.objects.filter(current_quantity__gt=0, expiry_date__isnull=False)
          .select_related('product', 'warehouse').order_by('expiry_date'))

    expired, d30, d60, d90 = [], [], [], []
    for b in qs:
        days = (b.expiry_date - today).days
        rec = {'b': b, 'days': days}
        if days < 0:
            expired.append(rec)
        elif days <= 30:
            d30.append(rec)
        elif days <= 60:
            d60.append(rec)
        elif days <= 90:
            d90.append(rec)
    return render(request, 'products/expiry_report.html', {
        'title': 'تقرير الصلاحية', 'today': today,
        'expired': expired, 'd30': d30, 'd60': d60, 'd90': d90,
        'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_granular_action('inventory', 'ap_aging', 'products', 'view')
def ap_aging_report(request):
    """Supplier payables aging (Phase 7.4)."""
    from datetime import datetime
    from .aging import ap_aging
    as_of = None
    raw = request.GET.get('as_of')
    if raw:
        try:
            as_of = datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            as_of = None
    rows, totals = ap_aging(as_of)
    if request.GET.get('export') == 'csv':
        from financial.reports import csv_response
        header = ['المورد', 'حالي', '31-60', '61-90', '90+', 'الإجمالي']
        data = [[r['supplier'].name, r['current'], r['d30'], r['d60'], r['d90'], r['total']] for r in rows]
        return csv_response('ap_aging', header, data)
    return render(request, 'products/ap_aging.html', {
        'title': 'أعمار ديون الموردين', 'rows': rows, 'totals': totals,
        'as_of': raw or '', 'print_mode': request.GET.get('print') == '1',
    })


@login_required
@require_granular_action('inventory', 'audit', 'products', 'view')
def stocktake_list(request):
    from .models import StockCount
    counts = StockCount.objects.select_related('warehouse', 'created_by').all()
    from .models import Warehouse
    return render(request, 'products/stocktake_list.html', {
        'title': 'الجرد', 'counts': counts,
        'warehouses': Warehouse.objects.filter(is_active=True),
    })


@login_required
@require_permission('products', 'edit')
@require_POST
def stocktake_create(request):
    """Open a stocktake session — snapshots current system quantities for the warehouse."""
    from .models import StockCount, StockCountItem, Warehouse, WarehouseStock
    wh = get_object_or_404(Warehouse, id=request.POST.get('warehouse_id'))
    count = StockCount.objects.create(warehouse=wh, created_by=request.user,
                                      note=request.POST.get('note', ''))
    # Menu-category items are prepared on demand and never carry a real stock count —
    # including them would pollute every variance list with phantom "shortages". Raw
    # materials are always counted though, even if their category fell back to the
    # generic "بدون قسم" (is_menu_category=True).
    rows = (WarehouseStock.objects.filter(warehouse=wh)
            .exclude(product__category__is_menu_category=True, product__is_raw_material=False)
            .select_related('product'))
    StockCountItem.objects.bulk_create([
        StockCountItem(count=count, product=ws.product, system_qty=ws.quantity or 0)
        for ws in rows if ws.product
    ])
    messages.success(request, f"تم فتح جلسة جرد #{count.id} للمخزن {wh.name}.")
    return redirect('stocktake_detail', pk=count.id)


@login_required
@require_permission('products', 'view')
def stocktake_detail(request, pk):
    from .models import StockCount
    count = get_object_or_404(StockCount.objects.select_related('warehouse'), pk=pk)
    if request.method == 'POST' and not count.is_applied:
        from accounts.permissions import has_permission
        if not has_permission(request.user, 'products', 'edit'):
            raise PermissionDenied()
        for item in count.items.all():
            raw = request.POST.get(f'counted_{item.id}', '').strip()
            if raw != '':
                try:
                    item.counted_qty = Decimal(raw)
                    item.save(update_fields=['counted_qty'])
                except (InvalidOperation, TypeError):
                    pass
        messages.success(request, "تم حفظ الجرد. راجع الفروقات ثم اعتمد الجرد.")
        return redirect('stocktake_detail', pk=count.id)

    items = list(count.items.select_related('product').order_by('product__name'))
    variance_items = [i for i in items if i.variance is not None and i.variance != 0]
    return render(request, 'products/stocktake_detail.html', {
        'title': f'جرد #{count.id}', 'count': count, 'items': items,
        'variance_count': len(variance_items),
        'counted_count': len([i for i in items if i.counted_qty is not None]),
    })


@login_required
@require_permission('products', 'edit')
@require_POST
def stocktake_apply(request, pk):
    """Apply the count — adjusts stock to the counted quantities via the inventory service."""
    from .models import StockCount
    from .inventory_services import adjust_to
    from django.utils import timezone
    count = get_object_or_404(StockCount, pk=pk)
    if count.is_applied:
        messages.error(request, "هذا الجرد معتمد بالفعل.")
        return redirect('stocktake_detail', pk=pk)

    applied = 0
    with db_transaction.atomic():
        for item in count.items.select_related('product'):
            if item.counted_qty is None:
                continue
            delta = adjust_to(item.product, count.warehouse, item.counted_qty,
                              user=request.user, reference=f'COUNT-{count.id}',
                              note=f'تسوية جرد #{count.id}')
            if delta != 0:
                applied += 1
        count.status = StockCount.STATUS_APPLIED
        count.applied_by = request.user
        count.applied_at = timezone.now()
        count.save(update_fields=['status', 'applied_by', 'applied_at'])
    messages.success(request, f"تم اعتماد الجرد وتسوية {applied} صنف.")
    return redirect('stocktake_detail', pk=pk)


@login_required
@require_granular_action('inventory', 'price_comparison', 'products', 'view')
def supplier_price_comparison(request):
    """Compare suppliers' last purchase prices per product (Phase 7.5). Read-only."""
    from .models import SupplierProduct
    from collections import OrderedDict

    q = request.GET.get('q', '').strip()
    only_multi = request.GET.get('multi') == '1'

    links = (SupplierProduct.objects.filter(last_purchase_price__gt=0)
             .select_related('product', 'supplier').order_by('product__name', 'last_purchase_price'))
    if q:
        links = links.filter(Q(product__name__icontains=q) | Q(product__sku__icontains=q))

    groups = OrderedDict()
    for link in links:
        if not link.product:
            continue
        g = groups.setdefault(link.product_id, {'product': link.product, 'offers': []})
        g['offers'].append(link)

    rows = []
    for g in groups.values():
        offers = g['offers']
        if only_multi and len(offers) < 2:
            continue
        cheapest = min(o.last_purchase_price for o in offers)
        dearest = max(o.last_purchase_price for o in offers)
        for o in offers:
            o.is_cheapest = (o.last_purchase_price == cheapest)
        rows.append({'product': g['product'], 'offers': offers,
                     'cheapest': cheapest, 'spread': dearest - cheapest, 'count': len(offers)})

    return render(request, 'products/supplier_price_comparison.html', {
        'title': 'مقارنة أسعار الموردين', 'rows': rows,
        'q': q, 'only_multi': only_multi,
    })


@login_required
@require_permission('products', 'edit')
@require_POST
def price_break_add(request, pk):
    """Add a quantity-break price to a product (Phase 6.6)."""
    from .models import ProductPriceBreak
    product = get_object_or_404(Product, pk=pk)
    try:
        min_qty = Decimal(str(request.POST.get('min_quantity', '0')))
        unit_price = Decimal(str(request.POST.get('unit_price', '0')))
    except (InvalidOperation, TypeError):
        messages.error(request, "قيم غير صحيحة.")
        return redirect('product_detail', pk=pk)
    ctype = request.POST.get('customer_type', '')
    if ctype not in ('', 'retail', 'semi_wholesale', 'wholesale'):
        ctype = ''
    if min_qty <= 0 or unit_price <= 0:
        messages.error(request, "الكمية والسعر يجب أن يكونا أكبر من صفر.")
        return redirect('product_detail', pk=pk)
    ProductPriceBreak.objects.create(product=product, min_quantity=min_qty,
                                     unit_price=unit_price, customer_type=ctype)
    messages.success(request, "تمت إضافة شريحة السعر.")
    return redirect('product_detail', pk=pk)


@login_required
@require_permission('products', 'edit')
@require_POST
def price_break_delete(request, pk):
    from .models import ProductPriceBreak
    br = get_object_or_404(ProductPriceBreak, pk=pk)
    product_id = br.product_id
    br.delete()
    messages.success(request, "تم حذف شريحة السعر.")
    return redirect('product_detail', pk=product_id)


@login_required
@require_permission('products', 'view')
def inventory_insights_view(request):
    """Demand & dead-stock insights (Phase 10.5)."""
    from .insights import inventory_insights
    try:
        window = int(request.GET.get('window', '90'))
    except (ValueError, TypeError):
        window = 90
    window = max(7, min(window, 365))
    data = inventory_insights(window_days=window)
    return render(request, 'products/inventory_insights.html', {
        'title': 'تحليل المخزون والطلب', 'd': data, 'window': window,
        'print_mode': request.GET.get('print') == '1',
    })


# ── Fashion variants (size × color) ───────────────────────────────────────────
@login_required
@require_permission('products', 'edit')
def product_variants(request, pk):
    """Manage a product's variants (size × color) — list, add one, bulk-generate, delete."""
    from .models import Product, ProductVariant, Size
    from decimal import Decimal
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            size_id = request.POST.get('size') or None
            color = (request.POST.get('color') or '').strip()
            try:
                ProductVariant.objects.get_or_create(
                    product=product,
                    size_id=size_id or None,
                    color=color,
                    defaults={
                        'barcode': (request.POST.get('barcode') or '').strip(),
                        'price_override': (Decimal(request.POST['price_override'])
                                           if request.POST.get('price_override') else None),
                        'stock_quantity': Decimal(request.POST.get('stock_quantity') or '0'),
                    })
                messages.success(request, 'تمت إضافة الخيار.')
            except Exception as e:
                messages.error(request, f'تعذّر الحفظ: {e}')
        elif action == 'bulk':
            # sizes = multi-select ids; colors = comma/newline separated names → matrix
            size_ids = request.POST.getlist('sizes') or [None]
            colors = [c.strip() for c in (request.POST.get('colors') or '').replace('\n', ',').split(',') if c.strip()] or ['']
            created = 0
            for sid in size_ids:
                for col in colors:
                    _, was_new = ProductVariant.objects.get_or_create(
                        product=product, size_id=(sid or None), color=col,
                        defaults={'stock_quantity': Decimal('0')})
                    created += 1 if was_new else 0
            if not product.has_variants:
                product.has_variants = True
                product.save(update_fields=['has_variants'])
            messages.success(request, f'تم إنشاء {created} خيار جديد.')
        elif action == 'set_stock':
            v = ProductVariant.objects.filter(pk=request.POST.get('variant_id'), product=product).first()
            if v:
                try:
                    v.stock_quantity = Decimal(request.POST.get('stock_quantity') or '0')
                    v.price_override = (Decimal(request.POST['price_override'])
                                        if request.POST.get('price_override') else None)
                    v.save(update_fields=['stock_quantity', 'price_override'])
                    messages.success(request, 'تم التحديث.')
                except Exception as e:
                    messages.error(request, f'قيمة غير صحيحة: {e}')
        elif action == 'delete':
            ProductVariant.objects.filter(pk=request.POST.get('variant_id'), product=product).delete()
            messages.success(request, 'تم حذف الخيار.')
        return redirect('product_variants', pk=pk)

    variants = product.variants.select_related('size').all()
    sizes = Size.objects.filter(is_active=True).order_by('sort_order', 'name')
    return render(request, 'products/product_variants.html', {
        'product': product, 'variants': variants, 'sizes': sizes,
        'title': f'خيارات: {product.name}',
    })


@login_required
def product_variants_api(request, pk):
    """JSON list of a product's sellable variants — used by the POS variant picker."""
    from .models import Product, ProductVariant
    product = get_object_or_404(Product, pk=pk)
    rows = []
    for v in product.variants.select_related('size').filter(is_active=True):
        rows.append({
            'id': v.id, 'label': v.label,
            'size': v.size.name if v.size else '', 'color': v.color,
            'price': float(v.price), 'stock': float(v.stock_quantity),
            'barcode': v.barcode,
        })
    return JsonResponse({'product_id': product.id, 'name': product.name, 'variants': rows})
